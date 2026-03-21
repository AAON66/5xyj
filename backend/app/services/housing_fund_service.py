from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from backend.app.services.header_normalizer import HeaderMappingDecision, HeaderNormalizationResult
from backend.app.services.normalization_service import NormalizedPreviewRecord, StandardizationResult

HEADER_PATTERNS: dict[str, tuple[str, ...]] = {
    "person_name": ("\u59d3\u540d", "\u804c\u5de5\u59d3\u540d"),
    "id_number": ("\u8bc1\u4ef6\u53f7\u7801",),
    "id_type": ("\u8bc1\u4ef6\u7c7b\u578b",),
    "company_name": ("\u5355\u4f4d\u540d\u79f0", "\u7528\u5de5\u5355\u4f4d\u540d\u79f0"),
    "billing_period": ("\u7f34\u5b58\u65f6\u6bb5", "\u4e1a\u52a1\u5e74\u6708", "\u6c47\u8865\u7f34\u5e74\u6708", "\u6c47\u7f34\u5e74\u6708"),
    "housing_fund_account": ("\u516c\u79ef\u91d1\u8d26\u53f7", "\u4e2a\u4eba\u8d26\u53f7", "\u4e2a\u4eba\u5ba2\u6237\u53f7", "\u804c\u5de5\u8d26\u53f7"),
    "housing_fund_base": ("\u7f34\u5b58\u57fa\u6570", "\u7f34\u5b58\u57fa\u6570\uff08\u5143\uff09"),
    "housing_fund_personal": ("\u4e2a\u4eba\u7f34\u5b58\u989d", "\u4e2a\u4eba\u6708\u7f34\u5b58\u989d(\u5143)", "\u804c\u5de5\u6708\u7f34\u989d"),
    "housing_fund_company": ("\u5355\u4f4d\u7f34\u5b58\u989d", "\u5355\u4f4d\u6708\u7f34\u5b58\u989d(\u5143)", "\u5355\u4f4d\u6708\u7f34\u989d"),
    "housing_fund_total": ("\u91d1\u989d\u5408\u8ba1\uff08\u5143\uff09", "\u5408\u8ba1\u6708\u7f34\u5b58\u989d(\u5143)", "\u603b\u91d1\u989d", "\u53d1\u751f\u989d"),
}
RATE_PATTERNS: dict[str, tuple[str, ...]] = {
    "company_rate": ("\u5355\u4f4d\u7f34\u5b58\u6bd4\u4f8b",),
    "personal_rate": ("\u4e2a\u4eba\u7f34\u5b58\u6bd4\u4f8b",),
}
PLACEHOLDER_STRINGS = {"", "-", "--", "\u2014\u2014", "none", "null", "(\u7a7a\u767d)"}
NON_DETAIL_NAME_PATTERNS = ("经办网点", "打印日期", "管理中心", "制表人", "说明", "备注", "汇总", "合计")


@dataclass(slots=True)
class HousingFundWorkbookAnalysis:
    normalization: HeaderNormalizationResult
    standardized: StandardizationResult


@dataclass(slots=True)
class _WorkbookCandidate:
    sheet_name: str
    header_row: int
    headers: list[str]
    raw_header_signature: str


def analyze_housing_fund_workbook(
    path: str | Path,
    *,
    region: str | None = None,
    company_name: str | None = None,
    source_file_name: str | None = None,
) -> HousingFundWorkbookAnalysis:
    workbook_path = Path(path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        candidate = _detect_workbook_candidate(workbook, workbook_path.name)
        sheet = workbook[candidate.sheet_name]
        workbook_company_name = company_name or _extract_company_name(sheet, candidate.header_row)
        decisions = [_build_decision(header) for header in candidate.headers]
        normalization = HeaderNormalizationResult(
            source_file=source_file_name or workbook_path.name,
            sheet_name=candidate.sheet_name,
            raw_header_signature=candidate.raw_header_signature,
            decisions=decisions,
            unmapped_headers=[decision.raw_header_signature for decision in decisions if decision.canonical_field is None],
        )

        records: list[NormalizedPreviewRecord] = []
        row_number = candidate.header_row + 1
        while row_number <= sheet.max_row:
            row_values = [sheet.cell(row_number, index + 1).value for index in range(len(candidate.headers))]
            preview = _build_preview_record(
                headers=candidate.headers,
                values=row_values,
                source_file_name=source_file_name or workbook_path.name,
                sheet_name=candidate.sheet_name,
                raw_header_signature=candidate.raw_header_signature,
                row_number=row_number,
                region=region,
                company_name=workbook_company_name,
            )
            if preview is not None:
                records.append(preview)
            row_number += 1

        standardized = StandardizationResult(
            source_file=source_file_name or workbook_path.name,
            sheet_name=candidate.sheet_name,
            raw_header_signature=candidate.raw_header_signature,
            records=records,
            filtered_rows=[],
            unmapped_headers=normalization.unmapped_headers,
        )
        return HousingFundWorkbookAnalysis(normalization=normalization, standardized=standardized)
    finally:
        workbook.close()


def standardize_housing_fund_workbook(
    path: str | Path,
    *,
    region: str | None = None,
    company_name: str | None = None,
    source_file_name: str | None = None,
) -> StandardizationResult:
    return analyze_housing_fund_workbook(
        path,
        region=region,
        company_name=company_name,
        source_file_name=source_file_name,
    ).standardized


def _detect_workbook_candidate(workbook, source_file_name: str) -> _WorkbookCandidate:
    best_candidate: _WorkbookCandidate | None = None
    best_score = -1
    for sheet in workbook.worksheets:
        for row_number in range(1, min(sheet.max_row, 12) + 1):
            headers = [_clean_text(cell.value) for cell in sheet[row_number]]
            score = _score_headers(headers)
            if score <= best_score:
                continue
            best_score = score
            filtered_headers = [header or f"column_{index + 1}" for index, header in enumerate(headers)]
            best_candidate = _WorkbookCandidate(
                sheet_name=sheet.title,
                header_row=row_number,
                headers=filtered_headers,
                raw_header_signature=" | ".join(filtered_headers),
            )

    if best_candidate is None or best_score < 8:
        raise ValueError(f"Could not detect a valid housing fund header row in {source_file_name}.")
    return best_candidate


def _score_headers(headers: list[str | None]) -> int:
    score = 0
    normalized = [item or "" for item in headers]
    if any(_header_matches(header, HEADER_PATTERNS["person_name"]) for header in normalized):
        score += 4
    if any(_header_matches(header, HEADER_PATTERNS["id_number"]) for header in normalized):
        score += 4
    if any(_header_matches(header, HEADER_PATTERNS["housing_fund_total"]) for header in normalized):
        score += 3
    if any(_header_matches(header, HEADER_PATTERNS["housing_fund_personal"]) for header in normalized):
        score += 2
    if any(_header_matches(header, HEADER_PATTERNS["housing_fund_company"]) for header in normalized):
        score += 2
    if any(_header_matches(header, HEADER_PATTERNS["housing_fund_base"]) for header in normalized):
        score += 1
    return score


def _build_decision(header: str) -> HeaderMappingDecision:
    canonical_field = None
    for field_name, patterns in HEADER_PATTERNS.items():
        if _header_matches(header, patterns):
            canonical_field = field_name
            break
    return HeaderMappingDecision(
        raw_header=header,
        raw_header_signature=header,
        canonical_field=canonical_field,
        mapping_source="rule" if canonical_field else "unmapped",
        confidence=0.98 if canonical_field else None,
        candidate_fields=[canonical_field] if canonical_field else [],
        matched_rules=[header] if canonical_field else [],
    )


def _extract_company_name(sheet, header_row: int) -> str | None:
    for row_number in range(1, min(header_row, 6) + 1):
        values = [_clean_text(cell.value) for cell in sheet[row_number]]
        for index, value in enumerate(values):
            if not value:
                continue
            if "\u5355\u4f4d\u540d\u79f0" in value and "\uff1a" in value:
                return value.split("\uff1a", 1)[-1].strip() or None
            if value == "\u5355\u4f4d\u540d\u79f0\uff1a" and index + 1 < len(values):
                neighbor = values[index + 1]
                if neighbor:
                    return neighbor
    return None


def _build_preview_record(
    *,
    headers: list[str],
    values: list[object],
    source_file_name: str,
    sheet_name: str,
    raw_header_signature: str,
    row_number: int,
    region: str | None,
    company_name: str | None,
) -> NormalizedPreviewRecord | None:
    raw_values = {header: _clean_cell_value(value) for header, value in zip(headers, values, strict=True)}
    person_name = _find_first_value(raw_values, HEADER_PATTERNS["person_name"])
    id_number = _find_first_value(raw_values, HEADER_PATTERNS["id_number"])
    if not person_name and not id_number:
        return None
    if _looks_like_non_detail_record(person_name, id_number):
        return None

    billing_period = _normalize_period(_find_first_value(raw_values, HEADER_PATTERNS["billing_period"]))
    period_start, period_end = _derive_period_bounds(_find_first_value(raw_values, HEADER_PATTERNS["billing_period"]))
    housing_account = _find_first_value(raw_values, HEADER_PATTERNS["housing_fund_account"])
    housing_base = _to_decimal(_find_first_value(raw_values, HEADER_PATTERNS["housing_fund_base"]))
    personal_amount = _to_decimal(_find_first_value(raw_values, HEADER_PATTERNS["housing_fund_personal"]))
    company_amount = _to_decimal(_find_first_value(raw_values, HEADER_PATTERNS["housing_fund_company"]))
    total_amount = _to_decimal(_find_first_value(raw_values, HEADER_PATTERNS["housing_fund_total"]))
    personal_rate = _to_decimal(_find_first_value(raw_values, RATE_PATTERNS["personal_rate"]))
    company_rate = _to_decimal(_find_first_value(raw_values, RATE_PATTERNS["company_rate"]))

    inference_notes: list[str] = []
    personal_amount, company_amount, total_amount = _resolve_housing_amounts(
        personal_amount=personal_amount,
        company_amount=company_amount,
        total_amount=total_amount,
        base_amount=housing_base,
        personal_rate=personal_rate,
        company_rate=company_rate,
        inference_notes=inference_notes,
    )

    record_company_name = _find_first_value(raw_values, HEADER_PATTERNS["company_name"]) or company_name
    raw_payload = {
        "raw_values": {key: _json_safe(value) for key, value in raw_values.items()},
        "unmapped_values": {},
    }
    if inference_notes:
        raw_payload["housing_fund_inference_notes"] = inference_notes

    values_payload: dict[str, Any] = {
        "person_name": person_name,
        "id_number": id_number,
        "id_type": _find_first_value(raw_values, HEADER_PATTERNS["id_type"]),
        "company_name": record_company_name,
        "region": region,
        "billing_period": billing_period,
        "period_start": period_start,
        "period_end": period_end,
        "housing_fund_account": housing_account,
        "housing_fund_base": housing_base,
        "housing_fund_personal": personal_amount,
        "housing_fund_company": company_amount,
        "housing_fund_total": total_amount,
        "raw_sheet_name": sheet_name,
        "raw_header_signature": raw_header_signature,
        "source_file_name": source_file_name,
    }
    return NormalizedPreviewRecord(
        source_row_number=row_number,
        values={key: value for key, value in values_payload.items() if value is not None},
        unmapped_values={},
        raw_values=raw_values,
        raw_payload=raw_payload,
    )


def _looks_like_non_detail_record(person_name: str | None, id_number: str | None) -> bool:
    normalized_name = (person_name or '').replace(' ', '')
    normalized_id = (id_number or '').strip()
    if any(pattern in normalized_name for pattern in NON_DETAIL_NAME_PATTERNS):
        return True
    if normalized_name and len(normalized_name) > 20 and not normalized_id:
        return True
    return False


def _resolve_housing_amounts(
    *,
    personal_amount: Decimal | None,
    company_amount: Decimal | None,
    total_amount: Decimal | None,
    base_amount: Decimal | None,
    personal_rate: Decimal | None,
    company_rate: Decimal | None,
    inference_notes: list[str],
) -> tuple[Decimal | None, Decimal | None, Decimal | None]:
    quant = Decimal("0.01")

    if total_amount is None and personal_amount is not None and company_amount is not None:
        total_amount = (personal_amount + company_amount).quantize(quant)

    if personal_amount is None and company_amount is None and total_amount is not None:
        if base_amount is not None and personal_rate is not None and company_rate is not None and (personal_rate + company_rate) != 0:
            total_rate = personal_rate + company_rate
            personal_amount = (total_amount * personal_rate / total_rate).quantize(quant, rounding=ROUND_HALF_UP)
            company_amount = (total_amount - personal_amount).quantize(quant, rounding=ROUND_HALF_UP)
            inference_notes.append("split_from_total_and_rates")
        else:
            personal_amount = (total_amount / Decimal("2")).quantize(quant, rounding=ROUND_HALF_UP)
            company_amount = (total_amount - personal_amount).quantize(quant, rounding=ROUND_HALF_UP)
            inference_notes.append("split_equally_from_total")

    if personal_amount is None and total_amount is not None and company_amount is not None:
        personal_amount = (total_amount - company_amount).quantize(quant, rounding=ROUND_HALF_UP)
        inference_notes.append("derived_personal_from_total")
    if company_amount is None and total_amount is not None and personal_amount is not None:
        company_amount = (total_amount - personal_amount).quantize(quant, rounding=ROUND_HALF_UP)
        inference_notes.append("derived_company_from_total")

    if personal_amount is None and base_amount is not None and personal_rate is not None:
        personal_amount = (base_amount * personal_rate).quantize(quant, rounding=ROUND_HALF_UP)
        inference_notes.append("derived_personal_from_base_rate")
    if company_amount is None and base_amount is not None and company_rate is not None:
        company_amount = (base_amount * company_rate).quantize(quant, rounding=ROUND_HALF_UP)
        inference_notes.append("derived_company_from_base_rate")

    if total_amount is None and personal_amount is not None and company_amount is not None:
        total_amount = (personal_amount + company_amount).quantize(quant, rounding=ROUND_HALF_UP)
    return personal_amount, company_amount, total_amount


def _find_first_value(raw_values: dict[str, Any], patterns: tuple[str, ...]) -> Any:
    for header, value in raw_values.items():
        if _header_matches(header, patterns):
            return value
    return None


def _header_matches(header: str | None, patterns: tuple[str, ...]) -> bool:
    normalized = (header or "").replace(" ", "")
    return any(pattern.replace(" ", "") in normalized for pattern in patterns)


def _normalize_period(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if "-" in text and len(text.replace("-", "")) == 12:
        start = text.split("-", 1)[0]
        return f"{start[:4]}-{start[4:6]}"
    if len(text) == 6 and text.isdigit():
        return f"{text[:4]}-{text[4:6]}"
    if len(text) == 7 and text[4] == "-":
        return text
    return text


def _derive_period_bounds(value: Any) -> tuple[str | None, str | None]:
    if value is None:
        return None, None
    text = str(value).strip()
    if not text:
        return None, None
    if "-" in text and len(text.replace("-", "")) == 12:
        start_raw, end_raw = text.split("-", 1)
        start = _normalize_period(start_raw)
        end = _normalize_period(end_raw)
        return start, end
    normalized = _normalize_period(text)
    return normalized, normalized


def _clean_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _clean_cell_value(value: object) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.lower() in PLACEHOLDER_STRINGS:
            return None
        return stripped
    return value


def _to_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        return Decimal(str(value))
    candidate = str(value).strip().replace(",", "")
    if not candidate or candidate.lower() in PLACEHOLDER_STRINGS:
        return None
    try:
        return Decimal(candidate)
    except InvalidOperation:
        return None


def _json_safe(value: Any) -> Any:
    if isinstance(value, Decimal):
        return format(value, "f")
    return value
