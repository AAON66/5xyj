from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal
from io import BytesIO
from typing import TYPE_CHECKING, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, selectinload

from backend.app.models import ImportBatch, NormalizedRecord, SourceFile
from backend.app.schemas.compare import (
    BatchCompareRead,
    CompareBatchMetaRead,
    CompareExportRequest,
    CompareRecordSideRead,
    CompareRowRead,
)
from backend.app.services.import_service import BatchNotFoundError

if TYPE_CHECKING:
    from backend.app.schemas.compare import CompareRowInput


PREFERRED_COMPARE_FIELDS = [
    'person_name',
    'employee_id',
    'id_number',
    'social_security_number',
    'housing_fund_account',
    'company_name',
    'region',
    'billing_period',
    'period_start',
    'period_end',
    'payment_base',
    'payment_salary',
    'housing_fund_base',
    'housing_fund_personal',
    'housing_fund_company',
    'housing_fund_total',
    'total_amount',
    'company_total_amount',
    'personal_total_amount',
    'pension_company',
    'pension_personal',
    'medical_company',
    'medical_personal',
    'medical_maternity_company',
    'maternity_amount',
    'unemployment_company',
    'unemployment_personal',
    'injury_company',
    'supplementary_medical_company',
    'supplementary_pension_company',
    'large_medical_personal',
    'late_fee',
    'interest',
    'raw_sheet_name',
    'raw_header_signature',
    'source_file_name',
]

LEFT_ONLY_STATUS = 'left_only'
RIGHT_ONLY_STATUS = 'right_only'
CHANGED_STATUS = 'changed'
SAME_STATUS = 'same'

DIFF_ROW_FILL = PatternFill(fill_type='solid', fgColor='FDE7E9')
HEADER_FILL = PatternFill(fill_type='solid', fgColor='EAF0F9')
HEADER_FONT = Font(bold=True)


@dataclass(frozen=True, slots=True)
class CompareIdentity:
    basis: str
    value: str


def compare_batches(db: Session, left_batch_id: str, right_batch_id: str) -> BatchCompareRead:
    left_batch = _get_batch_with_records(db, left_batch_id)
    right_batch = _get_batch_with_records(db, right_batch_id)

    left_groups = _group_records_by_identity(left_batch.normalized_records)
    right_groups = _group_records_by_identity(right_batch.normalized_records)
    all_keys = sorted(set(left_groups) | set(right_groups), key=_identity_sort_key)

    rows: list[CompareRowRead] = []
    used_fields: set[str] = set()
    counters = {
        SAME_STATUS: 0,
        CHANGED_STATUS: 0,
        LEFT_ONLY_STATUS: 0,
        RIGHT_ONLY_STATUS: 0,
    }

    for identity in all_keys:
        left_records = sorted(left_groups.get(identity, []), key=_record_sort_key)
        right_records = sorted(right_groups.get(identity, []), key=_record_sort_key)
        for index in range(max(len(left_records), len(right_records))):
            left_record = left_records[index] if index < len(left_records) else None
            right_record = right_records[index] if index < len(right_records) else None

            row_fields = _collect_fields(left_record, right_record)
            used_fields.update(row_fields)
            different_fields = _different_fields(row_fields, left_record, right_record)

            if left_record is None:
                diff_status = RIGHT_ONLY_STATUS
            elif right_record is None:
                diff_status = LEFT_ONLY_STATUS
            elif different_fields:
                diff_status = CHANGED_STATUS
            else:
                diff_status = SAME_STATUS
            counters[diff_status] += 1

            rows.append(
                CompareRowRead(
                    compare_key=_build_compare_row_key(identity, index),
                    match_basis=identity.basis,
                    diff_status=diff_status,
                    different_fields=different_fields,
                    left=_serialize_record_side(left_record, row_fields),
                    right=_serialize_record_side(right_record, row_fields),
                )
            )

    fields = _order_fields(used_fields)
    rows = [_align_row_fields(row, fields) for row in rows]

    return BatchCompareRead(
        left_batch=CompareBatchMetaRead(
            id=left_batch.id,
            batch_name=left_batch.batch_name,
            status=left_batch.status.value,
            record_count=len(left_batch.normalized_records),
        ),
        right_batch=CompareBatchMetaRead(
            id=right_batch.id,
            batch_name=right_batch.batch_name,
            status=right_batch.status.value,
            record_count=len(right_batch.normalized_records),
        ),
        fields=fields,
        total_row_count=len(rows),
        same_row_count=counters[SAME_STATUS],
        changed_row_count=counters[CHANGED_STATUS],
        left_only_count=counters[LEFT_ONLY_STATUS],
        right_only_count=counters[RIGHT_ONLY_STATUS],
        rows=rows,
    )


def build_compare_export_workbook(payload: CompareExportRequest) -> tuple[bytes, str]:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = '差异概览'
    _write_summary_sheet(summary_sheet, payload)

    left_sheet = workbook.create_sheet('左侧数据')
    right_sheet = workbook.create_sheet('右侧数据')
    _write_data_sheet(left_sheet, payload.rows, payload.fields, side='left')
    _write_data_sheet(right_sheet, payload.rows, payload.fields, side='right')

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    file_name = _build_compare_export_name(payload.left_batch_name, payload.right_batch_name)
    return buffer.getvalue(), file_name


def _get_batch_with_records(db: Session, batch_id: str) -> ImportBatch:
    batch = (
        db.query(ImportBatch)
        .options(
            selectinload(ImportBatch.normalized_records).selectinload(NormalizedRecord.source_file),
            selectinload(ImportBatch.source_files).load_only(SourceFile.id, SourceFile.file_name),
        )
        .filter(ImportBatch.id == batch_id)
        .first()
    )
    if batch is None:
        raise BatchNotFoundError(f"Import batch '{batch_id}' was not found.")
    return batch


def _group_records_by_identity(records: list[NormalizedRecord]) -> dict[CompareIdentity, list[NormalizedRecord]]:
    grouped: dict[CompareIdentity, list[NormalizedRecord]] = defaultdict(list)
    for record in records:
        grouped[_build_compare_identity(record)].append(record)
    return grouped


def _build_compare_identity(record: NormalizedRecord) -> CompareIdentity:
    candidates = (
        ('employee_id', _normalize_identity_value(record.employee_id)),
        ('id_number', _normalize_identity_value(record.id_number)),
        ('social_security_number', _normalize_identity_value(record.social_security_number)),
        ('housing_fund_account', _normalize_identity_value(record.housing_fund_account)),
        (
            'person_name_company',
            _normalize_identity_value(
                '|'.join(
                    item
                    for item in (
                        _normalize_identity_value(record.person_name),
                        _normalize_identity_value(record.company_name),
                    )
                    if item
                )
            ),
        ),
        ('person_name', _normalize_identity_value(record.person_name)),
    )
    for basis, value in candidates:
        if value:
            return CompareIdentity(basis=basis, value=value)
    fallback = f'{record.source_file_name or record.source_file.file_name}#{record.source_row_number}'
    return CompareIdentity(basis='source_row', value=fallback)


def _normalize_identity_value(value: Optional[str]) -> str:
    if value is None:
        return ''
    return str(value).strip().lower()


def _identity_sort_key(identity: CompareIdentity) -> tuple[str, str]:
    return identity.value, identity.basis


def _record_sort_key(record: NormalizedRecord) -> tuple[str, int, str]:
    source_name = record.source_file_name or record.source_file.file_name or ''
    return source_name, record.source_row_number, record.person_name or ''


def _collect_fields(left_record: Optional[NormalizedRecord], right_record: Optional[NormalizedRecord]) -> list[str]:
    fields: list[str] = []
    for field in PREFERRED_COMPARE_FIELDS:
        if _record_has_value(left_record, field) or _record_has_value(right_record, field):
            fields.append(field)
    return fields


def _record_has_value(record: Optional[NormalizedRecord], field: str) -> bool:
    if record is None:
        return False
    value = getattr(record, field, None)
    return _normalize_compare_value(value) is not None


def _different_fields(fields: list[str], left_record: Optional[NormalizedRecord], right_record: Optional[NormalizedRecord]) -> list[str]:
    if left_record is None or right_record is None:
        return fields
    return [
        field
        for field in fields
        if _normalize_compare_value(getattr(left_record, field, None))
        != _normalize_compare_value(getattr(right_record, field, None))
    ]


def _serialize_record_side(record: Optional[NormalizedRecord], fields: list[str]) -> CompareRecordSideRead:
    if record is None:
        return CompareRecordSideRead(
            record_id=None,
            source_file_id=None,
            source_file_name=None,
            source_row_number=None,
            values={field: None for field in fields},
        )
    return CompareRecordSideRead(
        record_id=record.id,
        source_file_id=record.source_file_id,
        source_file_name=record.source_file_name or record.source_file.file_name,
        source_row_number=record.source_row_number,
        values={field: _serialize_value(getattr(record, field, None)) for field in fields},
    )


def _align_row_fields(row: CompareRowRead, fields: list[str]) -> CompareRowRead:
    left_values = {field: row.left.values.get(field) for field in fields}
    right_values = {field: row.right.values.get(field) for field in fields}
    return row.model_copy(
        update={
            'left': row.left.model_copy(update={'values': left_values}),
            'right': row.right.model_copy(update={'values': right_values}),
        }
    )


def _order_fields(fields: set[str]) -> list[str]:
    ordered = [field for field in PREFERRED_COMPARE_FIELDS if field in fields]
    extras = sorted(field for field in fields if field not in PREFERRED_COMPARE_FIELDS)
    return [*ordered, *extras]


def _build_compare_row_key(identity: CompareIdentity, index: int) -> str:
    return f'{identity.basis}:{identity.value}:{index + 1}'


def _serialize_value(value: Optional[object]) -> Optional[object]:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return format(value, 'f')
    return value


def _normalize_compare_value(value: Optional[object]) -> Optional[str]:
    serialized = _serialize_value(value)
    if serialized is None:
        return None
    normalized = str(serialized).strip()
    return normalized or None


def _write_summary_sheet(sheet, payload: CompareExportRequest) -> None:
    rows = payload.rows
    same_count = sum(1 for row in rows if row.diff_status == SAME_STATUS)
    changed_count = sum(1 for row in rows if row.diff_status == CHANGED_STATUS)
    left_only_count = sum(1 for row in rows if row.diff_status == LEFT_ONLY_STATUS)
    right_only_count = sum(1 for row in rows if row.diff_status == RIGHT_ONLY_STATUS)

    summary_lines = [
        ('左侧批次', payload.left_batch_name),
        ('右侧批次', payload.right_batch_name),
        ('字段数', len(payload.fields)),
        ('总对比行数', len(rows)),
        ('完全一致', same_count),
        ('存在差异', changed_count),
        ('仅左侧存在', left_only_count),
        ('仅右侧存在', right_only_count),
    ]
    for index, (label, value) in enumerate(summary_lines, start=1):
        sheet.cell(row=index, column=1, value=label)
        sheet.cell(row=index, column=2, value=value)
    for cell in sheet[1]:
        cell.font = HEADER_FONT
    sheet.column_dimensions['A'].width = 16
    sheet.column_dimensions['B'].width = 36


def _write_data_sheet(sheet, rows: list[CompareRowInput], fields: list[str], *, side: str) -> None:
    headers = ['compare_key', 'diff_status', 'different_fields', 'source_file_name', 'source_row_number', *fields]
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for row_index, row in enumerate(rows, start=2):
        side_record = row.left if side == 'left' else row.right
        values = [
            row.compare_key,
            row.diff_status,
            ', '.join(row.different_fields),
            side_record.source_file_name,
            side_record.source_row_number,
            *[_serialize_export_cell(side_record.values.get(field)) for field in fields],
        ]
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column_index, value=value)
        if row.diff_status != SAME_STATUS:
            for column_index in range(1, len(headers) + 1):
                sheet.cell(row=row_index, column=column_index).fill = DIFF_ROW_FILL

    sheet.freeze_panes = 'A2'
    for index, header in enumerate(headers, start=1):
        width = 14 if index <= 5 else min(max(len(header) + 4, 12), 24)
        sheet.column_dimensions[get_column_letter(index)].width = width


def _serialize_export_cell(value: Optional[object]) -> Optional[object]:
    if value is None:
        return None
    if isinstance(value, (str, int, float)):
        return value
    return str(value)


def _build_compare_export_name(left_batch_name: str, right_batch_name: str) -> str:
    left_name = _sanitize_file_stem(left_batch_name)
    right_name = _sanitize_file_stem(right_batch_name)
    return f'compare_{left_name}_vs_{right_name}.xlsx'


def _sanitize_file_stem(value: str) -> str:
    sanitized = ''.join(char if char not in '\\/:*?"<>|' else '_' for char in value).strip()
    sanitized = '_'.join(part for part in sanitized.split() if part)
    sanitized = sanitized.replace('__', '_')
    sanitized = sanitized.strip('_.')
    return (sanitized or 'batch')[:24]
