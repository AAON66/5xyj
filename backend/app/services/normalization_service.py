from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from backend.app.models.enums import SourceFileKind
from backend.app.models.normalized_record import NormalizedRecord
from backend.app.parsers import HeaderExtraction, extract_header_structure
from backend.app.services.header_normalizer import (
    HeaderMappingDecision,
    HeaderNormalizationResult,
    normalize_header_extraction,
    normalize_header_extraction_with_fallback,
)
from backend.app.validators import RowFilterDecision, classify_row

AMOUNT_FIELDS = {
    "payment_base",
    "payment_salary",
    "housing_fund_base",
    "housing_fund_personal",
    "housing_fund_company",
    "housing_fund_total",
    "total_amount",
    "company_total_amount",
    "personal_total_amount",
    "pension_company",
    "pension_personal",
    "medical_company",
    "medical_personal",
    "medical_maternity_company",
    "maternity_amount",
    "unemployment_company",
    "unemployment_personal",
    "injury_company",
    "supplementary_medical_company",
    "supplementary_pension_company",
    "large_medical_personal",
    "late_fee",
    "interest",
}

PLACEHOLDER_STRINGS = {"", "-", "--", "??", "none", "null", "(??)"}
MERGE_VALUE_FIELDS = (
    "person_name",
    "id_type",
    "id_number",
    "employee_id",
    "social_security_number",
    "company_name",
    "region",
    "billing_period",
    "period_start",
    "period_end",
    "payment_base",
    "payment_salary",
    "housing_fund_account",
    "housing_fund_base",
    "housing_fund_personal",
    "housing_fund_company",
    "housing_fund_total",
    "total_amount",
    "company_total_amount",
    "personal_total_amount",
    "pension_company",
    "pension_personal",
    "medical_company",
    "medical_personal",
    "medical_maternity_company",
    "maternity_amount",
    "unemployment_company",
    "unemployment_personal",
    "injury_company",
    "supplementary_medical_company",
    "supplementary_pension_company",
    "large_medical_personal",
    "late_fee",
    "interest",
    "raw_sheet_name",
    "raw_header_signature",
    "source_file_name",
)


@dataclass(slots=True)
class NormalizedPreviewRecord:
    source_row_number: int
    values: dict[str, Any]
    unmapped_values: dict[str, Any]
    raw_values: dict[str, Any]
    raw_payload: dict[str, Any]


@dataclass(slots=True)
class StandardizationResult:
    source_file: str
    sheet_name: str
    raw_header_signature: str
    records: list[NormalizedPreviewRecord]
    filtered_rows: list[RowFilterDecision]
    unmapped_headers: list[str]


@dataclass(slots=True)
class SourceRecordBundle:
    source_file_id: str
    source_file_name: str
    source_kind: str
    standardized: StandardizationResult


@dataclass(slots=True)
class _MergedRecordEntry:
    source_file_id: str
    source_row_number: int
    values: dict[str, Any]
    raw_payload: dict[str, Any]


def standardize_workbook(
    path: str | Path,
    *,
    region: str | None = None,
    company_name: str | None = None,
    source_file_name: str | None = None,
    extraction: HeaderExtraction | None = None,
    normalization: HeaderNormalizationResult | None = None,
) -> StandardizationResult:
    workbook_path = Path(path)
    runtime_extraction = extraction or extract_header_structure(workbook_path)
    runtime_normalization = normalization or normalize_header_extraction(runtime_extraction, region=region)
    return _standardize_rows(
        workbook_path,
        runtime_extraction,
        runtime_normalization,
        region=region,
        company_name=company_name,
        source_file_name=source_file_name,
    )


async def standardize_workbook_with_fallback(
    path: str | Path,
    *,
    region: str | None = None,
    company_name: str | None = None,
    source_file_name: str | None = None,
    confidence_threshold: float = 0.8,
    extraction: HeaderExtraction | None = None,
    normalization: HeaderNormalizationResult | None = None,
) -> StandardizationResult:
    workbook_path = Path(path)
    runtime_extraction = extraction or extract_header_structure(workbook_path)
    runtime_normalization = normalization or await normalize_header_extraction_with_fallback(
        runtime_extraction,
        region=region,
        confidence_threshold=confidence_threshold,
    )
    return _standardize_rows(
        workbook_path,
        runtime_extraction,
        runtime_normalization,
        region=region,
        company_name=company_name,
        source_file_name=source_file_name,
    )


def build_normalized_models(
    result: StandardizationResult,
    *,
    batch_id: str,
    source_file_id: str,
) -> list[NormalizedRecord]:
    records: list[NormalizedRecord] = []
    for preview in result.records:
        model_kwargs = {
            "batch_id": batch_id,
            "source_file_id": source_file_id,
            "source_row_number": preview.source_row_number,
            "raw_sheet_name": result.sheet_name,
            "raw_header_signature": result.raw_header_signature,
            "source_file_name": result.source_file,
            "raw_payload": preview.raw_payload,
        }
        model_kwargs.update(preview.values)
        records.append(NormalizedRecord(**model_kwargs))
    return records


def merge_batch_standardized_records(
    source_bundles: list[SourceRecordBundle],
    *,
    batch_id: str,
) -> list[NormalizedRecord]:
    social_bundles = [item for item in source_bundles if item.source_kind == SourceFileKind.SOCIAL_SECURITY.value]
    housing_bundles = [item for item in source_bundles if item.source_kind == SourceFileKind.HOUSING_FUND.value]

    entries: list[_MergedRecordEntry] = []
    index: dict[tuple[str, ...], list[_MergedRecordEntry]] = {}

    base_bundles = social_bundles if social_bundles else housing_bundles
    for bundle in base_bundles:
        for record in bundle.standardized.records:
            entry = _create_entry(bundle, record)
            entries.append(entry)
            key = _merge_key(record)
            if key is not None:
                index.setdefault(key, []).append(entry)

    if social_bundles:
        overlay_bundles = housing_bundles
    else:
        overlay_bundles = []

    for bundle in overlay_bundles:
        for record in bundle.standardized.records:
            key = _merge_key(record)
            matches = index.get(key, []) if key is not None else []
            if len(matches) == 1:
                _merge_entry(matches[0], bundle, record)
                continue
            entry = _create_entry(bundle, record)
            entries.append(entry)
            if key is not None:
                index.setdefault(key, []).append(entry)

    models: list[NormalizedRecord] = []
    for entry in entries:
        model_kwargs = {
            "batch_id": batch_id,
            "source_file_id": entry.source_file_id,
            "source_row_number": entry.source_row_number,
            "raw_payload": entry.raw_payload,
        }
        for field_name in MERGE_VALUE_FIELDS:
            value = entry.values.get(field_name)
            if value is not None:
                model_kwargs[field_name] = value
        models.append(NormalizedRecord(**model_kwargs))
    return models


def _create_entry(bundle: SourceRecordBundle, record: NormalizedPreviewRecord) -> _MergedRecordEntry:
    raw_payload = {
        **record.raw_payload,
        "merged_sources": [
            {
                "source_kind": bundle.source_kind,
                "source_file_id": bundle.source_file_id,
                "source_file_name": bundle.source_file_name,
                "sheet_name": bundle.standardized.sheet_name,
                "raw_header_signature": bundle.standardized.raw_header_signature,
                "source_row_number": record.source_row_number,
                "raw_values": {_key: _json_safe(_value) for _key, _value in record.raw_values.items()},
                "unmapped_values": {_key: _json_safe(_value) for _key, _value in record.unmapped_values.items()},
            }
        ],
    }
    return _MergedRecordEntry(
        source_file_id=bundle.source_file_id,
        source_row_number=record.source_row_number,
        values=dict(record.values),
        raw_payload=raw_payload,
    )


def _merge_entry(entry: _MergedRecordEntry, bundle: SourceRecordBundle, record: NormalizedPreviewRecord) -> None:
    for field_name, value in record.values.items():
        existing = entry.values.get(field_name)
        if existing is None and value is not None:
            entry.values[field_name] = value
            continue
        if value is None or existing == value:
            continue
        conflicts = entry.raw_payload.setdefault("merged_field_conflicts", {})
        values = conflicts.setdefault(field_name, [])
        for item in (existing, value):
            safe_item = _json_safe(item)
            if safe_item not in values:
                values.append(safe_item)

    merged_sources = entry.raw_payload.setdefault("merged_sources", [])
    merged_sources.append(
        {
            "source_kind": bundle.source_kind,
            "source_file_id": bundle.source_file_id,
            "source_file_name": bundle.source_file_name,
            "sheet_name": bundle.standardized.sheet_name,
            "raw_header_signature": bundle.standardized.raw_header_signature,
            "source_row_number": record.source_row_number,
            "raw_values": {key: _json_safe(value) for key, value in record.raw_values.items()},
            "unmapped_values": {key: _json_safe(value) for key, value in record.unmapped_values.items()},
        }
    )


def _merge_key(record: NormalizedPreviewRecord) -> tuple[str, ...] | None:
    id_number = _normalize_identity_value(record.values.get("id_number"))
    company_name = _normalize_identity_value(record.values.get("company_name"))
    person_name = _normalize_identity_value(record.values.get("person_name"))

    if id_number and company_name:
        return ("id_company", id_number, company_name)
    if id_number:
        return ("id", id_number)
    if person_name and company_name:
        return ("name_company", person_name, company_name)
    if person_name:
        return ("name", person_name)
    return None


def _normalize_identity_value(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip().lower()
    return text or None


def _standardize_rows(
    workbook_path: Path,
    extraction: HeaderExtraction,
    normalization: HeaderNormalizationResult,
    *,
    region: str | None,
    company_name: str | None,
    source_file_name: str | None,
) -> StandardizationResult:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook[extraction.sheet_name]
        max_column = max(column.column_index for column in extraction.columns)
        rows = list(
            sheet.iter_rows(
                min_row=extraction.data_start_row,
                max_row=sheet.max_row,
                max_col=max_column,
                values_only=True,
            )
        )
    finally:
        workbook.close()

    records: list[NormalizedPreviewRecord] = []
    filtered_rows: list[RowFilterDecision] = []
    row_number = extraction.data_start_row
    for row in rows:
        selected_values = [row[column.column_index - 1] for column in extraction.columns]
        filter_decision = classify_row(selected_values, row_number=row_number)
        if not filter_decision.keep:
            filtered_rows.append(filter_decision)
            row_number += 1
            continue

        records.append(
            _build_preview_record(
                selected_values,
                extraction=extraction,
                decisions=normalization.decisions,
                region=region,
                company_name=company_name,
                source_file_name=source_file_name or workbook_path.name,
                row_number=row_number,
            )
        )
        row_number += 1

    return StandardizationResult(
        source_file=source_file_name or workbook_path.name,
        sheet_name=extraction.sheet_name,
        raw_header_signature=extraction.raw_header_signature,
        records=records,
        filtered_rows=filtered_rows,
        unmapped_headers=normalization.unmapped_headers,
    )


def _build_preview_record(
    selected_values: list[object],
    *,
    extraction: HeaderExtraction,
    decisions: list[HeaderMappingDecision],
    region: str | None,
    company_name: str | None,
    source_file_name: str,
    row_number: int,
) -> NormalizedPreviewRecord:
    raw_values: dict[str, Any] = {}
    values: dict[str, Any] = {
        "company_name": company_name,
        "region": region,
        "raw_sheet_name": extraction.sheet_name,
        "raw_header_signature": extraction.raw_header_signature,
        "source_file_name": source_file_name,
    }
    unmapped_values: dict[str, Any] = {}
    field_conflicts: dict[str, list[Any]] = {}

    for column, decision, raw_value in zip(extraction.columns, decisions, selected_values, strict=True):
        cleaned_raw = _clean_cell_value(raw_value)
        raw_values[column.signature] = cleaned_raw
        if decision.canonical_field is None:
            if cleaned_raw is not None:
                unmapped_values[column.signature] = cleaned_raw
            continue

        canonical_value = _coerce_canonical_value(decision.canonical_field, cleaned_raw)
        if canonical_value is None:
            continue
        _assign_canonical_value(values, field_conflicts, decision.canonical_field, canonical_value)

    _derive_period_fields(values)
    raw_payload = {
        "raw_values": {key: _json_safe(value) for key, value in raw_values.items()},
        "unmapped_values": {key: _json_safe(value) for key, value in unmapped_values.items()},
    }
    if field_conflicts:
        raw_payload["field_conflicts"] = {
            key: [_json_safe(value) for value in conflict_values]
            for key, conflict_values in field_conflicts.items()
        }

    return NormalizedPreviewRecord(
        source_row_number=row_number,
        values=values,
        unmapped_values=unmapped_values,
        raw_values=raw_values,
        raw_payload=raw_payload,
    )


def _assign_canonical_value(
    values: dict[str, Any],
    field_conflicts: dict[str, list[Any]],
    field_name: str,
    value: Any,
) -> None:
    existing = values.get(field_name)
    if existing is None:
        values[field_name] = value
        return
    if existing == value:
        return
    field_conflicts.setdefault(field_name, [existing]).append(value)


def _derive_period_fields(values: dict[str, Any]) -> None:
    billing_period = values.get("billing_period")
    period_start = values.get("period_start")
    period_end = values.get("period_end")

    if not billing_period and isinstance(period_start, str):
        if isinstance(period_end, str) and len(period_start) >= 7 and period_start[:7] == period_end[:7]:
            values["billing_period"] = period_start[:7]
        elif len(period_start) >= 7:
            values["billing_period"] = period_start[:7]

    if not period_start and isinstance(billing_period, str):
        values["period_start"] = billing_period
    if not period_end and isinstance(billing_period, str):
        values["period_end"] = billing_period


def _coerce_canonical_value(field_name: str, value: Any) -> Any:
    if value is None:
        return None
    if field_name in AMOUNT_FIELDS:
        return _to_decimal(value)
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return value


def _clean_cell_value(value: object) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.lower() in PLACEHOLDER_STRINGS:
            return None
        return stripped
    return value


def _to_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        return Decimal(str(value))
    if isinstance(value, str):
        candidate = value.strip().replace(",", "")
        if not candidate or candidate.lower() in PLACEHOLDER_STRINGS:
            return None
        try:
            return Decimal(candidate)
        except InvalidOperation:
            return None
    return None


def _json_safe(value: Any) -> Any:
    if isinstance(value, Decimal):
        return format(value, "f")
    return value
