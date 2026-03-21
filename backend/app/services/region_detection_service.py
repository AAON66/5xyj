from __future__ import annotations

import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import httpx
from openpyxl import load_workbook

from backend.app.core.config import get_settings

DEEPSEEK_COMPLETIONS_PATH = "/chat/completions"
DEFAULT_REGION_LLM_TIMEOUT = 45.0
REGION_LLM_CONFIDENCE_THRESHOLD = 0.8
STRONG_RULE_CONFIDENCE_THRESHOLD = 0.96

REGION_LABELS = {
    "guangzhou": "\u5e7f\u5dde",
    "hangzhou": "\u676d\u5dde",
    "xiamen": "\u53a6\u95e8",
    "shenzhen": "\u6df1\u5733",
    "wuhan": "\u6b66\u6c49",
    "changsha": "\u957f\u6c99",
}

REGION_KEYWORDS = {
    "guangzhou": ("\u5e7f\u5dde", "\u5e7f\u5206", "\u89c6\u64ad", "\u7a57"),
    "hangzhou": ("\u676d\u5dde", "\u805a\u53d8", "\u88c2\u53d8"),
    "xiamen": ("\u53a6\u95e8",),
    "shenzhen": ("\u6df1\u5733", "\u5218\u8273\u73b2"),
    "wuhan": ("\u6b66\u6c49",),
    "changsha": ("\u957f\u6c99",),
}

REGION_CONTENT_PATTERNS = {
    "guangzhou": (
        "\u793e\u4f1a\u4fdd\u9669\u8d39\u7533\u62a5\u4e2a\u4eba\u660e\u7ec6\u8868",
        "\u8d39\u6b3e\u6240\u5c5e\u671f\u8d77",
        "\u8d39\u6b3e\u6240\u5c5e\u671f\u6b62",
        "\u4e2a\u4eba\u793e\u4fdd\u53f7",
        "\u5355\u4f4d\u90e8\u5206\u5408\u8ba1",
        "\u4e2a\u4eba\u90e8\u5206\u5408\u8ba1",
        "\u5e94\u7f34\u91d1\u989d\u5408\u8ba1",
        "\u5e7f\u5dde\u4f4f\u623f\u516c\u79ef\u91d1\u7ba1\u7406\u4e2d\u5fc3",
    ),
    "hangzhou": (
        "\u5355\u4f4d\u793e\u4fdd\u8d39\u804c\u5de5\u5168\u9669\u79cd\u7533\u62a5\u660e\u7ec6",
        "\u57fa\u672c\u517b\u8001\u5e94\u7f34\u8d39\u989d",
        "\u673a\u5173\u517b\u8001\u5e94\u7f34\u8d39\u989d",
        "\u516c\u52a1\u5458\u8865\u52a9\u5e94\u7f34\u8d39\u989d",
        "\u5355\u4f4d\u90e8\u5206",
        "\u4e2a\u4eba\u90e8\u5206",
    ),
    "xiamen": (
        "\u804c\u5de5\u793e\u4fdd\u5bf9\u8d26\u5355\u660e\u7ec6\u67e5\u8be2",
        "\u5355\u4f4d\u603b\u989d",
        "\u4e2a\u4eba\u603b\u989d",
        "\u672c\u91d1\u5408\u8ba1",
        "\u6ede\u7eb3\u91d1",
        "\u5229\u606f",
    ),
    "shenzhen": (
        "\u7533\u62a5\u660e\u7ec6",
        "\u4e2a\u4eba\u793e\u4fdd\u5408\u8ba1",
        "\u5355\u4f4d\u793e\u4fdd\u5408\u8ba1",
        "\u5730\u65b9\u8865\u5145\u533b\u7597",
        "\u5730\u65b9\u8865\u5145\u517b\u8001",
        "\u5728\u804c\u4eba\u5458",
        "\u9000\u4f11\u4eba\u5458",
        "\u5bb6\u5c5e\u7edf\u7b79\u4eba\u5458",
        "\u5355\u7b14\u7f34\u5b58\u6e05\u5355",
    ),
    "wuhan": (
        "\u804c\u5de5\u660e\u7ec6",
        "\u5355\u4f4d\u7f34\u7eb3",
        "\u4e2a\u4eba\u7f34\u7eb3",
        "\u517b\u8001\u4fdd\u9669\u5e94\u7f34\u8d39\u989d",
        "\u4f01\u4e1a\u804c\u5de5\u57fa\u672c\u533b\u7597\u5e94\u7f34\u8d39\u989d",
    ),
    "changsha": (
        "\u804c\u5de5\u5927\u989d\u533b\u7597\u4e92\u52a9\u4fdd\u9669(\u4e2a\u4eba\u7f34\u7eb3)",
        "\u804c\u5de5\u57fa\u672c\u517b\u8001\u4fdd\u9669(\u5355\u4f4d\u7f34\u7eb3)",
        "\u804c\u5de5\u57fa\u672c\u533b\u7597\u4fdd\u9669(\u5355\u4f4d\u7f34\u7eb3)",
        "\u603b\u8ba1",
        "sheet4",
    ),
}

REGION_SHEET_HINTS = {
    "xiamen": ("\u804c\u5de5\u793e\u4fdd\u5bf9\u8d26\u5355\u660e\u7ec6\u67e5\u8be2",),
    "shenzhen": ("\u7533\u62a5\u660e\u7ec6", "\u5355\u7b14\u7f34\u5b58\u6e05\u5355"),
    "changsha": ("sheet4", "\u88681"),
}

PLACEHOLDER_TEXT = {"", "none", "null", "nan"}


@dataclass(slots=True)
class RegionDetectionResult:
    region: str | None
    confidence: float
    source: str
    reason: str
    local_confidence: float | None = None
    llm_confidence: float | None = None


@dataclass(slots=True)
class LLMRegionResult:
    region: str | None
    confidence: float | None
    candidate_regions: list[str]
    status: str
    reason: str


@dataclass(slots=True)
class WorkbookRegionContext:
    filename: str
    source_kind: str | None
    sheet_names: list[str]
    sample_text: str


def detect_region_from_filename(filename: str) -> str | None:
    return detect_region_with_local_rules(filename).region


def detect_region_for_workbook(
    workbook_path: str | Path,
    *,
    filename: str | None = None,
    source_kind: str | None = None,
) -> RegionDetectionResult:
    workbook_context = build_workbook_region_context(workbook_path, filename=filename, source_kind=source_kind)
    local_result = detect_region_with_local_rules(workbook_context.filename, workbook_context=workbook_context)
    if local_result.region and local_result.confidence >= STRONG_RULE_CONFIDENCE_THRESHOLD:
        return local_result
    llm_result = detect_region_with_llm_sync(workbook_context)
    return merge_region_detection_results(local_result, llm_result)


def detect_region_with_local_rules(
    filename: str,
    *,
    workbook_context: WorkbookRegionContext | None = None,
) -> RegionDetectionResult:
    scores = {region: 0.0 for region in REGION_LABELS}
    reasons = {region: [] for region in REGION_LABELS}

    for region, label in REGION_LABELS.items():
        if label in filename:
            scores[region] += 0.72
            reasons[region].append(f"filename_label:{label}")

    for region, keywords in REGION_KEYWORDS.items():
        for keyword in keywords:
            if keyword in filename:
                increment = 0.18 if keyword == REGION_LABELS[region] else min(0.16, 0.08 + len(keyword) * 0.02)
                scores[region] += increment
                reasons[region].append(f"filename_keyword:{keyword}")

    if workbook_context is not None:
        sample_text = workbook_context.sample_text
        normalized_text = _normalize_text(sample_text)
        for region, label in REGION_LABELS.items():
            if label in sample_text:
                scores[region] += 0.65
                reasons[region].append(f"workbook_label:{label}")

        for region, patterns in REGION_CONTENT_PATTERNS.items():
            hits = [pattern for pattern in patterns if _normalize_text(pattern) in normalized_text]
            for pattern in hits[:4]:
                scores[region] += 0.18
                reasons[region].append(f"content:{pattern}")

        joined_sheet_names = " | ".join(workbook_context.sheet_names)
        normalized_sheet_names = _normalize_text(joined_sheet_names)
        for region, hints in REGION_SHEET_HINTS.items():
            for hint in hints:
                if _normalize_text(hint) in normalized_sheet_names:
                    scores[region] += 0.22
                    reasons[region].append(f"sheet:{hint}")

    ranked = sorted(scores.items(), key=lambda item: item[1], reverse=True)
    best_region, best_score = ranked[0]
    second_score = ranked[1][1] if len(ranked) > 1 else 0.0
    margin = best_score - second_score

    if best_score < 0.3:
        return RegionDetectionResult(
            region=None,
            confidence=0.0,
            source="rule",
            reason="No local region signals were strong enough.",
            local_confidence=0.0,
        )

    confidence = min(0.99, best_score if margin >= 0.12 else best_score - 0.08)
    confidence = max(confidence, 0.31)
    return RegionDetectionResult(
        region=best_region,
        confidence=confidence,
        source="rule",
        reason=", ".join(reasons[best_region]) or "local_signals",
        local_confidence=confidence,
    )


def build_workbook_region_context(
    workbook_path: str | Path,
    *,
    filename: str | None = None,
    source_kind: str | None = None,
) -> WorkbookRegionContext:
    path = Path(workbook_path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_names = [sheet.title for sheet in workbook.worksheets[:6]]
        fragments: list[str] = []
        for sheet in workbook.worksheets[:6]:
            fragments.append(sheet.title)
            max_row = min(sheet.max_row, 14)
            max_column = min(sheet.max_column, 12)
            for row_index in range(1, max_row + 1):
                row_parts: list[str] = []
                for column_index in range(1, max_column + 1):
                    value = sheet.cell(row=row_index, column=column_index).value
                    cleaned = _clean_cell_text(value)
                    if cleaned is not None:
                        row_parts.append(cleaned)
                if row_parts:
                    fragments.append(" | ".join(row_parts))
        sample_text = "\n".join(fragments)
    finally:
        workbook.close()

    return WorkbookRegionContext(
        filename=filename or path.name,
        source_kind=source_kind,
        sheet_names=sheet_names,
        sample_text=sample_text,
    )


def detect_region_with_llm_sync(workbook_context: WorkbookRegionContext) -> LLMRegionResult:
    settings = get_settings()
    if not settings.enable_llm_fallback:
        return LLMRegionResult(None, None, [], "disabled", "LLM fallback is disabled.")
    if not settings.deepseek_api_key:
        return LLMRegionResult(None, None, [], "skipped_no_api_key", "DeepSeek API key is not configured.")

    payload = {
        "model": settings.deepseek_model,
        "temperature": 0.0,
        "response_format": {"type": "json_object"},
        "messages": [
            {
                "role": "system",
                "content": (
                    "???????? Excel ???????"
                    "??????????????? region????? null?"
                    "????????????sheet ??????????????"
                    "????? JSON?????? region, confidence, candidate_regions, reason?"
                ),
            },
            {
                "role": "user",
                "content": json.dumps(
                    {
                        "filename": workbook_context.filename,
                        "source_kind": workbook_context.source_kind,
                        "sheet_names": workbook_context.sheet_names,
                        "sample_text": workbook_context.sample_text[:4000],
                        "candidate_regions": list(REGION_LABELS),
                    },
                    ensure_ascii=False,
                ),
            },
        ],
    }

    try:
        with httpx.Client(base_url=settings.deepseek_api_base_url, timeout=DEFAULT_REGION_LLM_TIMEOUT) as client:
            response = client.post(
                DEEPSEEK_COMPLETIONS_PATH,
                headers={"Authorization": f"Bearer {settings.deepseek_api_key}"},
                json=payload,
            )
            response.raise_for_status()
    except httpx.HTTPError as exc:
        return LLMRegionResult(None, None, [], "error", str(exc))

    try:
        content = response.json()["choices"][0]["message"]["content"]
        parsed = json.loads(_extract_json_content(content))
    except (KeyError, IndexError, TypeError, json.JSONDecodeError):
        return LLMRegionResult(None, None, [], "invalid_response", "DeepSeek returned an invalid response payload.")

    region = parsed.get("region")
    candidate_regions = [item for item in parsed.get("candidate_regions", []) if item in REGION_LABELS]
    confidence = parsed.get("confidence")
    if region not in REGION_LABELS:
        region = None
    if region and region not in candidate_regions:
        candidate_regions.insert(0, region)

    return LLMRegionResult(
        region=region,
        confidence=float(confidence) if confidence is not None else None,
        candidate_regions=candidate_regions,
        status="success",
        reason=str(parsed.get("reason", "")),
    )


def merge_region_detection_results(local_result: RegionDetectionResult, llm_result: LLMRegionResult) -> RegionDetectionResult:
    llm_confidence = llm_result.confidence or 0.0
    local_confidence = local_result.confidence or 0.0

    if local_result.region and llm_result.region and local_result.region == llm_result.region:
        return RegionDetectionResult(
            region=local_result.region,
            confidence=max(local_confidence, llm_confidence, 0.9),
            source="llm+rule",
            reason=f"rule={local_result.reason}; llm={llm_result.reason or llm_result.status}",
            local_confidence=local_confidence,
            llm_confidence=llm_result.confidence,
        )

    if local_result.region and local_confidence >= STRONG_RULE_CONFIDENCE_THRESHOLD and llm_result.region != local_result.region:
        return RegionDetectionResult(
            region=local_result.region,
            confidence=local_confidence,
            source="rule",
            reason=local_result.reason,
            local_confidence=local_confidence,
            llm_confidence=llm_result.confidence,
        )

    if llm_result.region and llm_confidence >= REGION_LLM_CONFIDENCE_THRESHOLD:
        return RegionDetectionResult(
            region=llm_result.region,
            confidence=llm_confidence,
            source="llm",
            reason=llm_result.reason or llm_result.status,
            local_confidence=local_confidence or None,
            llm_confidence=llm_result.confidence,
        )

    return RegionDetectionResult(
        region=local_result.region,
        confidence=local_confidence,
        source=local_result.source,
        reason=local_result.reason,
        local_confidence=local_confidence,
        llm_confidence=llm_result.confidence,
    )


def _extract_json_content(content: Any) -> str:
    if isinstance(content, list):
        text = "".join(part.get("text", "") if isinstance(part, dict) else str(part) for part in content)
    else:
        text = str(content)
    stripped = text.strip()
    fenced = re.search(r"```(?:json)?\s*(\{.*\})\s*```", stripped, re.DOTALL)
    if fenced:
        return fenced.group(1)
    return stripped


def _clean_cell_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() in PLACEHOLDER_TEXT:
        return None
    return text


def _normalize_text(value: str) -> str:
    return "".join(value.lower().replace("?", "(").replace("?", ")").split())
