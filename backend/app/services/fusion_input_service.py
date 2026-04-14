from __future__ import annotations

from collections import Counter
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from backend.app.models.sync_config import SyncConfig
from backend.app.schemas.fusion_inputs import FusionBurdenDiagnostics, FusionBurdenRow
from backend.app.services.feishu_client import FeishuClient


HEADER_ALIASES = {
    "employee_id": {"工号", "员工工号"},
    "id_number": {"身份证号", "身份证", "证件号码"},
    "personal_social_burden": {"个人社保承担额", "社保个人承担额"},
    "personal_housing_burden": {"个人公积金承担额", "公积金个人承担额"},
}
FEISHU_ALLOWED_FIELDS = set(HEADER_ALIASES)


def _normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    normalized = str(value).strip()
    return normalized or None


def _to_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, AttributeError):
        return None


def build_burden_key(employee_id: str | None, id_number: str | None) -> tuple[str, str] | None:
    employee_id_value = _normalize_text(employee_id)
    id_number_value = _normalize_text(id_number)
    if employee_id_value:
        return employee_id_value, ""
    if id_number_value:
        return "", id_number_value
    return None


def _resolve_header_map(header_row: list[Any]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for index, raw_value in enumerate(header_row):
        normalized = _normalize_text(raw_value)
        if normalized is None:
            continue
        for field_name, aliases in HEADER_ALIASES.items():
            if normalized in aliases:
                mapping[index] = field_name
                break
    return mapping


def _diagnostics_from_rows(
    raw_rows: list[dict[str, Any]],
    *,
    source_kind: str,
    source_ref: str,
) -> tuple[list[FusionBurdenRow], FusionBurdenDiagnostics]:
    messages: list[str] = []
    missing_key_rows = 0
    duplicate_key_rows = 0
    key_counts: Counter[tuple[str, str]] = Counter()
    normalized_rows: list[tuple[tuple[str, str], FusionBurdenRow]] = []

    for row in raw_rows:
        key = build_burden_key(row.get("employee_id"), row.get("id_number"))
        if key is None:
            missing_key_rows += 1
            continue
        burden_row = FusionBurdenRow(
            employee_id=_normalize_text(row.get("employee_id")),
            id_number=_normalize_text(row.get("id_number")),
            personal_social_burden=_to_decimal(row.get("personal_social_burden")),
            personal_housing_burden=_to_decimal(row.get("personal_housing_burden")),
            source_kind=source_kind,
            source_ref=source_ref,
        )
        key_counts[key] += 1
        normalized_rows.append((key, burden_row))

    duplicate_keys = {key for key, count in key_counts.items() if count > 1}
    if missing_key_rows:
        messages.append(f"{source_ref}: skipped {missing_key_rows} row(s) without employee_id/id_number.")
    if duplicate_keys:
        duplicate_key_rows = sum(key_counts[key] for key in duplicate_keys)
        messages.append(f"{source_ref}: skipped {duplicate_key_rows} row(s) with duplicate employee keys.")

    rows = [row for key, row in normalized_rows if key not in duplicate_keys]
    return rows, FusionBurdenDiagnostics(
        missing_key_rows=missing_key_rows,
        duplicate_key_rows=duplicate_key_rows,
        unmatched_rows=0,
        messages=messages,
    )


def parse_burden_workbook(file_bytes: bytes, filename: str) -> tuple[list[FusionBurdenRow], FusionBurdenDiagnostics]:
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    try:
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            for header_index, row in enumerate(rows[:5]):
                header_map = _resolve_header_map(list(row))
                if not header_map:
                    continue
                raw_rows: list[dict[str, Any]] = []
                for data_row in rows[header_index + 1:]:
                    if not any(cell not in (None, "") for cell in data_row):
                        continue
                    mapped: dict[str, Any] = {}
                    for column_index, field_name in header_map.items():
                        if column_index < len(data_row):
                            mapped[field_name] = data_row[column_index]
                    raw_rows.append(mapped)
                return _diagnostics_from_rows(raw_rows, source_kind="excel", source_ref=filename)
        return [], FusionBurdenDiagnostics(messages=[f"{filename}: no supported burden headers found."])
    finally:
        workbook.close()


async def load_burden_rows_from_feishu(
    db: Session,
    client: FeishuClient,
    config_id: str,
) -> tuple[list[FusionBurdenRow], FusionBurdenDiagnostics]:
    config = db.get(SyncConfig, config_id)
    if config is None:
        raise ValueError(f"SyncConfig '{config_id}' not found")

    mapped_fields = {feishu_column: system_field for feishu_column, system_field in config.field_mapping.items() if system_field in FEISHU_ALLOWED_FIELDS}
    if not mapped_fields:
        return [], FusionBurdenDiagnostics(messages=[f"{config.name}: no burden fields mapped in SyncConfig."])

    raw_rows: list[dict[str, Any]] = []
    page_token: str | None = None
    while True:
        result = await client.search_records(config.app_token, config.table_id, page_token=page_token)
        data = result.get("data", {})
        for item in data.get("items", []):
            fields = item.get("fields", {})
            mapped: dict[str, Any] = {}
            for feishu_column, system_field in mapped_fields.items():
                mapped[system_field] = fields.get(feishu_column)
            raw_rows.append(mapped)
        if not data.get("has_more"):
            break
        page_token = data.get("page_token")

    return _diagnostics_from_rows(raw_rows, source_kind="feishu", source_ref=config.name)
