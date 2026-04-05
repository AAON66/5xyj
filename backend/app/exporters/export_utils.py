from __future__ import annotations

from copy import copy, deepcopy
from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from pathlib import Path
import json
import re
from typing import Iterable, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from backend.app.core.config import Settings, get_settings
from backend.app.models.enums import TemplateType
from backend.app.models.normalized_record import NormalizedRecord

HEADER_LIKE_PERSON_VALUES = {
    '\u59d3\u540d',
    '\u5458\u5de5\u59d3\u540d',
    '\u8eab\u4efd\u8bc1\u53f7',
    '\u8eab\u4efd\u8bc1\u53f7\u7801',
    '\u8bc1\u4ef6\u53f7',
    '\u8bc1\u4ef6\u53f7\u7801',
    '\u5de5\u53f7',
    '\u7a7a\u767d',
    '(\u7a7a\u767d)',
    '\uff08\u7a7a\u767d\uff09',
}
ID_NUMBER_PATTERN = re.compile(r'^\d{15}$|^\d{17}[\dX]$')
NON_MAINLAND_ID_NUMBER_PATTERN = re.compile(r'^[A-Z]{1,2}\d{6,10}[A-Z0-9]?$')
EXPORT_AMOUNT_FIELDS = (
    'housing_fund_personal',
    'housing_fund_company',
    'housing_fund_total',
    'total_amount',
    'company_total_amount',
    'personal_total_amount',
    'pension_company',
    'pension_personal',
    'medical_company',
    'medical_personal',
    'medical_maternity_company',
    'maternity_amount',
    'unemployment_company',
    'unemployment_personal',
    'injury_company',
    'supplementary_medical_company',
    'supplementary_pension_company',
    'large_medical_personal',
    'late_fee',
    'interest',
)
EXPORT_TEXT_FIELDS = (
    'person_name',
    'id_type',
    'id_number',
    'employee_id',
    'social_security_number',
    'company_name',
    'region',
    'billing_period',
    'period_start',
    'period_end',
    'housing_fund_account',
    'raw_sheet_name',
    'raw_header_signature',
    'source_file_name',
)
SOCIAL_AMOUNT_FIELDS = (
    'total_amount',
    'company_total_amount',
    'personal_total_amount',
    'pension_company',
    'pension_personal',
    'medical_company',
    'medical_personal',
    'medical_maternity_company',
    'maternity_amount',
    'unemployment_company',
    'unemployment_personal',
    'injury_company',
    'supplementary_medical_company',
    'supplementary_pension_company',
    'large_medical_personal',
    'late_fee',
    'interest',
)
EXPLICIT_HOUSING_COMPANY_HEADER_TOKENS = (
    '单位',
    '单位缴存额',
    '单位月缴存额',
    '单位月缴存额(元)',
    '单位月缴额',
    '单位金额',
)
EXPLICIT_HOUSING_PERSONAL_HEADER_TOKENS = (
    '个人',
    '个人缴存额',
    '个人月缴存额',
    '个人月缴存额(元)',
    '职工月缴额',
    '个人金额',
)
HOUSING_SOURCE_KIND = 'housing_fund'
HOUSING_BURDEN_CANDIDATE_RATIO = Decimal('1.5')
SOCIAL_SOURCE_KIND = 'social_security'
SOCIAL_BURDEN_CANDIDATE_RATIO = Decimal('1.5')
HEADER_NORMALIZATION_TRANSLATION = str.maketrans(
    {
        '\u3000': ' ',
        '\uff08': '(',
        '\uff09': ')',
        '\u3010': '[',
        '\u3011': ']',
        '\uff1a': ':',
        '\uff0f': '/',
    }
)
HOUSING_METADATA_HEADER_MARKERS = (
    '\u6bd4\u4f8b',
    '\u57fa\u6570',
    '\u8d26\u53f7',
)
EXPLICIT_HOUSING_COMBINED_HEADER_TOKENS = (
    '\u5355\u4f4d\u53ca\u4e2a\u4eba',
)
XIAMEN_MEDICAL_COMPANY_HEADER = (
    '\u804c\u5de5\u57fa\u672c\u533b\u7597\u4fdd\u9669\u8d39',
    '\u5355\u4f4d\u5e94\u7f34',
)
XIAMEN_MATERNITY_HEADER = (
    '\u804c\u5de5\u57fa\u672c\u533b\u7597\u4fdd\u9669\u8d39(\u751f\u80b2)',
    '\u5355\u4f4d\u5e94\u7f34',
)
SOURCE_KIND_BY_AMOUNT_FIELD = {
    'housing_fund_personal': HOUSING_SOURCE_KIND,
    'housing_fund_company': HOUSING_SOURCE_KIND,
    'housing_fund_total': HOUSING_SOURCE_KIND,
    'total_amount': SOCIAL_SOURCE_KIND,
    'company_total_amount': SOCIAL_SOURCE_KIND,
    'personal_total_amount': SOCIAL_SOURCE_KIND,
    'pension_company': SOCIAL_SOURCE_KIND,
    'pension_personal': SOCIAL_SOURCE_KIND,
    'medical_company': SOCIAL_SOURCE_KIND,
    'medical_personal': SOCIAL_SOURCE_KIND,
    'medical_maternity_company': SOCIAL_SOURCE_KIND,
    'unemployment_company': SOCIAL_SOURCE_KIND,
    'unemployment_personal': SOCIAL_SOURCE_KIND,
    'injury_company': SOCIAL_SOURCE_KIND,
    'supplementary_medical_company': SOCIAL_SOURCE_KIND,
    'supplementary_pension_company': SOCIAL_SOURCE_KIND,
    'large_medical_personal': SOCIAL_SOURCE_KIND,
    'late_fee': SOCIAL_SOURCE_KIND,
    'interest': SOCIAL_SOURCE_KIND,
}

REGION_LABELS = {
    'guangzhou': '\u5e7f\u5dde',
    'hangzhou': '\u676d\u5dde',
    'xiamen': '\u53a6\u95e8',
    'shenzhen': '\u6df1\u5733',
    'wuhan': '\u6b66\u6c49',
    'changsha': '\u957f\u6c99',
}


class ExportServiceError(Exception):
    pass


@dataclass
class ExportArtifactResult:
    template_type: str
    status: str
    file_path: Optional[str]
    error_message: Optional[str]
    row_count: int = 0


@dataclass
class DualTemplateExportResult:
    status: str
    artifacts: list[ExportArtifactResult]


# ---------------------------------------------------------------------------
# Pure utility functions
# ---------------------------------------------------------------------------

def _amount(value: Optional[Decimal]) -> Decimal:
    return value if value is not None else Decimal('0')


def _normalize_export_text(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_export_header(value: Optional[str]) -> Optional[str]:
    text = _normalize_export_text(value)
    if text is None:
        return None
    return text.translate(HEADER_NORMALIZATION_TRANSLATION).replace(' ', '').lower()


def _normalize_id_number(value: Optional[str]) -> Optional[str]:
    text = _normalize_export_text(value)
    if text is None:
        return None
    compact = text.replace(' ', '').upper()
    if compact in HEADER_LIKE_PERSON_VALUES:
        return None
    if ID_NUMBER_PATTERN.fullmatch(compact):
        return compact
    if NON_MAINLAND_ID_NUMBER_PATTERN.fullmatch(compact):
        return compact
    return None


def _decimal_from_raw_value(value: object) -> Optional[Decimal]:
    if value in (None, ''):
        return None
    text = str(value).strip().replace(',', '')
    if not text:
        return None
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _region_label(value: Optional[str]) -> str:
    if value is None:
        return ''
    return REGION_LABELS.get(value, value)


def _is_exportable_record(record: NormalizedRecord) -> bool:
    person_name = _normalize_export_text(record.person_name)
    employee_id = _normalize_export_text(record.employee_id)
    id_number = _normalize_id_number(record.id_number)
    if person_name is None or person_name in HEADER_LIKE_PERSON_VALUES:
        return False
    if id_number is None and employee_id is None:
        return False
    if not _has_any_business_value(record):
        return False
    if _is_inferred_housing_only_record(record):
        return False
    return True


def _has_any_business_value(record: NormalizedRecord) -> bool:
    return any(_amount(getattr(record, field_name)) != Decimal('0') for field_name in EXPORT_AMOUNT_FIELDS)


def _is_inferred_housing_only_record(record: NormalizedRecord) -> bool:
    personal_housing, company_housing, housing_total = _resolved_housing_fund_values(record)
    has_housing = any(amount != Decimal('0') for amount in (personal_housing, company_housing, housing_total))
    if not has_housing:
        return False
    has_social = any(_amount(getattr(record, field_name, None)) != Decimal('0') for field_name in SOCIAL_AMOUNT_FIELDS)
    if has_social:
        return False
    return not _has_explicit_housing_breakdown(record)


def _has_explicit_housing_breakdown(record: NormalizedRecord) -> bool:
    raw_payload = record.raw_payload or {}
    merged_sources = raw_payload.get('merged_sources')
    if not isinstance(merged_sources, list):
        return False

    for source in merged_sources:
        if not isinstance(source, dict) or source.get('source_kind') != HOUSING_SOURCE_KIND:
            continue
        raw_values = source.get('raw_values')
        if not isinstance(raw_values, dict):
            continue
        has_company = _has_explicit_housing_amount_key(raw_values, EXPLICIT_HOUSING_COMPANY_HEADER_TOKENS)
        has_personal = _has_explicit_housing_amount_key(raw_values, EXPLICIT_HOUSING_PERSONAL_HEADER_TOKENS)
        has_combined = _has_explicit_housing_amount_key(raw_values, EXPLICIT_HOUSING_COMBINED_HEADER_TOKENS)
        if (has_company and has_personal) or has_combined:
            return True
    return False


def _has_explicit_housing_amount_key(raw_values: dict[object, object], header_tokens: tuple[str, ...]) -> bool:
    normalized_tokens = tuple(
        normalized
        for normalized in (_normalize_export_header(token) for token in header_tokens)
        if normalized is not None
    )
    for raw_key, raw_value in raw_values.items():
        normalized_header = _normalize_export_header(str(raw_key) if raw_key is not None else None)
        if normalized_header is None:
            continue
        if raw_value in (None, ''):
            continue
        header_candidates = {normalized_header, *filter(None, re.split(r'[:/\[\]()]+', normalized_header))}
        if any(
            candidate == token or candidate.endswith(token)
            for candidate in header_candidates
            for token in normalized_tokens
        ):
            return True
        if any(marker in normalized_header for marker in HOUSING_METADATA_HEADER_MARKERS):
            continue
    return False


# ---------------------------------------------------------------------------
# Merge functions
# ---------------------------------------------------------------------------

def _merge_export_records(records: list[NormalizedRecord]) -> list[NormalizedRecord]:
    merged_records: list[NormalizedRecord] = []
    index_by_key: dict[tuple[str, str], int] = {}

    for record in records:
        merge_key = _export_merge_key(record)
        if merge_key is None:
            merged_records.append(record)
            continue
        existing_index = index_by_key.get(merge_key)
        if existing_index is None:
            merged_records.append(_copy_export_record(record))
            index_by_key[merge_key] = len(merged_records) - 1
            continue
        merged_records[existing_index] = _merge_two_records(merged_records[existing_index], record)

    return merged_records


def _export_merge_key(record: NormalizedRecord) -> tuple[str, str] | None:
    employee_id = _normalize_export_text(record.employee_id)
    if employee_id:
        return ('employee_id', employee_id)
    id_number = _normalize_id_number(record.id_number)
    if id_number:
        return ('id_number', id_number)
    return None


def _copy_export_record(record: NormalizedRecord) -> NormalizedRecord:
    clone = NormalizedRecord()
    for field_name in EXPORT_TEXT_FIELDS:
        setattr(clone, field_name, getattr(record, field_name, None))
    for field_name in EXPORT_AMOUNT_FIELDS:
        setattr(clone, field_name, getattr(record, field_name, None))
    clone.raw_payload = deepcopy(record.raw_payload) if record.raw_payload is not None else None
    return clone


def _merge_two_records(base: NormalizedRecord, incoming: NormalizedRecord) -> NormalizedRecord:
    for field_name in EXPORT_TEXT_FIELDS:
        setattr(
            base,
            field_name,
            _merge_text_value(
                getattr(base, field_name, None),
                getattr(incoming, field_name, None),
                prefer_longer=field_name in {'company_name', 'raw_sheet_name', 'source_file_name'},
                prefer_valid_id=field_name == 'id_number',
            ),
        )
    for field_name in EXPORT_AMOUNT_FIELDS:
        setattr(
            base,
            field_name,
            _merge_amount_value(
                getattr(base, field_name, None),
                getattr(incoming, field_name, None),
                field_name=field_name,
                current_payload=base.raw_payload,
                incoming_payload=incoming.raw_payload,
            ),
        )
    base.raw_payload = _merge_raw_payloads(base.raw_payload, incoming.raw_payload)
    return base


def _merge_raw_payloads(
    current: dict[str, object] | None,
    incoming: dict[str, object] | None,
) -> dict[str, object] | None:
    if current is None:
        return deepcopy(incoming) if incoming is not None else None
    if incoming is None:
        return current

    merged = deepcopy(current)
    merged_sources = merged.setdefault('merged_sources', [])
    if not isinstance(merged_sources, list):
        merged_sources = []
        merged['merged_sources'] = merged_sources

    existing_keys = {
        (
            item.get('source_kind'),
            item.get('source_file_name'),
            item.get('source_row_number'),
        )
        for item in merged_sources
        if isinstance(item, dict)
    }
    for item in incoming.get('merged_sources', []):
        if not isinstance(item, dict):
            continue
        item_key = (
            item.get('source_kind'),
            item.get('source_file_name'),
            item.get('source_row_number'),
        )
        if item_key in existing_keys:
            continue
        merged_sources.append(deepcopy(item))
        existing_keys.add(item_key)
    return merged


def _merge_text_value(
    current: Optional[str],
    incoming: Optional[str],
    *,
    prefer_longer: bool = False,
    prefer_valid_id: bool = False,
) -> Optional[str]:
    normalized_current = _normalize_export_text(current)
    normalized_incoming = _normalize_export_text(incoming)
    if normalized_current is None:
        return normalized_incoming
    if normalized_incoming is None:
        return normalized_current
    if normalized_current == normalized_incoming:
        return normalized_current
    if prefer_valid_id:
        current_id = _normalize_id_number(normalized_current)
        incoming_id = _normalize_id_number(normalized_incoming)
        if current_id is None:
            return normalized_incoming
        if incoming_id is None:
            return normalized_current
    if normalized_current in HEADER_LIKE_PERSON_VALUES:
        return normalized_incoming
    if normalized_incoming in HEADER_LIKE_PERSON_VALUES:
        return normalized_current
    if prefer_longer and len(normalized_incoming) > len(normalized_current):
        return normalized_incoming
    return normalized_current


def _merge_amount_value(
    current: Optional[Decimal],
    incoming: Optional[Decimal],
    *,
    field_name: str,
    current_payload: dict[str, object] | None,
    incoming_payload: dict[str, object] | None,
) -> Optional[Decimal]:
    current_amount = _amount(current)
    incoming_amount = _amount(incoming)
    if current is None or current_amount == Decimal('0'):
        return incoming
    if incoming is None or incoming_amount == Decimal('0'):
        return current
    if _should_accumulate_amount(
        field_name=field_name,
        current_payload=current_payload,
        incoming_payload=incoming_payload,
    ):
        return current_amount + incoming_amount
    if current_amount == incoming_amount:
        return current
    return current if abs(current_amount) >= abs(incoming_amount) else incoming


def _should_accumulate_amount(
    *,
    field_name: str,
    current_payload: dict[str, object] | None,
    incoming_payload: dict[str, object] | None,
) -> bool:
    source_kind = SOURCE_KIND_BY_AMOUNT_FIELD.get(field_name)
    if source_kind is None:
        return False
    current_signatures = _source_signatures_for_kind(current_payload, source_kind)
    incoming_signatures = _source_signatures_for_kind(incoming_payload, source_kind)
    if not current_signatures or not incoming_signatures:
        return False
    return current_signatures != incoming_signatures


def _source_signatures_for_kind(
    raw_payload: dict[str, object] | None,
    source_kind: str,
) -> Optional[set[tuple[str], int | None]]:
    if not isinstance(raw_payload, dict):
        return set()
    merged_sources = raw_payload.get('merged_sources')
    if not isinstance(merged_sources, list):
        return set()

    signatures: Optional[set[tuple[str], int | None]] = set()
    for item in merged_sources:
        if not isinstance(item, dict):
            continue
        if item.get('source_kind') != source_kind:
            continue
        source_file_name = _normalize_export_text(item.get('source_file_name'))
        source_row_number = item.get('source_row_number')
        if isinstance(source_row_number, str) and source_row_number.isdigit():
            source_row_number = int(source_row_number)
        signatures.add((source_file_name, source_row_number if isinstance(source_row_number, int) else None))
    return signatures


def _source_signature_count_for_kind(
    raw_payload: dict[str, object] | None,
    source_kind: str,
) -> int:
    return len(_source_signatures_for_kind(raw_payload, source_kind))


# ---------------------------------------------------------------------------
# Sheet rewrite functions
# ---------------------------------------------------------------------------

def _rewrite_sheet_in_place(
    sheet: Worksheet,
    *,
    template_row: int,
    records: list[NormalizedRecord],
    value_builder,
) -> None:
    row_values = [value_builder(record) for record in records]
    template_snapshot = _snapshot_row(sheet, template_row)
    existing_last_row = _detect_existing_data_last_row(sheet, template_row, probe_columns=len(template_snapshot['cells']))
    target_last_row = max(existing_last_row, template_row + len(records) - 1)

    for target_row in range(template_row, target_last_row + 1):
        _apply_row_snapshot(sheet, template_snapshot, source_row=template_row, target_row=target_row)
        values = row_values[target_row - template_row] if target_row - template_row < len(row_values) else None
        _populate_output_row(sheet, template_snapshot, target_row=target_row, values=values)


def _detect_existing_data_last_row(sheet: Worksheet, start_row: int, *, probe_columns: int) -> int:
    last_populated = start_row
    consecutive_empty_rows = 0
    upper_bound = min(sheet.max_row, start_row + 2000)

    for row_number in range(start_row, upper_bound + 1):
        has_value = False
        for column_index in range(1, probe_columns + 1):
            cell = sheet.cell(row=row_number, column=column_index)
            if isinstance(cell, MergedCell):
                continue
            if cell.value not in (None, ''):
                has_value = True
                break
        if has_value:
            last_populated = row_number
            consecutive_empty_rows = 0
        else:
            consecutive_empty_rows += 1
            if consecutive_empty_rows >= 25:
                break

    return last_populated


def _snapshot_row(sheet: Worksheet, row_number: int) -> dict[str, object]:
    row_dimension = copy(sheet.row_dimensions[row_number]) if row_number in sheet.row_dimensions else None
    cells = []
    for column_index in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=row_number, column=column_index)
        cells.append(
            {
                'column_index': column_index,
                'value': cell.value,
                'style': copy(cell._style) if cell.has_style else None,
                'font': copy(cell.font),
                'fill': copy(cell.fill),
                'border': copy(cell.border),
                'alignment': copy(cell.alignment),
                'number_format': cell.number_format,
                'protection': copy(cell.protection),
            }
        )
    return {'row_dimension': row_dimension, 'cells': cells}


def _apply_row_snapshot(
    sheet: Worksheet,
    snapshot: dict[str, object],
    *,
    source_row: int,
    target_row: int,
) -> None:
    row_dimension = snapshot['row_dimension']
    if row_dimension is not None:
        sheet.row_dimensions[target_row] = copy(row_dimension)

    for cell_snapshot in snapshot['cells']:
        column_index = cell_snapshot['column_index']
        target_cell = sheet.cell(row=target_row, column=column_index)
        if isinstance(target_cell, MergedCell):
            continue
        value = cell_snapshot['value']
        if isinstance(value, str) and value.startswith('=') and target_row != source_row:
            origin = f"{get_column_letter(column_index)}{source_row}"
            destination = f"{get_column_letter(column_index)}{target_row}"
            value = Translator(value, origin=origin).translate_formula(destination)
        target_cell.value = value
        target_cell.font = copy(cell_snapshot['font'])
        target_cell.fill = copy(cell_snapshot['fill'])
        target_cell.border = copy(cell_snapshot['border'])
        target_cell.alignment = copy(cell_snapshot['alignment'])
        target_cell.number_format = cell_snapshot['number_format']
        target_cell.protection = copy(cell_snapshot['protection'])
        if cell_snapshot['style'] is not None:
            target_cell._style = copy(cell_snapshot['style'])


def _populate_output_row(
    sheet: Worksheet,
    snapshot: dict[str, object],
    *,
    target_row: int,
    values: Optional[list[object]],
) -> None:
    cell_snapshots = snapshot['cells']
    for cell_snapshot in cell_snapshots:
        column_index = cell_snapshot['column_index']
        template_value = cell_snapshot['value']
        is_formula = isinstance(template_value, str) and template_value.startswith('=')
        uses_external_reference = is_formula and '[' in template_value
        target_cell = sheet.cell(row=target_row, column=column_index)
        if isinstance(target_cell, MergedCell):
            continue

        if values is None:
            if not (is_formula and not uses_external_reference):
                target_cell.value = None
            continue

        if column_index > len(values):
            continue
        if is_formula and not uses_external_reference:
            continue
        target_cell.value = values[column_index - 1]


# ---------------------------------------------------------------------------
# Housing / burden functions
# ---------------------------------------------------------------------------

def _resolved_housing_fund_values(record: NormalizedRecord) -> tuple[Decimal, Decimal, Decimal]:
    personal = _amount(record.housing_fund_personal)
    company = _amount(record.housing_fund_company)
    total = _amount(record.housing_fund_total)
    quant = Decimal('0.01')

    if total != 0 and personal > 0 and personal <= Decimal('1') and company > Decimal('1'):
        personal = (total - company).quantize(quant, rounding=ROUND_HALF_UP)
        if personal <= 0:
            personal = company
    if total != 0 and company > 0 and company <= Decimal('1') and personal > Decimal('1'):
        company = (total - personal).quantize(quant, rounding=ROUND_HALF_UP)
        if company <= 0:
            company = personal

    if total == 0:
        total = personal + company
    if personal == 0 and company == 0 and total != 0:
        personal = (total / Decimal('2')).quantize(quant, rounding=ROUND_HALF_UP)
        company = (total - personal).quantize(quant, rounding=ROUND_HALF_UP)
    elif personal == 0 and total != 0 and company != 0:
        personal = (total - company).quantize(quant, rounding=ROUND_HALF_UP)
    elif company == 0 and total != 0 and personal != 0:
        company = (total - personal).quantize(quant, rounding=ROUND_HALF_UP)

    if total == 0:
        total = personal + company
    return personal, company, total


def _build_social_burden_context(records: list[NormalizedRecord]) -> dict[str, tuple[Decimal, ...]]:
    amount_counts_by_source: dict[str, dict[Decimal, int]] = {}

    for record in records:
        source_key = _social_burden_source_key(record)
        if source_key is None:
            continue
        company_medical = _amount(record.medical_maternity_company or record.medical_company)
        if company_medical <= 0:
            continue
        amount_counts = amount_counts_by_source.setdefault(source_key, {})
        amount_counts[company_medical] = amount_counts.get(company_medical, 0) + 1

    context: dict[str, tuple[Decimal, ...]] = {}
    for source_key, amount_counts in amount_counts_by_source.items():
        repeated_amounts = sorted(amount for amount, count in amount_counts.items() if count >= 2)
        if not repeated_amounts:
            continue
        baseline = repeated_amounts[0]
        candidates = tuple(
            amount for amount in repeated_amounts if amount <= baseline * SOCIAL_BURDEN_CANDIDATE_RATIO
        )
        if candidates:
            context[source_key] = candidates
    return context


def _build_housing_burden_context(records: list[NormalizedRecord]) -> dict[str, tuple[Decimal, ...]]:
    amount_counts_by_source: dict[str, dict[Decimal, int]] = {}

    for record in records:
        source_key = _housing_burden_source_key(record)
        if source_key is None:
            continue
        personal_housing, company_housing, _housing_total = _resolved_housing_fund_values(record)
        if personal_housing <= 0 or company_housing <= 0 or personal_housing != company_housing:
            continue
        amount_counts = amount_counts_by_source.setdefault(source_key, {})
        amount_counts[company_housing] = amount_counts.get(company_housing, 0) + 1

    context: dict[str, tuple[Decimal, ...]] = {}
    for source_key, amount_counts in amount_counts_by_source.items():
        repeated_amounts = sorted(amount for amount, count in amount_counts.items() if count >= 2)
        if not repeated_amounts:
            continue
        baseline = repeated_amounts[0]
        candidates = tuple(
            amount for amount in repeated_amounts if amount <= baseline * HOUSING_BURDEN_CANDIDATE_RATIO
        )
        if candidates:
            context[source_key] = candidates
    return context


def _resolved_personal_social_burden(
    record: NormalizedRecord,
    *,
    company_medical: Decimal,
    social_burden_context: dict[str, tuple[Decimal, ...]],
) -> Decimal:
    return Decimal('0')


def _resolved_personal_housing_burden(
    record: NormalizedRecord,
    *,
    company_housing: Decimal,
    housing_burden_context: dict[str, tuple[Decimal, ...]],
) -> Decimal:
    return Decimal('0')


def _select_burden_allowance(
    current_amount: Decimal,
    candidates: tuple[Decimal, ...],
) -> Optional[Decimal]:
    if not candidates:
        return None
    eligible = [candidate for candidate in candidates if candidate <= current_amount]
    if eligible:
        return eligible[-1]
    return candidates[0]


def _social_burden_source_key(record: NormalizedRecord) -> Optional[str]:
    return _source_key_for_kind(record, SOCIAL_SOURCE_KIND, ('medical_company', 'medical_maternity_company'))


def _housing_burden_source_key(record: NormalizedRecord) -> Optional[str]:
    return _source_key_for_kind(record, HOUSING_SOURCE_KIND, ('housing_fund_personal', 'housing_fund_company', 'housing_fund_total'))


def _source_key_for_kind(
    record: NormalizedRecord,
    source_kind: str,
    amount_fields: tuple[str, ...],
) -> Optional[str]:
    raw_payload = record.raw_payload or {}
    merged_sources = raw_payload.get('merged_sources')
    if isinstance(merged_sources, list):
        for source in merged_sources:
            if not isinstance(source, dict):
                continue
            if source.get('source_kind') != source_kind:
                continue
            source_file_name = _normalize_export_text(source.get('source_file_name'))
            if source_file_name:
                return source_file_name
    if any(_amount(getattr(record, field_name, None)) > 0 for field_name in amount_fields):
        return _normalize_export_text(record.source_file_name)
    return None


# ---------------------------------------------------------------------------
# Resolved amount helpers (shared by salary and tool)
# ---------------------------------------------------------------------------

def _resolved_personal_large_medical(record: NormalizedRecord) -> Decimal:
    amount = _amount(record.large_medical_personal)
    if amount <= 0:
        return amount
    return amount


def _resolved_company_medical(record: NormalizedRecord) -> Decimal:
    if record.region == 'xiamen' and _source_signature_count_for_kind(record.raw_payload, SOCIAL_SOURCE_KIND) > 1:
        medical_amounts = _extract_xiamen_social_amounts(record.raw_payload, target='medical')
        maternity_amounts = _extract_xiamen_social_amounts(record.raw_payload, target='maternity')
        if medical_amounts or maternity_amounts:
            return max(medical_amounts, default=_amount(record.medical_company)) + max(
                maternity_amounts,
                default=_amount(record.maternity_amount),
            )

    amount = _amount(record.medical_maternity_company or record.medical_company)
    if record.region == 'xiamen' and _source_signature_count_for_kind(record.raw_payload, SOCIAL_SOURCE_KIND) > 1:
        amount += _amount(record.maternity_amount)
    return amount


def _resolved_company_large_medical(record: NormalizedRecord) -> Decimal:
    if record.region == 'wuhan':
        return _amount(record.supplementary_medical_company)
    return _amount(record.supplementary_medical_company)


# ---------------------------------------------------------------------------
# Xiamen extraction
# ---------------------------------------------------------------------------

def _extract_xiamen_social_amounts(
    raw_payload: dict[str, object] | None,
    *,
    target: str,
) -> list[Decimal]:
    if not isinstance(raw_payload, dict):
        return []
    merged_sources = raw_payload.get('merged_sources')
    if not isinstance(merged_sources, list):
        return []

    amounts: list[Decimal] = []
    for source in merged_sources:
        if not isinstance(source, dict) or source.get('source_kind') != SOCIAL_SOURCE_KIND:
            continue
        raw_values = source.get('raw_values')
        if not isinstance(raw_values, dict):
            continue
        amount = _extract_xiamen_source_amount(raw_values, target=target)
        if amount is not None and amount > 0:
            amounts.append(amount)
    return amounts


def _extract_xiamen_source_amount(
    raw_values: dict[object, object],
    *,
    target: str,
) -> Optional[Decimal]:
    for raw_key, raw_value in raw_values.items():
        normalized_header = _normalize_export_header(str(raw_key) if raw_key is not None else None)
        if normalized_header is None:
            continue
        if target == 'medical':
            if XIAMEN_MATERNITY_HEADER[0] in normalized_header:
                continue
            if XIAMEN_MEDICAL_COMPANY_HEADER[0] in normalized_header and XIAMEN_MEDICAL_COMPANY_HEADER[1] in normalized_header:
                return _decimal_from_raw_value(raw_value)
            continue
        if target == 'maternity':
            if XIAMEN_MATERNITY_HEADER[0] in normalized_header and XIAMEN_MATERNITY_HEADER[1] in normalized_header:
                return _decimal_from_raw_value(raw_value)
    return None


# ---------------------------------------------------------------------------
# Template resolution
# ---------------------------------------------------------------------------

def _resolve_template_path(
    template_path: str | Path | None,
    template_type: TemplateType,
    *,
    settings: Optional[Settings] = None,
) -> Path:
    attempted_locations: list[str] = []
    if template_path:
        candidate = Path(template_path)
        if candidate.exists():
            return candidate
        attempted_locations.append(str(candidate))
        attempted_message = f" Attempted: {', '.join(attempted_locations)}."
        raise ExportServiceError(
            f'Explicit template path for {template_type.value} does not exist.{attempted_message}'
        )

    resolved_settings = settings or get_settings()
    configured = (
        resolved_settings.salary_template_file
        if template_type == TemplateType.SALARY
        else resolved_settings.final_tool_template_file
    )
    if configured and configured.exists():
        return configured
    if configured is not None:
        attempted_locations.append(str(configured))

    discovered = _discover_template_candidates(resolved_settings, template_type)
    if discovered:
        return discovered[0]

    attempted_message = f" Attempted: {', '.join(attempted_locations)}." if attempted_locations else ''
    raise ExportServiceError(f'No template could be resolved for {template_type.value}.{attempted_message}')


def _discover_template_candidates(settings, template_type: TemplateType) -> list[Path]:
    pattern = '*薪酬*.xlsx' if template_type == TemplateType.SALARY else '*最终版*.xlsx'
    candidates: dict[Path, tuple[float, str]] = {}

    root = settings.templates_path
    if not root.exists():
        return []

    for manifest_path in root.rglob('manifest.json'):
        for manifest_match in _discover_template_candidates_from_manifest(manifest_path, template_type):
            try:
                stat = manifest_match.stat()
                candidates[manifest_match.resolve()] = (stat.st_mtime, manifest_match.name)
            except OSError:
                continue

    for match in root.rglob(pattern):
        if not match.is_file():
            continue
        try:
            stat = match.stat()
            candidates[match.resolve()] = (stat.st_mtime, match.name)
        except OSError:
            continue

    return [
        path
        for path, _ in sorted(
            candidates.items(),
            key=lambda item: (item[1][0], item[1][1]),
            reverse=True,
        )
    ]


def _discover_template_candidates_from_manifest(
    manifest_path: Path,
    template_type: TemplateType,
) -> list[Path]:
    try:
        payload = json.loads(manifest_path.read_text(encoding='utf-8'))
    except (OSError, json.JSONDecodeError):
        return []
    if not isinstance(payload, dict):
        return []

    manifest_key = 'salary' if template_type == TemplateType.SALARY else 'final_tool'
    relative_path = payload.get(manifest_key)
    if not isinstance(relative_path, str) or not relative_path:
        return []

    candidate = (manifest_path.parent / relative_path).resolve()
    if not candidate.is_file():
        return []
    return [candidate]
