from __future__ import annotations

from decimal import Decimal

from openpyxl import Workbook

from backend.app.models.normalized_record import NormalizedRecord

from backend.app.exporters.export_utils import (
    _amount,
    _resolved_company_large_medical,
    _resolved_company_medical,
    _resolved_housing_fund_values,
    _resolved_personal_large_medical,
    _rewrite_sheet_in_place,
)

SALARY_SHEET_HEADER_ROW = 1
SALARY_DATA_START_ROW = 2

SALARY_HEADERS = [
    '\u5458\u5de5\u59d3\u540d', '\u5de5\u53f7', '\u4e2a\u4eba\u533b\u7597\u4fdd\u9669', '\u4e2a\u4eba\u5931\u4e1a\u4fdd\u9669', '\u4e2a\u4eba\u5927\u75c5\u533b\u7597', '\u4e2a\u4eba\u6b8b\u75be\u4fdd\u969c\u91d1', '\u4e2a\u4eba\u517b\u8001\u4fdd\u9669', '\u4e2a\u4eba\u516c\u79ef\u91d1',
    '\u516c\u53f8\u517b\u8001\u4fdd\u9669', '\u516c\u53f8\u533b\u7597\u4fdd\u9669', '\u516c\u53f8\u5931\u4e1a\u4fdd\u9669', '\u516c\u53f8\u5de5\u4f24\u4fdd\u9669', '\u516c\u53f8\u751f\u80b2\u4fdd\u9669', '\u516c\u53f8\u5927\u75c5\u533b\u7597', '\u516c\u53f8\u6b8b\u75be\u4fdd\u969c\u91d1', '\u516c\u53f8\u516c\u79ef\u91d1',
]


def _rewrite_salary_sheet(workbook: Workbook, records: list[NormalizedRecord]) -> None:
    sheet = workbook[workbook.sheetnames[0]]
    _strip_salary_burden_columns(sheet)
    _rewrite_sheet_in_place(
        sheet,
        template_row=SALARY_DATA_START_ROW,
        records=records,
        value_builder=_salary_row_values,
    )


def _salary_row_values(record: NormalizedRecord) -> list[object]:
    personal_medical = _amount(record.medical_personal)
    personal_unemployment = _amount(record.unemployment_personal)
    personal_large_medical = _resolved_personal_large_medical(record)
    personal_pension = _amount(record.pension_personal)
    company_pension = _amount(record.pension_company) + _amount(record.supplementary_pension_company)
    company_medical = _resolved_company_medical(record)
    company_unemployment = _amount(record.unemployment_company)
    company_injury = _amount(record.injury_company)
    company_maternity = _amount(record.maternity_amount)
    company_large_medical = _resolved_company_large_medical(record)
    personal_housing, company_housing, _housing_total = _resolved_housing_fund_values(record)

    return [
        record.person_name or '',
        record.employee_id or '',
        personal_medical,
        personal_unemployment,
        personal_large_medical,
        Decimal('0'),
        personal_pension,
        personal_housing,
        company_pension,
        company_medical,
        company_unemployment,
        company_injury,
        company_maternity,
        company_large_medical,
        Decimal('0'),
        company_housing,
    ]


def _strip_salary_burden_columns(sheet) -> None:
    if (
        sheet.max_column >= 18
        and sheet.cell(row=SALARY_SHEET_HEADER_ROW, column=17).value == '个人社保承担额'
        and sheet.cell(row=SALARY_SHEET_HEADER_ROW, column=18).value == '个人公积金承担额'
    ):
        sheet.delete_cols(17, 2)
