from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl.worksheet.worksheet import Worksheet

from backend.app.parsers.workbook_loader import load_workbook_compatible

HEADER_KEYWORDS = {
    "姓名",
    "证件",
    "工号",
    "社保",
    "保险",
    "费款",
    "所属期",
    "单位",
    "个人",
    "合计",
    "金额",
    "缴费",
    "基数",
    "工资",
}
PIVOT_HINT_KEYWORDS = {"求和项", "总计", "(空白)"}
TRANSACTIONAL_HINT_KEYWORDS = {"征收项目", "征收品目", "费率", "人员编号"}
GROUP_ROW_TOKENS = {"合计", "小计", "在职人员", "退休人员", "家属统筹人员"}
MAX_SCAN_ROWS = 30
MAX_SCAN_COLUMNS = 30
DETAIL_HEADER_KEYWORDS = {
    "\u59d3\u540d",
    "\u8eab\u4efd\u8bc1",
    "\u8bc1\u4ef6\u53f7",
    "\u5ba2\u6237\u540d\u79f0",
    "\u516c\u53f8\u5168\u79f0",
    "\u53c2\u4fdd\u5730",
    "\u5e10\u5355\u5e74\u6708",
    "\u8d39\u7528\u6240\u5c5e\u6708",
    "\u7f34\u7eb3\u6708",
    "\u5355\u4f4d\u7f34\u7eb3",
    "\u4e2a\u4eba\u7f34\u7eb3",
}
SUMMARY_HEADER_KEYWORDS = {
    "\u4ed8\u6b3e\u5355\u4f4d",
    "\u793e\u4fdd\u5c0f\u8ba1",
    "\u516c\u79ef\u91d1\u5c0f\u8ba1",
    "\u6b8b\u969c\u91d1\u5c0f\u8ba1",
    "\u670d\u52a1\u8d39",
    "\u5408\u4f5c\u4ea7\u54c1",
    "\u4ed8\u6b3e\u901a\u77e5\u4e66",
}
DETAIL_SHEET_TOKENS = ("\u660e\u7ec6",)
SUMMARY_SHEET_TOKENS = ("\u901a\u77e5\u4e66", "\u6c47\u603b", "\u5bf9\u8d26")


@dataclass(slots=True)
class SheetDiscovery:
    sheet_name: str
    score: int
    header_row_candidates: list[int]
    data_start_row: int | None
    non_empty_rows: int
    preview_rows: list[list[str]]
    reasoning: list[str]


@dataclass(slots=True)
class WorkbookDiscovery:
    source_file: str
    sheet_names: list[str]
    selected_sheet_name: str | None
    selected_data_start_row: int | None
    selected_header_row_candidates: list[int]
    discoveries: list[SheetDiscovery]
    failure_reason: str | None = None


class WorkbookDiscoveryError(Exception):
    pass


def discover_workbook(path: str | Path) -> WorkbookDiscovery:
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise WorkbookDiscoveryError(f"Workbook '{workbook_path}' does not exist.")

    workbook = load_workbook_compatible(workbook_path, read_only=True, data_only=True)
    try:
        discoveries = [_discover_sheet(sheet) for sheet in workbook.worksheets]
    finally:
        workbook.close()

    selected = max(discoveries, key=lambda item: item.score, default=None)
    if selected is None or selected.score <= 0:
        return WorkbookDiscovery(
            source_file=workbook_path.name,
            sheet_names=[item.sheet_name for item in discoveries],
            selected_sheet_name=None,
            selected_data_start_row=None,
            selected_header_row_candidates=[],
            discoveries=discoveries,
            failure_reason="No valid worksheet candidate was detected.",
        )

    return WorkbookDiscovery(
        source_file=workbook_path.name,
        sheet_names=[item.sheet_name for item in discoveries],
        selected_sheet_name=selected.sheet_name,
        selected_data_start_row=selected.data_start_row,
        selected_header_row_candidates=selected.header_row_candidates,
        discoveries=sorted(discoveries, key=lambda item: item.score, reverse=True),
    )


def _discover_sheet(sheet: Worksheet) -> SheetDiscovery:
    scanned_rows = _scan_rows(sheet)
    non_empty_rows = sum(1 for row in scanned_rows if row)
    if not scanned_rows:
        return SheetDiscovery(
            sheet_name=sheet.title,
            score=0,
            header_row_candidates=[],
            data_start_row=None,
            non_empty_rows=0,
            preview_rows=[],
            reasoning=["Sheet is empty in the scanned window."],
        )

    row_profiles = [_profile_row(row) for row in scanned_rows]
    header_row_candidates = _select_header_candidates(scanned_rows, row_profiles)
    data_start_row = _detect_data_start_row(scanned_rows, header_row_candidates)

    score = 0
    reasoning: list[str] = []

    title_score = _calculate_sheet_title_score(sheet.title)
    if title_score:
        score += title_score
        reasoning.append(f"Detected title hint from sheet name '{sheet.title}' ({title_score:+d}).")

    header_score = _calculate_header_score(row_profiles, header_row_candidates)
    if header_score:
        score += header_score
        reasoning.append(f"Detected header row candidates at {header_row_candidates} (+{header_score}).")

    pivot_hint_bonus = _calculate_pivot_hint_bonus(scanned_rows)
    if pivot_hint_bonus:
        score += pivot_hint_bonus
        reasoning.append(f"Detected pivot-style hint rows (+{pivot_hint_bonus}).")

    if data_start_row is not None:
        score += 24
        reasoning.append(f"Detected data start row at {data_start_row} (+24).")

    summary_penalty = _calculate_summary_penalty(scanned_rows, row_profiles, sheet.title)
    if summary_penalty:
        score -= summary_penalty
        reasoning.append(f"Detected summary-style sheet markers (-{summary_penalty}).")

    transactional_penalty = _calculate_transactional_penalty(scanned_rows, header_row_candidates)
    if transactional_penalty:
        score -= transactional_penalty
        reasoning.append(f"Detected transactional detail-sheet markers (-{transactional_penalty}).")

    if non_empty_rows >= 4:
        density_bonus = min(non_empty_rows, 10)
        score += density_bonus
        reasoning.append(f"Found {non_empty_rows} non-empty rows in the scan window (+{density_bonus}).")

    return SheetDiscovery(
        sheet_name=sheet.title,
        score=score,
        header_row_candidates=header_row_candidates,
        data_start_row=data_start_row,
        non_empty_rows=non_empty_rows,
        preview_rows=scanned_rows[:6],
        reasoning=reasoning or ["No spreadsheet-like structure detected."],
    )


def _scan_rows(sheet: Worksheet) -> list[list[str]]:
    rows: list[list[str]] = []
    for row in sheet.iter_rows(min_row=1, max_row=MAX_SCAN_ROWS, max_col=MAX_SCAN_COLUMNS, values_only=True):
        normalized = [_normalize_cell(value) for value in row]
        trimmed = [value for value in normalized if value]
        rows.append(trimmed)
    return rows


def _normalize_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _profile_row(row: list[str]) -> dict[str, int]:
    if not row:
        return {"width": 0, "keywords": 0, "detail_keywords": 0, "summary_keywords": 0, "numeric_like": 0, "text_like": 0, "score": 0}

    keyword_hits = sum(1 for cell in row for keyword in HEADER_KEYWORDS if keyword in cell)
    detail_keyword_hits = sum(1 for cell in row for keyword in DETAIL_HEADER_KEYWORDS if keyword in cell)
    summary_keyword_hits = sum(1 for cell in row for keyword in SUMMARY_HEADER_KEYWORDS if keyword in cell)
    numeric_like_cells = sum(1 for cell in row if _looks_numeric(cell))
    text_like_cells = sum(1 for cell in row if not _looks_numeric(cell))
    score = keyword_hits * 12 + detail_keyword_hits * 18 + text_like_cells * 2 - numeric_like_cells
    if any(token in cell for cell in row for token in GROUP_ROW_TOKENS):
        score -= 10
    if summary_keyword_hits and detail_keyword_hits == 0:
        score -= summary_keyword_hits * 8
    return {
        "width": len(row),
        "keywords": keyword_hits,
        "detail_keywords": detail_keyword_hits,
        "summary_keywords": summary_keyword_hits,
        "numeric_like": numeric_like_cells,
        "text_like": text_like_cells,
        "score": score,
    }


def _select_header_candidates(rows: list[list[str]], row_profiles: list[dict[str, int]]) -> list[int]:
    ranked: list[tuple[int, int]] = []
    for index, profile in enumerate(row_profiles, start=1):
        row = rows[index - 1]
        if profile["width"] < 2:
            continue
        if profile["keywords"] == 0:
            continue
        if any(cell.strip() in GROUP_ROW_TOKENS for cell in row if cell.strip()):
            continue

        continuity_bonus = 0
        if index < len(row_profiles):
            next_profile = row_profiles[index]
            next_row = rows[index]
            if (
                next_profile["width"] >= max(2, profile["width"] // 2)
                and next_profile["keywords"] > 0
                and not _is_detail_like_row(next_row)
            ):
                continuity_bonus = 18

        rank_score = profile["score"] + profile["width"] * 3 + continuity_bonus
        ranked.append((index, rank_score))

    if not ranked:
        return []

    ranked.sort(key=lambda item: item[1], reverse=True)
    best_row = ranked[0][0]
    candidates = [best_row]
    if best_row < len(row_profiles):
        next_profile = row_profiles[best_row]
        next_row = rows[best_row]
        if (
            next_profile["width"] >= max(2, row_profiles[best_row - 1]["width"] // 2)
            and next_profile["keywords"] > 0
            and not _is_detail_like_row(next_row)
        ):
            candidates.append(best_row + 1)
    return candidates


def _calculate_header_score(row_profiles: list[dict[str, int]], header_row_candidates: list[int]) -> int:
    if not header_row_candidates:
        return 0

    score = 0
    for row_number in header_row_candidates:
        profile = row_profiles[row_number - 1]
        score += profile["keywords"] * 16
        score += profile["detail_keywords"] * 22
        score += profile["width"] * 4
        score += max(profile["text_like"] - profile["numeric_like"], 0)
        if profile["detail_keywords"] >= 2:
            score += 26

    if len(header_row_candidates) > 1:
        score += 18
    return score


def _detect_data_start_row(rows: list[list[str]], header_row_candidates: list[int]) -> int | None:
    if not header_row_candidates:
        return None

    start_index = header_row_candidates[-1]
    for row_number, row in enumerate(rows[start_index:], start=start_index + 1):
        if _is_detail_like_row(row):
            return row_number
    return None


def _is_detail_like_row(row: Iterable[str]) -> bool:
    values = [value for value in row if value]
    if len(values) < 3:
        return False
    if any(token == values[0] for token in GROUP_ROW_TOKENS):
        return False

    leading_values = values[:8]
    numeric_like = sum(1 for value in values if _looks_numeric(value))
    text_like = sum(1 for value in leading_values if value and not _looks_numeric(value))
    has_identifier = any(_looks_identifier(value) for value in leading_values)
    first_value = values[0]
    looks_like_named_detail = bool(first_value) and not _looks_numeric(first_value) and numeric_like >= 2 and len(values) >= 5
    return has_identifier or looks_like_named_detail or (text_like >= 2 and numeric_like >= 1 and len(values) >= 6)


def _calculate_pivot_hint_bonus(rows: list[list[str]]) -> int:
    bonus = 0
    for row in rows[:6]:
        for cell in row:
            if any(keyword in cell for keyword in PIVOT_HINT_KEYWORDS):
                bonus += 18
    return bonus



def _calculate_sheet_title_score(sheet_title: str) -> int:
    score = 0
    if any(token in sheet_title for token in DETAIL_SHEET_TOKENS):
        score += 30
    if any(token in sheet_title for token in SUMMARY_SHEET_TOKENS):
        score -= 24
    return score


def _calculate_summary_penalty(rows: list[list[str]], row_profiles: list[dict[str, int]], sheet_title: str) -> int:
    penalty = 0
    if any(token in sheet_title for token in SUMMARY_SHEET_TOKENS):
        penalty += 16

    top_rows = rows[:6]
    top_profiles = row_profiles[:6]
    summary_hits = sum(profile.get('summary_keywords', 0) for profile in top_profiles)
    detail_hits = sum(profile.get('detail_keywords', 0) for profile in top_profiles)
    if summary_hits >= 2 and detail_hits == 0:
        penalty += min(summary_hits * 10, 36)
    if any(any(keyword in cell for keyword in SUMMARY_HEADER_KEYWORDS) for row in top_rows for cell in row):
        penalty += 12
    return penalty


def _calculate_transactional_penalty(rows: list[list[str]], header_row_candidates: list[int]) -> int:
    if not header_row_candidates:
        return 0

    matched_keywords: set[str] = set()
    for row_number in header_row_candidates:
        for cell in rows[row_number - 1]:
            for keyword in TRANSACTIONAL_HINT_KEYWORDS:
                if keyword in cell:
                    matched_keywords.add(keyword)
    if len(matched_keywords) < 2:
        return 0
    return len(matched_keywords) * 42


def _looks_identifier(value: str) -> bool:
    candidate = value.replace(' ', '').replace('-', '').upper()
    if 15 <= len(candidate) <= 18 and candidate[:-1].isdigit() and (candidate[-1].isdigit() or candidate[-1] == 'X'):
        return True
    return False


def _looks_numeric(value: str) -> bool:
    candidate = value.replace(",", "").replace("%", "")
    if not candidate:
        return False
    try:
        float(candidate)
        return True
    except ValueError:
        return False
