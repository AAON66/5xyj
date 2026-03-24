from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from backend.app.parsers.workbook_discovery import WorkbookDiscoveryError, discover_workbook
from backend.app.parsers.workbook_loader import load_workbook_compatible


@dataclass(slots=True)
class HeaderColumn:
    column_index: int
    excel_column: str
    raw_header_parts: list[str]
    signature: str


@dataclass(slots=True)
class HeaderTreeNode:
    label: str
    row_number: int
    start_column: int
    end_column: int
    children: list["HeaderTreeNode"]


@dataclass(slots=True)
class HeaderExtraction:
    source_file: str
    sheet_name: str
    header_rows: list[int]
    data_start_row: int
    raw_header_signature: str
    columns: list[HeaderColumn]
    header_tree: list[HeaderTreeNode]


class HeaderExtractionError(Exception):
    pass


def extract_header_structure(path: str | Path) -> HeaderExtraction:
    workbook_path = Path(path)
    discovery = discover_workbook(workbook_path)
    if discovery.selected_sheet_name is None or discovery.selected_data_start_row is None:
        raise HeaderExtractionError(discovery.failure_reason or "Workbook discovery did not find a valid header region.")

    workbook = load_workbook_compatible(workbook_path, read_only=False, data_only=True)
    try:
        sheet = workbook[discovery.selected_sheet_name]
        header_rows = _infer_header_rows(sheet, discovery.selected_header_row_candidates, discovery.selected_data_start_row)
        columns = _extract_columns(sheet, header_rows, discovery.selected_data_start_row)
        if not columns:
            raise HeaderExtractionError("No header columns could be extracted from the selected header rows.")
        header_tree = _build_header_tree(sheet, header_rows, [column.column_index for column in columns])
        return HeaderExtraction(
            source_file=workbook_path.name,
            sheet_name=discovery.selected_sheet_name,
            header_rows=header_rows,
            data_start_row=discovery.selected_data_start_row,
            raw_header_signature=" | ".join(column.signature for column in columns),
            columns=columns,
            header_tree=header_tree,
        )
    finally:
        workbook.close()


def _infer_header_rows(sheet: Worksheet, header_candidates: list[int], data_start_row: int) -> list[int]:
    if not header_candidates:
        raise HeaderExtractionError("No header row candidates were provided.")

    if len(header_candidates) > 1:
        return header_candidates

    candidate_row = header_candidates[0]
    rows: list[int] = [candidate_row]
    previous_row = candidate_row - 1
    if previous_row >= 1 and _looks_like_header_row(sheet, previous_row, candidate_row, data_start_row):
        rows.insert(0, previous_row)
        return rows

    next_row = candidate_row + 1
    if next_row < data_start_row and _looks_like_header_row(sheet, next_row, candidate_row, data_start_row):
        rows.append(next_row)
    return rows


def _looks_like_header_row(sheet: Worksheet, row_number: int, anchor_row: int, data_start_row: int) -> bool:
    values = [_normalize(_effective_cell_value(sheet, row_number, column)) for column in range(1, sheet.max_column + 1)]
    anchor_values = [_normalize(_effective_cell_value(sheet, anchor_row, column)) for column in range(1, sheet.max_column + 1)]
    non_empty = [value for value in values if value]
    if len(non_empty) < 2:
        return False
    if row_number >= data_start_row:
        return False
    if _row_looks_like_data(non_empty):
        return False

    shared_columns = sum(
        1
        for current, anchor in zip(values, anchor_values, strict=False)
        if current and anchor and current != anchor
    )
    merged_bonus = sum(1 for merged_range in sheet.merged_cells.ranges if merged_range.min_row <= row_number <= merged_range.max_row)
    return shared_columns >= 1 or merged_bonus > 0


def _extract_columns(sheet: Worksheet, header_rows: list[int], data_start_row: int) -> list[HeaderColumn]:
    relevant_columns = [
        column_index
        for column_index in range(1, sheet.max_column + 1)
        if _column_has_signal(sheet, column_index, header_rows, data_start_row)
    ]

    columns: list[HeaderColumn] = []
    for column_index in relevant_columns:
        parts = []
        for row_number in header_rows:
            value = _normalize(_effective_cell_value(sheet, row_number, column_index))
            if value and (not parts or parts[-1] != value):
                parts.append(value)
        if not parts:
            continue
        signature = " / ".join(parts)
        columns.append(
            HeaderColumn(
                column_index=column_index,
                excel_column=get_column_letter(column_index),
                raw_header_parts=parts,
                signature=signature,
            )
        )
    return columns


def _column_has_signal(sheet: Worksheet, column_index: int, header_rows: list[int], data_start_row: int) -> bool:
    for row_number in [*header_rows, data_start_row]:
        if _normalize(_effective_cell_value(sheet, row_number, column_index)):
            return True
    return False


def _build_header_tree(sheet: Worksheet, header_rows: list[int], column_indexes: list[int]) -> list[HeaderTreeNode]:
    if len(header_rows) == 1:
        return [
            HeaderTreeNode(
                label=_normalize(_effective_cell_value(sheet, header_rows[0], column_index)),
                row_number=header_rows[0],
                start_column=column_index,
                end_column=column_index,
                children=[],
            )
            for column_index in column_indexes
            if _normalize(_effective_cell_value(sheet, header_rows[0], column_index))
        ]

    parent_row, child_row = header_rows[0], header_rows[-1]
    nodes: list[HeaderTreeNode] = []
    current_node: HeaderTreeNode | None = None
    for column_index in column_indexes:
        parent_label = _normalize(_effective_cell_value(sheet, parent_row, column_index))
        child_label = _normalize(_effective_cell_value(sheet, child_row, column_index))
        if not parent_label and child_label:
            parent_label = child_label

        child_node = HeaderTreeNode(
            label=child_label or parent_label,
            row_number=child_row,
            start_column=column_index,
            end_column=column_index,
            children=[],
        )

        if current_node and current_node.label == parent_label:
            current_node.end_column = column_index
            current_node.children.append(child_node)
            continue

        current_node = HeaderTreeNode(
            label=parent_label,
            row_number=parent_row,
            start_column=column_index,
            end_column=column_index,
            children=[child_node],
        )
        nodes.append(current_node)
    return nodes


def _effective_cell_value(sheet: Worksheet, row_number: int, column_index: int) -> object:
    cell = sheet.cell(row=row_number, column=column_index)
    if isinstance(cell, MergedCell):
        return _merged_parent_value(sheet, row_number, column_index)
    if cell.value is not None:
        return cell.value
    return _merged_parent_value(sheet, row_number, column_index)


def _merged_parent_value(sheet: Worksheet, row_number: int, column_index: int) -> object:
    for merged_range in sheet.merged_cells.ranges:
        if merged_range.min_row <= row_number <= merged_range.max_row and merged_range.min_col <= column_index <= merged_range.max_col:
            return sheet.cell(merged_range.min_row, merged_range.min_col).value
    return None


def _normalize(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _row_looks_like_data(values: Iterable[str]) -> bool:
    values = [value for value in values if value]
    if len(values) < 2:
        return False
    numeric_like = sum(1 for value in values if _looks_numeric(value))
    return numeric_like >= max(2, len(values) // 2)


def _looks_numeric(value: str) -> bool:
    candidate = value.replace(",", "").replace("%", "")
    if not candidate:
        return False
    try:
        float(candidate)
        return True
    except ValueError:
        return False
