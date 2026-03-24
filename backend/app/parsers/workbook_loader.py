from __future__ import annotations

from datetime import time
from pathlib import Path

import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException


def load_workbook_compatible(
    path: str | Path,
    *,
    read_only: bool,
    data_only: bool,
):
    workbook_path = Path(path)
    if workbook_path.suffix.lower() == '.xls':
        return _load_xls_workbook(workbook_path)

    try:
        return load_workbook(workbook_path, read_only=read_only, data_only=data_only)
    except (InvalidFileException, OSError):
        return _load_xls_workbook(workbook_path)


def _load_xls_workbook(path: Path) -> Workbook:
    legacy_workbook = xlrd.open_workbook(path.as_posix(), formatting_info=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    for legacy_sheet in legacy_workbook.sheets():
        sheet = workbook.create_sheet(title=legacy_sheet.name[:31] or 'Sheet')
        for row_index in range(legacy_sheet.nrows):
            for column_index in range(legacy_sheet.ncols):
                cell = legacy_sheet.cell(row_index, column_index)
                sheet.cell(row=row_index + 1, column=column_index + 1, value=_convert_xls_cell_value(cell, legacy_workbook.datemode))

        for merged_range in getattr(legacy_sheet, 'merged_cells', []):
            start_row, end_row, start_column, end_column = merged_range
            if end_row - start_row <= 1 and end_column - start_column <= 1:
                continue
            sheet.merge_cells(
                start_row=start_row + 1,
                end_row=end_row,
                start_column=start_column + 1,
                end_column=end_column,
            )

    return workbook


def _convert_xls_cell_value(cell: xlrd.sheet.Cell, datemode: int):
    if cell.ctype == xlrd.XL_CELL_EMPTY or cell.ctype == xlrd.XL_CELL_BLANK:
        return None
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            date_value = xlrd.xldate_as_datetime(cell.value, datemode)
        except (OverflowError, ValueError):
            return cell.value
        if date_value.time() == time(0, 0):
            return date_value.date()
        return date_value
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        numeric_value = float(cell.value)
        if numeric_value.is_integer():
            return int(numeric_value)
        return numeric_value
    if cell.ctype == xlrd.XL_CELL_ERROR:
        return None
    return cell.value
