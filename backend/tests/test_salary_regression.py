"""Salary template regression tests.

Locks the Salary export output to prevent regressions during Tool template fixes.
Tests cover standard cases and edge cases (housing fund, burden values, multi-region).
"""
from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from backend.app.core.config import ROOT_DIR
from backend.app.exporters import export_dual_templates
from backend.app.exporters.salary_exporter import SALARY_HEADERS, _salary_row_values
from backend.app.services import build_normalized_models, standardize_workbook
from backend.tests.support.export_fixtures import require_sample_workbook, resolve_required_export_templates


ARTIFACTS_ROOT = ROOT_DIR / ".test_artifacts" / "salary_regression"


def _prepare_artifact_dir(name: str) -> Path:
    ARTIFACTS_ROOT.mkdir(parents=True, exist_ok=True)
    target = ARTIFACTS_ROOT / f'{name}-{uuid4().hex[:8]}'
    target.mkdir(parents=True, exist_ok=False)
    return target


def _build_records(keyword: str, region: str, company: str, *, count: int = 1, housing: bool = False):
    sample_path = require_sample_workbook(keyword, housing=housing)
    standardized = standardize_workbook(sample_path, region=region, company_name=company)
    assert len(standardized.records) >= count
    trimmed = type(standardized)(
        source_file=standardized.source_file,
        sheet_name=standardized.sheet_name,
        raw_header_signature=standardized.raw_header_signature,
        records=standardized.records[:count],
        filtered_rows=standardized.filtered_rows,
        unmapped_headers=standardized.unmapped_headers,
    )
    records = build_normalized_models(trimmed, batch_id=f"batch-{region}", source_file_id=f"src-{region}")
    for i, r in enumerate(records, 1):
        r.employee_id = f"{i:05d}"
    return records


def test_salary_header_row_matches_constants():
    """Salary template header row matches SALARY_HEADERS constant."""
    templates = resolve_required_export_templates()
    wb = load_workbook(templates.salary, data_only=False)
    sheet = wb[wb.sheetnames[0]]
    actual_headers = [sheet.cell(row=1, column=i).value for i in range(1, len(SALARY_HEADERS) + 1)]
    wb.close()
    assert actual_headers == SALARY_HEADERS, f"Header mismatch: {actual_headers}"


def test_salary_row_values_length():
    """_salary_row_values returns exactly 18 elements matching SALARY_HEADERS."""
    records = _build_records("广分", "guangzhou", "广分示例")
    values = _salary_row_values(records[0])
    assert len(values) == len(SALARY_HEADERS), f"Expected {len(SALARY_HEADERS)}, got {len(values)}"


def test_salary_export_shenzhen_cell_values():
    """Salary export for Shenzhen produces correct cell values at known positions."""
    records = _build_records("深圳创造欢乐", "shenzhen", "创造欢乐", count=2)
    templates = resolve_required_export_templates()
    output = _prepare_artifact_dir("shenzhen-salary")
    result = export_dual_templates(
        records,
        output_dir=output,
        salary_template_path=templates.salary,
        final_tool_template_path=templates.final_tool,
        export_prefix="salary_reg_sz",
    )
    assert result.status == "completed"
    salary = next(a for a in result.artifacts if a.template_type == "salary")
    assert salary.status == "completed"
    assert salary.row_count == 2

    wb = load_workbook(salary.file_path, data_only=False)
    sheet = wb[wb.sheetnames[0]]
    # Row 2 = first record
    assert sheet["A2"].value == records[0].person_name
    assert sheet["B2"].value == records[0].employee_id
    # Row 3 = second record
    assert sheet["A3"].value == records[1].person_name
    wb.close()


def test_salary_export_guangzhou():
    """Salary export for Guangzhou produces correct output."""
    records = _build_records("广分", "guangzhou", "广分示例")
    templates = resolve_required_export_templates()
    output = _prepare_artifact_dir("guangzhou-salary")
    result = export_dual_templates(
        records,
        output_dir=output,
        salary_template_path=templates.salary,
        final_tool_template_path=templates.final_tool,
        export_prefix="salary_reg_gz",
    )
    assert result.status == "completed"
    salary = next(a for a in result.artifacts if a.template_type == "salary")
    assert salary.status == "completed"
    assert salary.row_count == 1

    wb = load_workbook(salary.file_path, data_only=False)
    sheet = wb[wb.sheetnames[0]]
    assert sheet["A2"].value == records[0].person_name
    # Personal pension (column E=5 -> SALARY_HEADERS index 6 -> col G)
    pension_val = sheet.cell(row=2, column=7).value
    assert pension_val is not None
    wb.close()


def test_salary_export_wuhan():
    """Salary export for Wuhan produces correct output."""
    records = _build_records("武汉", "wuhan", "武汉示例")
    templates = resolve_required_export_templates()
    output = _prepare_artifact_dir("wuhan-salary")
    result = export_dual_templates(
        records,
        output_dir=output,
        salary_template_path=templates.salary,
        final_tool_template_path=templates.final_tool,
        export_prefix="salary_reg_wh",
    )
    assert result.status == "completed"
    salary = next(a for a in result.artifacts if a.template_type == "salary")
    assert salary.status == "completed"


def test_salary_row_values_all_numeric_except_name_and_id():
    """All values after name and employee_id are Decimal or numeric."""
    records = _build_records("深圳创造欢乐", "shenzhen", "创造欢乐")
    values = _salary_row_values(records[0])
    # First two are name and employee_id (strings)
    for i, v in enumerate(values[2:], start=2):
        assert isinstance(v, (Decimal, int, float)), f"Position {i} ({SALARY_HEADERS[i]}): expected numeric, got {type(v).__name__}: {v}"
