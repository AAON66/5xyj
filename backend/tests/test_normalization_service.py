from __future__ import annotations

from datetime import datetime
from pathlib import Path
from decimal import Decimal
import sqlite3

import pytest
from openpyxl import Workbook

from backend.app.core.config import ROOT_DIR
from backend.app.models.enums import SourceFileKind
from backend.app.services import (
    SourceRecordBundle,
    build_normalized_models,
    merge_batch_standardized_records,
    standardize_housing_fund_workbook,
    standardize_workbook,
)


SAMPLES_DIR = ROOT_DIR / 'data' / 'samples'
HOUSING_SAMPLES_DIR = SAMPLES_DIR / '公积金'
APP_DB = ROOT_DIR / 'data' / 'app.db'


def find_sample(keyword: str) -> Path:
    for path in sorted(SAMPLES_DIR.glob('*.xlsx')):
        if keyword in path.name:
            return path
    pytest.skip(f'Sample containing {keyword!r} was not found in {SAMPLES_DIR}.')


def find_housing_sample(keyword: str) -> Path:
    for path in sorted(HOUSING_SAMPLES_DIR.glob('*.xlsx')):
        if keyword in path.name:
            return path
    pytest.skip(f'Housing sample containing {keyword!r} was not found in {HOUSING_SAMPLES_DIR}.')


def find_uploaded_sample(filename: str) -> Path:
    if not APP_DB.exists():
        pytest.skip(f'Application database was not found at {APP_DB}.')

    connection = sqlite3.connect(APP_DB)
    try:
        row = connection.execute(
            'select file_path from source_files where file_name = ? order by uploaded_at desc limit 1',
            (filename,),
        ).fetchone()
    finally:
        connection.close()

    if row is None:
        pytest.skip(f'Uploaded sample {filename!r} was not found in {APP_DB}.')

    sample_path = Path(row[0])
    if not sample_path.exists():
        pytest.skip(f'Uploaded sample path no longer exists: {sample_path}.')
    return sample_path


def test_standardize_workbook_on_real_guangzhou_sample() -> None:
    sample_path = find_sample('\u5e7f\u5206')

    result = standardize_workbook(sample_path, region='guangzhou')

    first = result.records[0]
    assert result.sheet_name == 'sheet1'
    assert first.source_row_number == 9
    assert first.values['person_name']
    assert first.values['id_number']
    assert first.values['social_security_number']
    assert first.values['billing_period'] == '2026-02'
    assert first.values['company_total_amount'] is not None
    assert first.values['personal_total_amount'] is not None
    assert first.values['pension_company'] is not None
    assert first.values['medical_maternity_company'] is not None
    assert first.values['injury_company'] is not None
    assert first.raw_values['\u5e8f\u53f7'] == 1
    assert first.unmapped_values['\u5e8f\u53f7'] == 1


def test_standardize_workbook_on_real_hangzhou_sample() -> None:
    sample_path = find_sample('\u676d\u5dde\u805a\u53d8')

    result = standardize_workbook(sample_path, region='hangzhou')

    first = result.records[0]
    assert len(result.records) > 0
    assert any(row.reason == 'summary_total' for row in result.filtered_rows)
    assert first.values['person_name']
    assert first.values['id_number']
    assert first.values['pension_company'] is not None
    assert first.values['pension_personal'] is not None
    assert first.values['medical_company'] is not None
    assert first.values['medical_personal'] is not None
    assert first.values['unemployment_company'] is not None
    assert first.values['injury_company'] is not None
    assert '\u5408\u8ba1' in first.unmapped_values


def test_standardize_workbook_on_real_xiamen_sample() -> None:
    sample_path = find_sample('\u53a6\u95e8202602\u793e\u4fdd\u8d26\u5355.xlsx')

    result = standardize_workbook(sample_path, region='xiamen')

    first = result.records[0]
    assert len(result.records) > 0
    assert any(row.row_number == 37 and row.reason == 'summary_total' for row in result.filtered_rows)
    assert first.values['billing_period'] == '2026-02'
    assert first.values['period_start'] == '2026-02-01'
    assert first.values['period_end'] == '2026-02-28'
    assert first.values['total_amount'] is not None
    assert first.values['company_total_amount'] is not None
    assert first.values['personal_total_amount'] is not None
    assert first.values['maternity_amount'] == Decimal('31.03')
    assert first.values['late_fee'] == Decimal('0')
    assert first.values['interest'] == Decimal('0')
    assert '\u53c2\u4fdd\u4eba\u5458\u8eab\u4efd' in first.unmapped_values


def test_standardize_workbook_on_real_shenzhen_sample() -> None:
    sample_path = find_sample('\u6df1\u5733\u521b\u9020\u6b22\u4e50')

    result = standardize_workbook(sample_path, region='shenzhen')

    first = result.records[0]
    filtered_reasons = {(row.reason, row.first_value) for row in result.filtered_rows}
    assert ('summary_subtotal', '\u5c0f\u8ba1') in filtered_reasons
    assert ('group_header', '\u9000\u4f11\u4eba\u5458') in filtered_reasons
    assert ('group_header', '\u5bb6\u5c5e\u7edf\u7b79\u4eba\u5458') in filtered_reasons
    assert first.source_row_number == 4
    assert first.values['total_amount'] is not None
    assert first.values['company_total_amount'] is not None
    assert first.values['personal_total_amount'] is not None
    assert first.values['supplementary_pension_company'] is not None
    assert '\u57fa\u672c\u517b\u8001\u4fdd\u9669\uff08\u5355\u4f4d\uff09 / \u8d39\u7387' in first.unmapped_values


def test_standardize_workbook_on_real_wuhan_sample() -> None:
    sample_path = find_sample('\u6b66\u6c49')

    result = standardize_workbook(sample_path, region='wuhan')

    first = result.records[0]
    assert len(result.records) == 1
    assert any(row.row_number == 4 and row.reason == 'summary_total' for row in result.filtered_rows)
    assert first.values['person_name']
    assert first.values['payment_base'] is not None
    assert first.values['payment_salary'] is not None
    assert first.values['pension_company'] is not None
    assert first.values['pension_personal'] is not None
    assert first.values['medical_company'] is not None
    assert first.values['medical_personal'] is not None


def test_standardize_workbook_on_wuhan_transactional_sample() -> None:
    artifacts_dir = ROOT_DIR / '.test_artifacts' / 'normalization_service'
    artifacts_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = artifacts_dir / 'wuhan_transactional_202603.xlsx'

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Sheet1'
    sheet.append(
        [
            '\u59d3\u540d',
            '\u8bc1\u4ef6\u53f7\u7801',
            '\u8d39\u6b3e\u6240\u5c5e\u671f',
            '\u9669\u79cd',
            '\u5e94\u7f34\u8d39\u989d\u5408\u8ba1',
            '\u5355\u4f4d\u90e8\u5206',
            None,
            None,
            None,
            None,
            '\u4e2a\u4eba\u90e8\u5206',
            None,
            None,
            None,
            None,
            '\u6570\u636e\u6765\u6e90',
            '\u7f34\u8d39\u7c7b\u578b',
            '\u4e3b\u7ba1\u7a0e\u52a1\u673a\u5173',
            '\u793e\u4fdd\u7ecf\u529e\u673a\u6784',
        ]
    )
    sheet.append(
        [
            None,
            None,
            None,
            None,
            None,
            '\u7f34\u8d39\u4eba\u6570',
            '\u7f34\u8d39\u5de5\u8d44',
            '\u7f34\u8d39\u57fa\u6570',
            '\u8d39\u7387',
            '\u5e94\u7f34\u8d39\u989d\uff08\u5143\uff09',
            '\u7f34\u8d39\u4eba\u6570',
            '\u7f34\u8d39\u5de5\u8d44',
            '\u7f34\u8d39\u57fa\u6570',
            '\u8d39\u7387',
            '\u5e94\u7f34\u8d39\u989d\uff08\u5143\uff09',
            None,
            None,
            None,
            None,
        ]
    )
    sheet.append(
        [
            '\u5d14\u4e9a\u946b',
            '411326199709232828',
            datetime(2026, 3, 1),
            '\u4f01\u4e1a\u804c\u5de5\u57fa\u672c\u517b\u8001\u4fdd\u9669',
            1079.52,
            1,
            4498,
            4498,
            0.16,
            719.68,
            1,
            4498,
            4498,
            0.08,
            359.84,
            '\u65e5\u5e38\u7f34\u8d39',
            None,
            '\u56fd\u5bb6\u7a0e\u52a1\u603b\u5c40\u6b66\u6c49\u4e1c\u6e56\u65b0\u6280\u672f\u5f00\u53d1\u533a\u7a0e\u52a1\u5c40',
            '\u4e1c\u6e56\u9ad8\u65b0\u5f00\u53d1\u533a\u4eba\u529b\u8d44\u6e90\u548c\u793e\u4f1a\u4fdd\u969c\u5c40(\u517b\u8001\u5931\u4e1a\u5de5\u4f24)',
        ]
    )
    sheet.append([None] * 19)
    sheet.append(
        [
            None,
            None,
            datetime(2026, 3, 1),
            '\u5931\u4e1a\u4fdd\u9669',
            44.98,
            1,
            4498,
            4498,
            0.007,
            31.49,
            1,
            4498,
            4498,
            0.003,
            13.49,
            '\u65e5\u5e38\u7f34\u8d39',
            None,
            '\u56fd\u5bb6\u7a0e\u52a1\u603b\u5c40\u6b66\u6c49\u4e1c\u6e56\u65b0\u6280\u672f\u5f00\u53d1\u533a\u7a0e\u52a1\u5c40',
            '\u4e1c\u6e56\u9ad8\u65b0\u5f00\u53d1\u533a\u4eba\u529b\u8d44\u6e90\u548c\u793e\u4f1a\u4fdd\u969c\u5c40(\u517b\u8001\u5931\u4e1a\u5de5\u4f24)',
        ]
    )
    sheet.append([None] * 19)
    sheet.append(
        [
            None,
            None,
            datetime(2026, 3, 1),
            '\u5de5\u4f24\u4fdd\u9669',
            9,
            1,
            4498,
            4498,
            0.002,
            9,
            '--',
            '--',
            '--',
            '--',
            '--',
            '\u65e5\u5e38\u7f34\u8d39',
            None,
            '\u56fd\u5bb6\u7a0e\u52a1\u603b\u5c40\u6b66\u6c49\u4e1c\u6e56\u65b0\u6280\u672f\u5f00\u53d1\u533a\u7a0e\u52a1\u5c40',
            '\u4e1c\u6e56\u9ad8\u65b0\u5f00\u53d1\u533a\u4eba\u529b\u8d44\u6e90\u548c\u793e\u4f1a\u4fdd\u969c\u5c40(\u517b\u8001\u5931\u4e1a\u5de5\u4f24)',
        ]
    )
    sheet.append([None] * 19)
    sheet.append(
        [
            None,
            None,
            datetime(2026, 3, 1),
            '\u4f01\u4e1a\u804c\u5de5\u57fa\u672c\u533b\u7597\u4fdd\u9669',
            481.29,
            1,
            4498,
            4498,
            0.087,
            391.33,
            1,
            4498,
            4498,
            0.02,
            89.96,
            '\u65e5\u5e38\u7f34\u8d39',
            None,
            '\u56fd\u5bb6\u7a0e\u52a1\u603b\u5c40\u6b66\u6c49\u4e1c\u6e56\u65b0\u6280\u672f\u5f00\u53d1\u533a\u7a0e\u52a1\u5c40',
            '\u4e1c\u6e56\u9ad8\u65b0\u5f00\u53d1\u533a\u533b\u7597\u4fdd\u969c\u5c40(\u533b\u7597)',
        ]
    )
    sheet.append([None] * 19)
    sheet.append(
        [
            None,
            None,
            datetime(2026, 3, 1),
            '\u804c\u5de5\u5927\u989d\u533b\u7597\u4e92\u52a9\u4fdd\u9669',
            7,
            1,
            4498,
            '--',
            1,
            '--',
            1,
            4498,
            7,
            1,
            7,
            '\u65e5\u5e38\u7f34\u8d39',
            None,
            '\u56fd\u5bb6\u7a0e\u52a1\u603b\u5c40\u6b66\u6c49\u4e1c\u6e56\u65b0\u6280\u672f\u5f00\u53d1\u533a\u7a0e\u52a1\u5c40',
            '\u4e1c\u6e56\u9ad8\u65b0\u5f00\u53d1\u533a\u533b\u7597\u4fdd\u969c\u5c40(\u533b\u7597)',
        ]
    )
    for column in ('A', 'B', 'C', 'D', 'E', 'P', 'Q', 'R', 'S'):
        sheet.merge_cells(f'{column}1:{column}2')
    sheet.merge_cells('F1:J1')
    sheet.merge_cells('K1:O1')
    workbook.save(workbook_path)

    result = standardize_workbook(workbook_path, region='wuhan', company_name='\u6b66\u6c49\u793a\u4f8b\u516c\u53f8')

    assert len(result.records) == 1
    assert len(result.filtered_rows) == 4
    first = result.records[0]
    assert first.source_row_number == 3
    assert first.values['person_name'] == '\u5d14\u4e9a\u946b'
    assert first.values['id_number'] == '411326199709232828'
    assert first.values['billing_period'] == '2026-03'
    assert first.values['payment_base'] == Decimal('4498')
    assert first.values['payment_salary'] == Decimal('4498')
    assert first.values['company_total_amount'] == Decimal('1151.50')
    assert first.values['personal_total_amount'] == Decimal('470.29')
    assert first.values['total_amount'] == Decimal('1621.79')
    assert first.values['pension_company'] == Decimal('719.68')
    assert first.values['pension_personal'] == Decimal('359.84')
    assert first.values['unemployment_company'] == Decimal('31.49')
    assert first.values['unemployment_personal'] == Decimal('13.49')
    assert first.values['injury_company'] == Decimal('9')
    assert first.values['medical_company'] == Decimal('391.33')
    assert first.values['medical_personal'] == Decimal('89.96')
    assert first.values['large_medical_personal'] == Decimal('7')
    assert first.raw_payload['merge_strategy'] == 'wuhan_transactional_by_insurance_item'
    assert len(first.raw_payload['merged_sources']) == 5
    assert first.raw_payload['merged_sources'][0]['raw_values']['费款所属期'] == '2026-03-01T00:00:00'


def test_standardize_workbook_on_real_changsha_sample() -> None:
    sample_path = find_sample('\u957f\u6c99')

    result = standardize_workbook(sample_path, region='changsha')

    first = result.records[0]
    assert result.sheet_name == 'Sheet4'
    assert first.source_row_number == 5
    assert first.values['person_name']
    assert first.values['pension_company'] is not None
    assert first.values['medical_company'] is not None
    assert first.values['unemployment_personal'] is not None
    assert first.values['large_medical_personal'] is not None
    assert first.values['total_amount'] is not None


def test_standardize_workbook_on_uploaded_changsha_transactional_sample_maps_large_medical() -> None:
    sample_path = find_uploaded_sample('长沙202602社保账单.xlsx')

    result = standardize_workbook(sample_path, region='changsha')

    match = next(
        record
        for record in result.records
        if record.raw_values.get('征收品目') == '职工大额医疗互助保险(个人缴纳)'
    )
    assert match.values['person_name'] == '余宸瑶'
    assert match.values['total_amount'] == Decimal('15')
    assert match.values['large_medical_personal'] == Decimal('15')


def test_build_normalized_models_preserves_provenance() -> None:
    sample_path = find_sample('\u5e7f\u5206')

    result = standardize_workbook(sample_path, region='guangzhou')
    models = build_normalized_models(result, batch_id='batch-1', source_file_id='source-1')

    assert models
    assert models[0].batch_id == 'batch-1'
    assert models[0].source_file_id == 'source-1'
    assert models[0].source_file_name == sample_path.name
    assert models[0].raw_sheet_name == 'sheet1'
    assert models[0].raw_header_signature == result.raw_header_signature
    assert models[0].raw_payload is not None
    assert models[0].raw_payload['raw_values']['\u59d3\u540d'] == result.records[0].values['person_name']


def test_merge_batch_standardized_records_merges_name_only_housing_rows_into_hangzhou_social_rows() -> None:
    social_sample = find_sample('\u676d\u5dde\u88c2\u53d8')
    housing_sample = find_housing_sample('\u676d\u5dde\u88c2\u53d8')

    social_result = standardize_workbook(social_sample, region='hangzhou', company_name='\u96f6\u4e00\u88c2\u53d8')
    housing_result = standardize_housing_fund_workbook(housing_sample, region='hangzhou')

    social_names = {record.values.get('person_name') for record in social_result.records}
    target_housing_record = next(
        record
        for record in housing_result.records
        if record.values.get('person_name') in social_names and record.values.get('housing_fund_company') is not None
    )

    merged = merge_batch_standardized_records(
        [
            SourceRecordBundle(
                source_file_id='social-source',
                source_file_name=social_sample.name,
                source_kind=SourceFileKind.SOCIAL_SECURITY.value,
                standardized=social_result,
            ),
            SourceRecordBundle(
                source_file_id='housing-source',
                source_file_name=housing_sample.name,
                source_kind=SourceFileKind.HOUSING_FUND.value,
                standardized=housing_result,
            ),
        ],
        batch_id='batch-merge',
    )

    merged_matches = [record for record in merged if record.person_name == target_housing_record.values.get('person_name')]
    assert len(merged_matches) == 1
    merged_record = merged_matches[0]
    assert merged_record.pension_company is not None
    assert merged_record.housing_fund_company == target_housing_record.values.get('housing_fund_company')
    assert merged_record.housing_fund_personal == target_housing_record.values.get('housing_fund_personal')
