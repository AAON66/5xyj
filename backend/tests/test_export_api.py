from __future__ import annotations

from typing import Optional

import json
import re
import shutil
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from backend.app.core.config import ROOT_DIR, Settings
from backend.app.core.database import create_database_engine, create_session_factory
from backend.app.dependencies import get_db
from backend.app.main import create_app
from backend.app.models import Base
from backend.app.models.employee_master import EmployeeMaster
from backend.app.services import standardize_workbook
from backend.tests.support.export_fixtures import (
    create_placeholder_template_pair,
    require_sample_workbook,
    resolve_required_export_templates,
)


ARTIFACTS_ROOT = ROOT_DIR / '.test_artifacts' / 'export_api'

SAMPLE_KEYWORD = '\u6df1\u5733\u521b\u9020\u6b22\u4e50'
COMPANY_NAME = '\u521b\u9020\u6b22\u4e50'
APP_NAME = '\u5bfc\u51fa\u6d4b\u8bd5'


def build_test_context(test_name: str, *, salary_template: Optional[Path], final_tool_template: Optional[Path]):
    artifacts_dir = ARTIFACTS_ROOT / test_name
    if artifacts_dir.exists():
        shutil.rmtree(artifacts_dir)
    artifacts_dir.mkdir(parents=True, exist_ok=True)

    database_path = artifacts_dir / 'export.db'
    settings = Settings(
        app_name=APP_NAME,
        app_version='0.2.0',
        auth_enabled=False,
        database_url=f'sqlite:///{database_path.as_posix()}',
        upload_dir=str(artifacts_dir / 'uploads'),
        samples_dir=str(artifacts_dir / 'samples'),
        templates_dir=str(artifacts_dir / 'templates'),
        outputs_dir=str(artifacts_dir / 'outputs'),
        salary_template_path=str(salary_template) if salary_template is not None else None,
        final_tool_template_path=str(final_tool_template) if final_tool_template is not None else None,
        log_format='plain',
    )

    engine = create_database_engine(settings)
    session_factory = create_session_factory(settings)
    Base.metadata.create_all(engine)

    app = create_app(settings)

    def override_get_db():
        db: Session = session_factory()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_get_db
    return TestClient(app), settings, session_factory


def create_batch(client: TestClient, sample_path: Path, *, batch_name: Optional[str] = 'export-batch') -> str:
    data = {
        'regions': json.dumps(['shenzhen'], ensure_ascii=False),
        'company_names': json.dumps([COMPANY_NAME], ensure_ascii=False),
    }
    if batch_name is not None:
        data['batch_name'] = batch_name

    response = client.post(
        '/api/v1/imports',
        data=data,
        files=[('files', (sample_path.name, sample_path.read_bytes(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))],
    )
    assert response.status_code == 201
    return response.json()['data']['id']


def seed_employee_for_first_record(session_factory, sample_path: Path) -> None:
    standardized = standardize_workbook(sample_path, region='shenzhen', company_name=COMPANY_NAME)
    first = standardized.records[0]
    db: Session = session_factory()
    try:
        db.add(
            EmployeeMaster(
                employee_id='01620',
                person_name=first.values['person_name'],
                id_number=first.values['id_number'],
                company_name=COMPANY_NAME,
                active=True,
            )
        )
        db.commit()
    finally:
        db.close()


def test_export_batch_endpoint_writes_both_template_outputs() -> None:
    templates = resolve_required_export_templates()
    sample_path = require_sample_workbook(SAMPLE_KEYWORD)
    client, _settings, session_factory = build_test_context(
        'export_success',
        salary_template=templates.salary,
        final_tool_template=templates.final_tool,
    )
    seed_employee_for_first_record(session_factory, sample_path)

    with client:
        batch_id = create_batch(client, sample_path)
        match_response = client.post(f'/api/v1/imports/{batch_id}/match')
        export_response = client.post(f'/api/v1/imports/{batch_id}/export')
        get_response = client.get(f'/api/v1/imports/{batch_id}/export')
        detail_response = client.get(f'/api/v1/imports/{batch_id}')

    assert match_response.status_code == 200
    assert export_response.status_code == 200
    payload = export_response.json()['data']
    assert payload['batch_id'] == batch_id
    assert payload['status'] == 'exported'
    assert payload['export_status'] == 'completed'
    assert len(payload['artifacts']) == 2
    assert all(item['status'] == 'completed' for item in payload['artifacts'])
    assert all(Path(item['file_path']).exists() for item in payload['artifacts'])
    assert Path(next(item for item in payload['artifacts'] if item['template_type'] == 'salary')['file_path']).name == 'export-batch_salary.xlsx'
    assert Path(next(item for item in payload['artifacts'] if item['template_type'] == 'final_tool')['file_path']).name == 'export-batch_final_tool.xlsx'

    get_payload = get_response.json()['data']
    assert get_payload['export_job_id'] == payload['export_job_id']
    assert get_payload['export_status'] == 'completed'
    assert len(get_payload['artifacts']) == 2
    assert detail_response.json()['data']['status'] == 'exported'


def test_export_batch_endpoint_uses_timestamp_filename_when_batch_name_is_implicit() -> None:
    templates = resolve_required_export_templates()
    sample_path = require_sample_workbook(SAMPLE_KEYWORD)
    client, _settings, session_factory = build_test_context(
        'export_timestamp_name',
        salary_template=templates.salary,
        final_tool_template=templates.final_tool,
    )
    seed_employee_for_first_record(session_factory, sample_path)

    with client:
        batch_id = create_batch(client, sample_path, batch_name=None)
        match_response = client.post(f'/api/v1/imports/{batch_id}/match')
        export_response = client.post(f'/api/v1/imports/{batch_id}/export')
        salary_download = client.get(f'/api/v1/imports/{batch_id}/export/salary/download')
        tool_download = client.get(f'/api/v1/imports/{batch_id}/export/final_tool/download')

    assert match_response.status_code == 200
    assert export_response.status_code == 200
    payload = export_response.json()['data']
    salary_name = Path(next(item for item in payload['artifacts'] if item['template_type'] == 'salary')['file_path']).name
    tool_name = Path(next(item for item in payload['artifacts'] if item['template_type'] == 'final_tool')['file_path']).name
    assert re.fullmatch(r'\d{8}-\d{6}_salary\.xlsx', salary_name)
    assert re.fullmatch(r'\d{8}-\d{6}_final_tool\.xlsx', tool_name)
    assert 'filename=' in salary_download.headers['content-disposition']
    assert salary_name in salary_download.headers['content-disposition']
    assert tool_name in tool_download.headers['content-disposition']


def test_export_batch_endpoint_sanitizes_and_truncates_batch_name_for_export_files() -> None:
    templates = resolve_required_export_templates()
    sample_path = require_sample_workbook(SAMPLE_KEYWORD)
    client, _settings, session_factory = build_test_context(
        'export_sanitized_name',
        salary_template=templates.salary,
        final_tool_template=templates.final_tool,
    )
    seed_employee_for_first_record(session_factory, sample_path)

    noisy_batch_name = '  2026/02:深圳*社保?公积金<>|  导出___批次___名字特别特别特别长特别特别长  '

    with client:
        batch_id = create_batch(client, sample_path, batch_name=noisy_batch_name)
        match_response = client.post(f'/api/v1/imports/{batch_id}/match')
        export_response = client.post(f'/api/v1/imports/{batch_id}/export')

    assert match_response.status_code == 200
    assert export_response.status_code == 200
    payload = export_response.json()['data']
    salary_name = Path(next(item for item in payload['artifacts'] if item['template_type'] == 'salary')['file_path']).name
    tool_name = Path(next(item for item in payload['artifacts'] if item['template_type'] == 'final_tool')['file_path']).name
    assert salary_name.endswith('_salary.xlsx')
    assert tool_name.endswith('_final_tool.xlsx')
    assert len(salary_name) <= 44
    assert len(tool_name) <= 48
    assert all(char not in salary_name for char in '\\/:*?"<>|')
    assert all(char not in tool_name for char in '\\/:*?"<>|')
    assert '__' not in salary_name
    assert '__' not in tool_name


def test_export_batch_endpoint_falls_back_to_timestamp_when_sanitized_batch_name_is_empty() -> None:
    templates = resolve_required_export_templates()
    sample_path = require_sample_workbook(SAMPLE_KEYWORD)
    client, _settings, session_factory = build_test_context(
        'export_empty_after_sanitize',
        salary_template=templates.salary,
        final_tool_template=templates.final_tool,
    )
    seed_employee_for_first_record(session_factory, sample_path)

    with client:
        batch_id = create_batch(client, sample_path, batch_name='  <>:"/\\\\|?*  ')
        match_response = client.post(f'/api/v1/imports/{batch_id}/match')
        export_response = client.post(f'/api/v1/imports/{batch_id}/export')

    assert match_response.status_code == 200
    assert export_response.status_code == 200
    payload = export_response.json()['data']
    salary_name = Path(next(item for item in payload['artifacts'] if item['template_type'] == 'salary')['file_path']).name
    tool_name = Path(next(item for item in payload['artifacts'] if item['template_type'] == 'final_tool')['file_path']).name
    assert re.fullmatch(r'\d{8}-\d{6}_salary\.xlsx', salary_name)
    assert re.fullmatch(r'\d{8}-\d{6}_final_tool\.xlsx', tool_name)


def test_export_batch_endpoint_discovers_templates_from_templates_dir_when_config_is_omitted() -> None:
    templates = resolve_required_export_templates()
    sample_path = require_sample_workbook(SAMPLE_KEYWORD)
    client, settings, session_factory = build_test_context(
        'export_template_discovery_fallback',
        salary_template=None,
        final_tool_template=None,
    )
    seed_employee_for_first_record(session_factory, sample_path)

    discovered_dir = settings.templates_path / 'nested'
    discovered_dir.mkdir(parents=True, exist_ok=True)
    shutil.copy2(templates.salary, discovered_dir / templates.salary.name)
    shutil.copy2(templates.final_tool, discovered_dir / templates.final_tool.name)
    shutil.copy2(templates.manifest, discovered_dir / templates.manifest.name)

    with client:
        batch_id = create_batch(client, sample_path, batch_name='export-template-discovery-fallback')
        match_response = client.post(f'/api/v1/imports/{batch_id}/match')
        export_response = client.post(f'/api/v1/imports/{batch_id}/export')

    assert match_response.status_code == 200
    assert export_response.status_code == 200
    payload = export_response.json()['data']
    assert payload['export_status'] == 'completed'
    assert all(item['status'] == 'completed' for item in payload['artifacts'])
    assert all(Path(item['file_path']).exists() for item in payload['artifacts'])


def test_export_batch_endpoint_prefers_valid_explicit_template_paths() -> None:
    sample_path = require_sample_workbook(SAMPLE_KEYWORD)
    explicit_templates = create_placeholder_template_pair(ARTIFACTS_ROOT / 'explicit_template_override')

    salary_workbook = load_workbook(explicit_templates.salary)
    salary_sheet = salary_workbook[salary_workbook.sheetnames[0]]
    salary_sheet['A1'] = 'EXPLICIT SALARY TEMPLATE'
    salary_workbook.save(explicit_templates.salary)
    salary_workbook.close()

    tool_workbook = load_workbook(explicit_templates.final_tool)
    tool_sheet = tool_workbook[tool_workbook.sheetnames[0]]
    tool_sheet['A6'] = 'EXPLICIT TOOL TEMPLATE'
    tool_workbook.save(explicit_templates.final_tool)
    tool_workbook.close()

    client, _settings, session_factory = build_test_context(
        'export_explicit_template_override',
        salary_template=explicit_templates.salary,
        final_tool_template=explicit_templates.final_tool,
    )
    seed_employee_for_first_record(session_factory, sample_path)

    with client:
        batch_id = create_batch(client, sample_path, batch_name='export-explicit-template-override')
        match_response = client.post(f'/api/v1/imports/{batch_id}/match')
        export_response = client.post(f'/api/v1/imports/{batch_id}/export')

    assert match_response.status_code == 200
    assert export_response.status_code == 200
    payload = export_response.json()['data']
    salary_artifact = next(item for item in payload['artifacts'] if item['template_type'] == 'salary')
    tool_artifact = next(item for item in payload['artifacts'] if item['template_type'] == 'final_tool')
    assert salary_artifact['status'] == 'completed'
    assert tool_artifact['status'] == 'completed'

    salary_output = load_workbook(salary_artifact['file_path'], data_only=False)
    salary_output_sheet = salary_output[salary_output.sheetnames[0]]
    assert salary_output_sheet['A1'].value == 'EXPLICIT SALARY TEMPLATE'
    salary_output.close()

    tool_output = load_workbook(tool_artifact['file_path'], data_only=False)
    tool_output_sheet = tool_output[tool_output.sheetnames[0]]
    assert tool_output_sheet['A6'].value == 'EXPLICIT TOOL TEMPLATE'
    tool_output.close()


def test_export_batch_endpoint_fails_when_any_template_is_missing() -> None:
    templates = resolve_required_export_templates()
    sample_path = require_sample_workbook(SAMPLE_KEYWORD)
    missing_tool = ARTIFACTS_ROOT / 'missing-template.xlsx'
    client, _settings, session_factory = build_test_context(
        'export_failed',
        salary_template=templates.salary,
        final_tool_template=missing_tool,
    )
    seed_employee_for_first_record(session_factory, sample_path)

    with client:
        batch_id = create_batch(client, sample_path)
        match_response = client.post(f'/api/v1/imports/{batch_id}/match')
        export_response = client.post(f'/api/v1/imports/{batch_id}/export')
        get_response = client.get(f'/api/v1/imports/{batch_id}/export')
        detail_response = client.get(f'/api/v1/imports/{batch_id}')

    assert match_response.status_code == 200
    assert export_response.status_code == 200
    payload = export_response.json()['data']
    assert payload['status'] == 'failed'
    assert payload['export_status'] == 'failed'
    salary_artifact = next(item for item in payload['artifacts'] if item['template_type'] == 'salary')
    tool_artifact = next(item for item in payload['artifacts'] if item['template_type'] == 'final_tool')
    assert salary_artifact['status'] == 'completed'
    assert tool_artifact['status'] == 'failed'
    assert tool_artifact['error_message']

    get_payload = get_response.json()['data']
    assert get_payload['export_status'] == 'failed'
    assert detail_response.json()['data']['status'] == 'failed'


def test_export_batch_endpoint_blocks_when_matching_has_not_run() -> None:
    templates = resolve_required_export_templates()
    sample_path = require_sample_workbook(SAMPLE_KEYWORD)
    client, _settings, _session_factory = build_test_context(
        'export_blocked',
        salary_template=templates.salary,
        final_tool_template=templates.final_tool,
    )

    with client:
        batch_id = create_batch(client, sample_path)
        export_response = client.post(f'/api/v1/imports/{batch_id}/export')
        get_response = client.get(f'/api/v1/imports/{batch_id}/export')

    assert export_response.status_code == 409
    assert export_response.json()['error']['code'] == 'http_error'
    assert 'matching' in export_response.json()['error']['message'].lower()
    assert get_response.status_code == 200
    assert get_response.json()['data']['export_job_id'] is None
    assert get_response.json()['data']['blocked_reason']

