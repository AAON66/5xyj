from __future__ import annotations

import shutil
from pathlib import Path

import pytest
from openpyxl import Workbook

from backend.app.core.config import ROOT_DIR
from backend.app.parsers import extract_header_structure


ARTIFACTS_ROOT = ROOT_DIR / ".test_artifacts" / "header_extraction"
SAMPLES_DIR = ROOT_DIR / "data" / "samples"


def create_artifact_dir(test_name: str) -> Path:
    target = ARTIFACTS_ROOT / test_name
    if target.exists():
        shutil.rmtree(target)
    target.mkdir(parents=True, exist_ok=True)
    return target


def find_sample(keyword: str) -> Path:
    for path in sorted(SAMPLES_DIR.glob("*.xlsx")):
        if keyword in path.name:
            return path
    pytest.skip(f"Sample containing '{keyword}' was not found in {SAMPLES_DIR}.")


def test_extract_header_structure_merges_parent_and_child_rows() -> None:
    artifacts_dir = create_artifact_dir("compound_header")
    workbook_path = artifacts_dir / "compound_header.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "申报明细"
    sheet.append(["标题"])
    sheet.append(["姓名", "证件号码", "基本养老保险", None, "失业保险", None])
    sheet.append([None, None, "缴费基数", "应缴金额", "缴费基数", "应缴金额"])
    sheet.append(["张三", "4401", 5000, 800, 3500, 28])
    sheet.merge_cells("A2:A3")
    sheet.merge_cells("B2:B3")
    sheet.merge_cells("C2:D2")
    sheet.merge_cells("E2:F2")
    workbook.save(workbook_path)

    extraction = extract_header_structure(workbook_path)

    assert extraction.header_rows == [2, 3]
    assert extraction.data_start_row == 4
    assert [column.signature for column in extraction.columns] == [
        "姓名",
        "证件号码",
        "基本养老保险 / 缴费基数",
        "基本养老保险 / 应缴金额",
        "失业保险 / 缴费基数",
        "失业保险 / 应缴金额",
    ]
    assert extraction.header_tree[0].label == "姓名"
    assert extraction.header_tree[2].label == "基本养老保险"
    assert [child.label for child in extraction.header_tree[2].children] == ["缴费基数", "应缴金额"]


def test_extract_header_structure_on_real_guangzhou_sample() -> None:
    sample_path = find_sample("广分")

    extraction = extract_header_structure(sample_path)

    signatures = [column.signature for column in extraction.columns[:12]]
    assert extraction.sheet_name == "sheet1"
    assert extraction.header_rows == [7, 8]
    assert extraction.data_start_row == 9
    assert signatures[:7] == [
        "序号",
        "姓名",
        "证件号码",
        "证件类型",
        "个人社保号",
        "费款所属期起",
        "费款所属期止",
    ]
    assert "基本养老保险(单位缴纳) / 缴费基数" in signatures
    assert "基本养老保险(单位缴纳) / 应缴金额" in signatures


def test_extract_header_structure_on_real_xiamen_sample() -> None:
    sample_path = find_sample("厦门202602社保账单.xlsx")

    extraction = extract_header_structure(sample_path)

    signatures = [column.signature for column in extraction.columns[:20]]
    assert extraction.sheet_name == "职工社保对账单明细查询"
    assert extraction.header_rows == [3, 4]
    assert extraction.data_start_row == 5
    assert signatures[:9] == [
        "证件号码",
        "姓名",
        "参保人员身份",
        "总金额",
        "单位缴费总金额",
        "个人缴费总金额",
        "建账年月",
        "费款所属期起",
        "费款所属期止",
    ]
    assert "城镇企业职工基本养老保险费 / 缴费工资" in signatures
    assert "城镇企业职工基本养老保险费 / 单位应缴" in signatures


def test_extract_header_structure_on_real_shenzhen_sample() -> None:
    sample_path = find_sample("深圳创造欢乐")

    extraction = extract_header_structure(sample_path)

    signatures = [column.signature for column in extraction.columns[:20]]
    assert extraction.sheet_name == "申报明细"
    assert extraction.header_rows == [1, 2]
    assert extraction.data_start_row == 4
    assert signatures[:8] == [
        "序号",
        "姓名",
        "证件号码",
        "费款所属期起",
        "费款所属期止",
        "应收金额",
        "个人社保合计",
        "单位社保合计",
    ]
    assert "基本养老保险（单位） / 费率" in signatures
    assert "基本养老保险（单位） / 应缴费额" in signatures


def test_extract_header_structure_on_real_wuhan_sample() -> None:
    sample_path = find_sample("武汉")

    extraction = extract_header_structure(sample_path)

    signatures = [column.signature for column in extraction.columns[:20]]
    assert extraction.sheet_name == "Sheet1"
    assert extraction.data_start_row == 3
    assert "职工明细 / 姓名" in signatures or "姓名" in signatures
    assert any("养老保险应缴费额" in signature for signature in signatures)