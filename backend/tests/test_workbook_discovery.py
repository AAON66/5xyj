from __future__ import annotations

import shutil
from pathlib import Path

import pytest
from openpyxl import Workbook

from backend.app.core.config import ROOT_DIR
from backend.app.parsers import discover_workbook


ARTIFACTS_ROOT = ROOT_DIR / ".test_artifacts" / "workbook_discovery"
SAMPLES_DIR = ROOT_DIR / "data" / "samples"


def create_artifact_dir(test_name: str) -> Path:
    target = ARTIFACTS_ROOT / test_name
    if target.exists():
        shutil.rmtree(target)
    target.mkdir(parents=True, exist_ok=True)
    return target


def find_sample(keyword: str) -> Path:
    for path in sorted(SAMPLES_DIR.glob("*.xlsx")):
        if keyword in path.name:
            return path
    pytest.skip(f"Sample containing '{keyword}' was not found in {SAMPLES_DIR}.")


def test_discover_workbook_prefers_sheet_with_header_and_data_region() -> None:
    artifacts_dir = create_artifact_dir("guangzhou_like")
    workbook_path = artifacts_dir / "guangzhou_like.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "申报明细"
    sheet.append(["2026年2月社会保险费申报个人明细表"])
    sheet.append([])
    sheet.append(["单位名称", "广分"])
    sheet.append(["费款所属期", "2026-02"])
    sheet.append([])
    sheet.append(["姓名", "证件号码", "个人社保号", "基本养老保险", "基本医疗保险（含生育）"])
    sheet.append(["", "", "", "单位缴纳", "单位缴纳"])
    sheet.append(["张三", "440101199001010011", "SS001", 120.5, 88])
    sheet.append(["李四", "440101199202020022", "SS002", 121, 89])
    workbook.save(workbook_path)

    discovery = discover_workbook(workbook_path)

    assert discovery.selected_sheet_name == "申报明细"
    assert discovery.selected_header_row_candidates == [6, 7]
    assert discovery.selected_data_start_row == 8
    assert discovery.discoveries[0].score > 0
    assert discovery.failure_reason is None


def test_discover_workbook_finds_valid_sheet_when_first_sheet_is_noise() -> None:
    artifacts_dir = create_artifact_dir("changsha_like")
    workbook_path = artifacts_dir / "changsha_like.xlsx"

    workbook = Workbook()
    workbook.active.title = "Sheet1"
    workbook["Sheet1"].append(["说明"])
    workbook.create_sheet("Sheet2").append(["封面"])
    workbook.create_sheet("Sheet3").append(["统计"])
    target_sheet = workbook.create_sheet("Sheet4")
    target_sheet.append(["长沙202602社保账单"])
    target_sheet.append([])
    target_sheet.append(["姓名", "工伤保险", "失业保险(单位缴纳)", "职工基本养老保险(个人缴纳)", "总计"])
    target_sheet.append(["王五", 16, 22, 140, 178])
    target_sheet.append(["赵六", 18, 21, 142, 181])
    workbook.save(workbook_path)

    discovery = discover_workbook(workbook_path)

    assert discovery.selected_sheet_name == "Sheet4"
    assert discovery.selected_header_row_candidates == [3]
    assert discovery.selected_data_start_row == 4
    assert discovery.discoveries[0].sheet_name == "Sheet4"
    assert any("Detected data start row" in reason for reason in discovery.discoveries[0].reasoning)


def test_discover_workbook_returns_failure_for_empty_workbook() -> None:
    artifacts_dir = create_artifact_dir("empty_workbook")
    workbook_path = artifacts_dir / "empty_workbook.xlsx"

    workbook = Workbook()
    workbook.active.title = "Sheet1"
    workbook.save(workbook_path)

    discovery = discover_workbook(workbook_path)

    assert discovery.selected_sheet_name is None
    assert discovery.failure_reason == "No valid worksheet candidate was detected."
    assert discovery.discoveries[0].score == 0


def test_discover_workbook_on_real_guangzhou_sample() -> None:
    sample_path = find_sample("广分")

    discovery = discover_workbook(sample_path)

    assert discovery.selected_sheet_name == "sheet1"
    assert discovery.selected_header_row_candidates == [8]
    assert discovery.selected_data_start_row == 9
    assert discovery.failure_reason is None


def test_discover_workbook_on_real_changsha_sample() -> None:
    sample_path = find_sample("长沙")

    discovery = discover_workbook(sample_path)

    assert discovery.selected_sheet_name == "Sheet4"
    assert discovery.selected_header_row_candidates == [4]
    assert discovery.selected_data_start_row == 5
    assert discovery.discoveries[0].sheet_name == "Sheet4"