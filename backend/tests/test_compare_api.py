from __future__ import annotations

from decimal import Decimal
from io import BytesIO
import shutil

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from backend.app.core.config import ROOT_DIR, Settings
from backend.app.core.database import create_database_engine, create_session_factory
from backend.app.dependencies import get_db
from backend.app.main import create_app
from backend.app.models import Base, ImportBatch, NormalizedRecord, SourceFile
from backend.app.models.enums import BatchStatus, SourceFileKind


ARTIFACTS_ROOT = ROOT_DIR / '.test_artifacts' / 'compare_api'


def build_test_context(test_name: str):
    artifacts_dir = ARTIFACTS_ROOT / test_name
    if artifacts_dir.exists():
        shutil.rmtree(artifacts_dir)
    artifacts_dir.mkdir(parents=True, exist_ok=True)

    database_path = artifacts_dir / 'compare.db'
    settings = Settings(
        app_name='对比测试',
        app_version='0.2.0',
        auth_enabled=False,
        database_url=f'sqlite:///{database_path.as_posix()}',
        upload_dir=str(artifacts_dir / 'uploads'),
        samples_dir=str(artifacts_dir / 'samples'),
        templates_dir=str(artifacts_dir / 'templates'),
        outputs_dir=str(artifacts_dir / 'outputs'),
        log_format='plain',
    )

    engine = create_database_engine(settings)
    session_factory = create_session_factory(settings)
    Base.metadata.create_all(engine)

    app = create_app(settings)

    def override_get_db():
        db: Session = session_factory()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_get_db
    return TestClient(app), session_factory


def seed_compare_batches(session_factory) -> tuple[str, str]:
    db: Session = session_factory()
    try:
        left_batch = ImportBatch(batch_name='2026-03 新融合', status=BatchStatus.NORMALIZED)
        right_batch = ImportBatch(batch_name='2026-02 上月归档', status=BatchStatus.NORMALIZED)
        db.add_all([left_batch, right_batch])
        db.flush()

        left_source = SourceFile(
            batch_id=left_batch.id,
            file_name='left.xlsx',
            file_path='D:/left.xlsx',
            file_size=1,
            source_kind=SourceFileKind.SOCIAL_SECURITY.value,
            region='guangzhou',
            company_name='示例公司',
        )
        right_source = SourceFile(
            batch_id=right_batch.id,
            file_name='right.xlsx',
            file_path='D:/right.xlsx',
            file_size=1,
            source_kind=SourceFileKind.SOCIAL_SECURITY.value,
            region='guangzhou',
            company_name='示例公司',
        )
        db.add_all([left_source, right_source])
        db.flush()

        db.add_all(
            [
                NormalizedRecord(
                    batch_id=left_batch.id,
                    source_file_id=left_source.id,
                    source_row_number=2,
                    person_name='张三',
                    employee_id='E001',
                    id_number='4401',
                    company_name='示例公司',
                    billing_period='2026-03',
                    total_amount=Decimal('100.00'),
                    source_file_name='left.xlsx',
                ),
                NormalizedRecord(
                    batch_id=right_batch.id,
                    source_file_id=right_source.id,
                    source_row_number=2,
                    person_name='张三',
                    employee_id='E001',
                    id_number='4401',
                    company_name='示例公司',
                    billing_period='2026-02',
                    total_amount=Decimal('80.00'),
                    source_file_name='right.xlsx',
                ),
                NormalizedRecord(
                    batch_id=left_batch.id,
                    source_file_id=left_source.id,
                    source_row_number=3,
                    person_name='李四',
                    employee_id='E002',
                    id_number='4402',
                    company_name='示例公司',
                    billing_period='2026-03',
                    total_amount=Decimal('90.00'),
                    source_file_name='left.xlsx',
                ),
                NormalizedRecord(
                    batch_id=right_batch.id,
                    source_file_id=right_source.id,
                    source_row_number=4,
                    person_name='王五',
                    employee_id='E003',
                    id_number='4403',
                    company_name='示例公司',
                    billing_period='2026-02',
                    total_amount=Decimal('70.00'),
                    source_file_name='right.xlsx',
                ),
            ]
        )
        db.commit()
        return left_batch.id, right_batch.id
    finally:
        db.close()


def test_compare_endpoint_returns_changed_and_missing_rows() -> None:
    client, session_factory = build_test_context('compare_rows')
    left_batch_id, right_batch_id = seed_compare_batches(session_factory)

    with client:
        response = client.get('/api/v1/compare', params={'left_batch_id': left_batch_id, 'right_batch_id': right_batch_id})

    assert response.status_code == 200
    payload = response.json()['data']
    assert payload['left_batch']['batch_name'] == '2026-03 新融合'
    assert payload['right_batch']['batch_name'] == '2026-02 上月归档'
    assert payload['total_row_count'] == 3
    assert payload['changed_row_count'] == 1
    assert payload['left_only_count'] == 1
    assert payload['right_only_count'] == 1

    rows = {row['compare_key']: row for row in payload['rows']}
    changed_row = next(row for row in rows.values() if row['diff_status'] == 'changed')
    assert 'total_amount' in changed_row['different_fields']
    assert changed_row['left']['values']['total_amount'] == '100.00'
    assert changed_row['right']['values']['total_amount'] == '80.00'

    assert any(row['diff_status'] == 'left_only' for row in rows.values())
    assert any(row['diff_status'] == 'right_only' for row in rows.values())


def test_compare_export_endpoint_returns_workbook_with_difference_highlights() -> None:
    client, session_factory = build_test_context('compare_export')
    left_batch_id, right_batch_id = seed_compare_batches(session_factory)

    with client:
        compare_response = client.get('/api/v1/compare', params={'left_batch_id': left_batch_id, 'right_batch_id': right_batch_id})
        payload = compare_response.json()['data']
        export_response = client.post(
            '/api/v1/compare/export',
            json={
                'left_batch_name': payload['left_batch']['batch_name'],
                'right_batch_name': payload['right_batch']['batch_name'],
                'fields': payload['fields'],
                'rows': payload['rows'],
            },
        )

    assert export_response.status_code == 200
    assert 'attachment; filename=' in export_response.headers['content-disposition']

    workbook = load_workbook(filename=BytesIO(export_response.content))
    assert workbook.sheetnames == ['差异概览', '左侧数据', '右侧数据']
    left_sheet = workbook['左侧数据']
    right_sheet = workbook['右侧数据']
    assert left_sheet['A2'].value
    assert right_sheet['A2'].value
    assert left_sheet['B2'].value == 'changed'


def test_compare_endpoint_returns_not_found_for_unknown_batch() -> None:
    client, _session_factory = build_test_context('compare_missing')

    with client:
        response = client.get('/api/v1/compare', params={'left_batch_id': 'missing', 'right_batch_id': 'missing-two'})

    assert response.status_code == 404
    assert response.json()['error']['code'] == 'not_found'
