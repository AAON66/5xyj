from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from backend.app.core.config import ROOT_DIR, get_settings
from backend.app.exporters import export_dual_templates
from backend.app.services import build_normalized_models, standardize_workbook


ARTIFACTS_ROOT = ROOT_DIR / '.test_artifacts' / 'template_exporter'
DESKTOP_ROOT = Path.home() / 'Desktop' / '202602社保公积金台账' / '202602社保公积金汇总'
SOCIAL_AMOUNT_FIELDS = (
    'total_amount',
    'company_total_amount',
    'personal_total_amount',
    'pension_company',
    'pension_personal',
    'medical_company',
    'medical_personal',
    'medical_maternity_company',
    'maternity_amount',
    'unemployment_company',
    'unemployment_personal',
    'injury_company',
    'supplementary_medical_company',
    'supplementary_pension_company',
    'large_medical_personal',
    'late_fee',
    'interest',
)


def find_template(keyword: str) -> Path:
    settings = get_settings()
    configured = [settings.salary_template_file, settings.final_tool_template_file]
    for path in configured:
        if path is not None and path.exists() and keyword in path.name:
            return path
    if DESKTOP_ROOT.exists():
        for path in sorted(DESKTOP_ROOT.glob('*.xlsx')):
            if keyword in path.name:
                return path
    pytest.skip(f'Template containing {keyword!r} was not found in configured paths or {DESKTOP_ROOT}.')


def find_sample(keyword: str) -> Path:
    samples_dir = ROOT_DIR / 'data' / 'samples'
    for path in sorted(samples_dir.glob('*.xlsx')):
        if keyword in path.name:
            return path
    pytest.skip(f'Sample containing {keyword!r} was not found in {samples_dir}.')


def clear_social_amounts(record) -> None:
    for field_name in SOCIAL_AMOUNT_FIELDS:
        setattr(record, field_name, Decimal('0'))


def test_export_dual_templates_writes_both_template_outputs() -> None:
    salary_template = find_template('薪酬')
    tool_template = find_template('最终版')
    sample_path = find_sample('深圳创造欢乐')

    standardized = standardize_workbook(sample_path, region='shenzhen', company_name='创造欢乐')
    trimmed = type(standardized)(
        source_file=standardized.source_file,
        sheet_name=standardized.sheet_name,
        raw_header_signature=standardized.raw_header_signature,
        records=standardized.records[:2],
        filtered_rows=standardized.filtered_rows,
        unmapped_headers=standardized.unmapped_headers,
    )
    records = build_normalized_models(trimmed, batch_id='batch-1', source_file_id='source-1')
    records[0].employee_id = '01620'
    records[1].employee_id = '01831'
    records[0].housing_fund_personal = records[0].housing_fund_personal or records[0].personal_total_amount
    records[0].housing_fund_company = records[0].housing_fund_company or records[0].personal_total_amount
    records[0].housing_fund_total = records[0].housing_fund_personal + records[0].housing_fund_company

    output_dir = ARTIFACTS_ROOT / 'successful_export'
    output_dir.mkdir(parents=True, exist_ok=True)

    result = export_dual_templates(
        records,
        output_dir=output_dir,
        salary_template_path=salary_template,
        final_tool_template_path=tool_template,
        export_prefix='match_batch',
    )

    assert result.status == 'completed'
    salary_artifact = next(item for item in result.artifacts if item.template_type == 'salary')
    tool_artifact = next(item for item in result.artifacts if item.template_type == 'final_tool')
    assert salary_artifact.status == 'completed'
    assert tool_artifact.status == 'completed'
    assert Path(salary_artifact.file_path).name == 'match_batch_salary.xlsx'
    assert Path(tool_artifact.file_path).name == 'match_batch_final_tool.xlsx'

    salary_wb = load_workbook(salary_artifact.file_path, data_only=False)
    salary_sheet = salary_wb[salary_wb.sheetnames[0]]
    assert salary_sheet['A2'].value == records[0].person_name
    assert salary_sheet['B2'].value == '01620'
    assert float(salary_sheet['C2'].value) == float(records[0].medical_personal)
    assert float(salary_sheet['I2'].value) == float(records[0].pension_company + records[0].supplementary_pension_company)
    assert float(salary_sheet['H2'].value) == float(records[0].housing_fund_personal)
    assert float(salary_sheet['P2'].value) == float(records[0].housing_fund_company)
    assert float(salary_sheet['Q2'].value) == 0.0
    salary_wb.close()

    tool_wb = load_workbook(tool_artifact.file_path, data_only=False)
    tool_sheet = tool_wb[tool_wb.sheetnames[0]]
    assert tool_sheet['A7'].value == '创造欢乐'
    assert tool_sheet['B7'].value == '深圳'
    assert tool_sheet['D7'].value == records[0].id_number
    assert tool_sheet['E7'].value == '01620'
    assert float(tool_sheet['O7'].value) == float(records[0].pension_company + records[0].supplementary_pension_company)
    assert float(tool_sheet['N7'].value) == float(records[0].housing_fund_personal)
    assert float(tool_sheet['V7'].value) == float(records[0].housing_fund_company)
    assert float(tool_sheet['W7'].value) == 0.0
    assert tool_sheet['AA7'].data_type == 'f'
    assert str(tool_sheet['AA7'].value).startswith('=')
    assert tool_sheet['AO7'].data_type == 'f'
    assert str(tool_sheet['AO7'].value).startswith('=')
    tool_wb.close()


def test_export_dual_templates_marks_overall_failure_when_any_template_is_missing() -> None:
    salary_template = find_template('薪酬')
    sample_path = find_sample('深圳创造欢乐')
    standardized = standardize_workbook(sample_path, region='shenzhen', company_name='创造欢乐')
    trimmed = type(standardized)(
        source_file=standardized.source_file,
        sheet_name=standardized.sheet_name,
        raw_header_signature=standardized.raw_header_signature,
        records=standardized.records[:1],
        filtered_rows=standardized.filtered_rows,
        unmapped_headers=standardized.unmapped_headers,
    )
    records = build_normalized_models(trimmed, batch_id='batch-1', source_file_id='source-1')
    records[0].employee_id = '01620'

    output_dir = ARTIFACTS_ROOT / 'failed_export'
    output_dir.mkdir(parents=True, exist_ok=True)

    result = export_dual_templates(
        records,
        output_dir=output_dir,
        salary_template_path=salary_template,
        final_tool_template_path=output_dir / 'missing.xlsx',
        export_prefix='missing_template',
    )

    assert result.status == 'failed'
    assert next(item for item in result.artifacts if item.template_type == 'salary').status == 'completed'
    failed_tool = next(item for item in result.artifacts if item.template_type == 'final_tool')
    assert failed_tool.status == 'failed'
    assert failed_tool.error_message


def test_export_dual_templates_filters_header_like_dirty_rows() -> None:
    salary_template = find_template('薪酬')
    tool_template = find_template('最终版')
    sample_path = find_sample('深圳创造欢乐')

    standardized = standardize_workbook(sample_path, region='shenzhen', company_name='创造欢乐')
    trimmed = type(standardized)(
        source_file=standardized.source_file,
        sheet_name=standardized.sheet_name,
        raw_header_signature=standardized.raw_header_signature,
        records=standardized.records[:1],
        filtered_rows=standardized.filtered_rows,
        unmapped_headers=standardized.unmapped_headers,
    )
    records = build_normalized_models(trimmed, batch_id='batch-1', source_file_id='source-1')
    records[0].employee_id = '01620'

    dirty = build_normalized_models(trimmed, batch_id='batch-1', source_file_id='source-1')[0]
    dirty.person_name = '姓名'
    dirty.id_number = '身份证号码'
    dirty.employee_id = None
    dirty.personal_total_amount = None
    dirty.housing_fund_personal = None
    records.append(dirty)

    output_dir = ARTIFACTS_ROOT / 'dirty_row_filter'
    output_dir.mkdir(parents=True, exist_ok=True)

    result = export_dual_templates(
        records,
        output_dir=output_dir,
        salary_template_path=salary_template,
        final_tool_template_path=tool_template,
        export_prefix='dirty_filter',
    )

    salary_artifact = next(item for item in result.artifacts if item.template_type == 'salary')
    tool_artifact = next(item for item in result.artifacts if item.template_type == 'final_tool')
    assert salary_artifact.row_count == 1
    assert tool_artifact.row_count == 1

    salary_wb = load_workbook(salary_artifact.file_path, data_only=False)
    salary_sheet = salary_wb[salary_wb.sheetnames[0]]
    assert salary_sheet['A2'].value == records[0].person_name
    assert salary_sheet['A3'].value in (None, '')
    salary_wb.close()

    tool_wb = load_workbook(tool_artifact.file_path, data_only=False)
    tool_sheet = tool_wb[tool_wb.sheetnames[0]]
    assert tool_sheet['C7'].value == records[0].person_name
    assert tool_sheet['C8'].value in (None, '')
    tool_wb.close()


def test_export_dual_templates_filters_zero_amount_rows_even_when_identity_exists() -> None:
    salary_template = find_template('薪酬')
    tool_template = find_template('最终版')
    sample_path = find_sample('深圳创造欢乐')

    standardized = standardize_workbook(sample_path, region='shenzhen', company_name='创造欢乐')
    trimmed = type(standardized)(
        source_file=standardized.source_file,
        sheet_name=standardized.sheet_name,
        raw_header_signature=standardized.raw_header_signature,
        records=standardized.records[:1],
        filtered_rows=standardized.filtered_rows,
        unmapped_headers=standardized.unmapped_headers,
    )
    zero_amount = build_normalized_models(trimmed, batch_id='batch-1', source_file_id='source-1')[0]
    zero_amount.person_name = '离职空行'
    zero_amount.id_number = '440100199001010011'
    zero_amount.employee_id = 'E0001'
    for field_name in (
        'housing_fund_personal',
        'housing_fund_company',
        'housing_fund_total',
        'total_amount',
        'company_total_amount',
        'personal_total_amount',
        'pension_company',
        'pension_personal',
        'medical_company',
        'medical_personal',
        'medical_maternity_company',
        'maternity_amount',
        'unemployment_company',
        'unemployment_personal',
        'injury_company',
        'supplementary_medical_company',
        'supplementary_pension_company',
        'large_medical_personal',
        'late_fee',
        'interest',
    ):
        setattr(zero_amount, field_name, Decimal('0'))

    output_dir = ARTIFACTS_ROOT / 'zero_amount_filter'
    output_dir.mkdir(parents=True, exist_ok=True)

    result = export_dual_templates(
        [zero_amount],
        output_dir=output_dir,
        salary_template_path=salary_template,
        final_tool_template_path=tool_template,
        export_prefix='zero_amount_filter',
    )

    salary_artifact = next(item for item in result.artifacts if item.template_type == 'salary')
    tool_artifact = next(item for item in result.artifacts if item.template_type == 'final_tool')
    assert salary_artifact.row_count == 0
    assert tool_artifact.row_count == 0

    salary_wb = load_workbook(salary_artifact.file_path, data_only=False)
    salary_sheet = salary_wb[salary_wb.sheetnames[0]]
    assert salary_sheet['A2'].value in (None, '')
    salary_wb.close()

    tool_wb = load_workbook(tool_artifact.file_path, data_only=False)
    tool_sheet = tool_wb[tool_wb.sheetnames[0]]
    assert tool_sheet['C7'].value in (None, '')
    tool_wb.close()


def test_export_dual_templates_filters_inferred_housing_only_when_mixed_with_explicit_split_rows() -> None:
    salary_template = find_template('薪酬')
    tool_template = find_template('最终版')
    sample_path = find_sample('深圳创造欢乐')

    standardized = standardize_workbook(sample_path, region='shenzhen', company_name='创造欢乐')
    trimmed = type(standardized)(
        source_file=standardized.source_file,
        sheet_name=standardized.sheet_name,
        raw_header_signature=standardized.raw_header_signature,
        records=standardized.records[:1],
        filtered_rows=standardized.filtered_rows,
        unmapped_headers=standardized.unmapped_headers,
    )

    inferred_only = build_normalized_models(trimmed, batch_id='batch-1', source_file_id='source-1')[0]
    inferred_only.person_name = '推导拆分'
    inferred_only.employee_id = 'E0002'
    inferred_only.id_number = '440100199001010022'
    clear_social_amounts(inferred_only)
    inferred_only.housing_fund_personal = Decimal('175')
    inferred_only.housing_fund_company = Decimal('175')
    inferred_only.housing_fund_total = Decimal('350')
    inferred_only.raw_payload = {
        'merged_sources': [
            {
                'source_kind': 'housing_fund',
                'raw_values': {
                    '姓名': '推导拆分',
                    '证件号码': '440100199001010022',
                    '缴存基数（元）': '3500',
                    '单位缴存比例': '0.05',
                    '个人缴存比例': '0.05',
                    '金额合计（元）': '350',
                },
            }
        ]
    }

    explicit_split = build_normalized_models(trimmed, batch_id='batch-1', source_file_id='source-2')[0]
    explicit_split.person_name = '显式拆分'
    explicit_split.employee_id = 'E0003'
    explicit_split.id_number = '440100199001010033'
    clear_social_amounts(explicit_split)
    explicit_split.housing_fund_personal = Decimal('175')
    explicit_split.housing_fund_company = Decimal('175')
    explicit_split.housing_fund_total = Decimal('350')
    explicit_split.raw_payload = {
        'merged_sources': [
            {
                'source_kind': 'housing_fund',
                'raw_values': {
                    '姓名': '显式拆分',
                    '证件号码': '440100199001010033',
                    '单位': 175,
                    '个人': 175,
                    '金额合计（元）': '350',
                },
            }
        ]
    }

    output_dir = ARTIFACTS_ROOT / 'housing_only_inference_filter'
    output_dir.mkdir(parents=True, exist_ok=True)

    result = export_dual_templates(
        [inferred_only, explicit_split],
        output_dir=output_dir,
        salary_template_path=salary_template,
        final_tool_template_path=tool_template,
        export_prefix='housing_only_inference_filter',
    )

    salary_artifact = next(item for item in result.artifacts if item.template_type == 'salary')
    tool_artifact = next(item for item in result.artifacts if item.template_type == 'final_tool')
    assert salary_artifact.row_count == 1
    assert tool_artifact.row_count == 1

    salary_wb = load_workbook(salary_artifact.file_path, data_only=False)
    salary_sheet = salary_wb[salary_wb.sheetnames[0]]
    assert salary_sheet['A2'].value == '显式拆分'
    assert salary_sheet['A3'].value in (None, '')
    salary_wb.close()

    tool_wb = load_workbook(tool_artifact.file_path, data_only=False)
    tool_sheet = tool_wb[tool_wb.sheetnames[0]]
    assert tool_sheet['C7'].value == '显式拆分'
    assert tool_sheet['C8'].value in (None, '')
    tool_wb.close()


def test_export_dual_templates_merges_records_with_same_employee_id_before_writing() -> None:
    salary_template = find_template('薪酬')
    tool_template = find_template('最终版')
    sample_path = find_sample('深圳创造欢乐')

    standardized = standardize_workbook(sample_path, region='shenzhen', company_name='创造欢乐')
    trimmed = type(standardized)(
        source_file=standardized.source_file,
        sheet_name=standardized.sheet_name,
        raw_header_signature=standardized.raw_header_signature,
        records=standardized.records[:1],
        filtered_rows=standardized.filtered_rows,
        unmapped_headers=standardized.unmapped_headers,
    )
    social_record = build_normalized_models(trimmed, batch_id='batch-1', source_file_id='source-1')[0]
    social_record.employee_id = '01620'
    social_record.housing_fund_personal = None
    social_record.housing_fund_company = None
    social_record.housing_fund_total = None

    housing_record = build_normalized_models(trimmed, batch_id='batch-1', source_file_id='source-2')[0]
    housing_record.employee_id = '01620'
    housing_record.medical_personal = None
    housing_record.unemployment_personal = None
    housing_record.large_medical_personal = None
    housing_record.pension_personal = None
    housing_record.pension_company = None
    housing_record.supplementary_pension_company = None
    housing_record.medical_company = None
    housing_record.medical_maternity_company = None
    housing_record.unemployment_company = None
    housing_record.injury_company = None
    housing_record.maternity_amount = None
    housing_record.supplementary_medical_company = None
    housing_record.personal_total_amount = None
    housing_record.company_total_amount = None
    housing_record.total_amount = None
    housing_record.housing_fund_personal = Decimal('500')
    housing_record.housing_fund_company = Decimal('500')
    housing_record.housing_fund_total = Decimal('1000')
    housing_record.raw_payload = {
        'merged_sources': [
            {
                'source_kind': 'housing_fund',
                'source_file_name': '深圳创造欢乐202602公积金账单.xlsx',
                'source_row_number': 12,
                'raw_values': {
                    '姓名': housing_record.person_name,
                    '证件号码': housing_record.id_number,
                    '单位': 500,
                    '个人': 500,
                    '金额合计（元）': '1000',
                },
            }
        ]
    }

    output_dir = ARTIFACTS_ROOT / 'merge_same_employee'
    output_dir.mkdir(parents=True, exist_ok=True)

    result = export_dual_templates(
        [social_record, housing_record],
        output_dir=output_dir,
        salary_template_path=salary_template,
        final_tool_template_path=tool_template,
        export_prefix='merge_same_employee',
    )

    salary_artifact = next(item for item in result.artifacts if item.template_type == 'salary')
    tool_artifact = next(item for item in result.artifacts if item.template_type == 'final_tool')
    assert salary_artifact.row_count == 1
    assert tool_artifact.row_count == 1

    salary_wb = load_workbook(salary_artifact.file_path, data_only=False)
    salary_sheet = salary_wb[salary_wb.sheetnames[0]]
    assert salary_sheet['A2'].value == social_record.person_name
    assert salary_sheet['B2'].value == '01620'
    assert float(salary_sheet['C2'].value) == float(social_record.medical_personal)
    assert float(salary_sheet['H2'].value) == 500.0
    assert float(salary_sheet['P2'].value) == 500.0
    assert salary_sheet['A3'].value in (None, '')
    salary_wb.close()

    tool_wb = load_workbook(tool_artifact.file_path, data_only=False)
    tool_sheet = tool_wb[tool_wb.sheetnames[0]]
    assert tool_sheet['C7'].value == social_record.person_name
    assert tool_sheet['E7'].value == '01620'
    assert float(tool_sheet['I7'].value) == float(social_record.medical_personal)
    assert float(tool_sheet['N7'].value) == 500.0
    assert float(tool_sheet['V7'].value) == 500.0
    assert tool_sheet['C8'].value in (None, '')
    tool_wb.close()


def test_export_dual_templates_derives_housing_burden_from_repeated_source_baseline() -> None:
    salary_template = find_template('薪酬')
    tool_template = find_template('最终版')
    sample_path = find_sample('深圳创造欢乐')

    standardized = standardize_workbook(sample_path, region='shenzhen', company_name='创造欢乐')
    trimmed = type(standardized)(
        source_file=standardized.source_file,
        sheet_name=standardized.sheet_name,
        raw_header_signature=standardized.raw_header_signature,
        records=standardized.records[:1],
        filtered_rows=standardized.filtered_rows,
        unmapped_headers=standardized.unmapped_headers,
    )
    seed = build_normalized_models(trimmed, batch_id='batch-1', source_file_id='source-1')[0]

    def build_record(employee_id: str, *, person_name: str, personal_housing: Decimal, company_housing: Decimal):
        record = build_normalized_models(trimmed, batch_id='batch-1', source_file_id='source-1')[0]
        record.person_name = person_name
        record.employee_id = employee_id
        record.id_number = f'44010019900101{employee_id[-4:]}'
        record.personal_total_amount = Decimal('0')
        record.housing_fund_personal = personal_housing
        record.housing_fund_company = company_housing
        record.housing_fund_total = personal_housing + company_housing
        record.raw_payload = {
            'merged_sources': [
                {
                    'source_kind': 'housing_fund',
                    'source_file_name': '广州公积金账单.xlsx',
                    'source_row_number': 10,
                }
            ]
        }
        return record

    standard_a = build_record('E1001', person_name='标准甲', personal_housing=Decimal('175'), company_housing=Decimal('175'))
    standard_b = build_record('E1002', person_name='标准乙', personal_housing=Decimal('175'), company_housing=Decimal('175'))
    target = build_record('E1003', person_name=seed.person_name or '目标员工', personal_housing=Decimal('2340'), company_housing=Decimal('975'))

    output_dir = ARTIFACTS_ROOT / 'housing_burden_baseline'
    output_dir.mkdir(parents=True, exist_ok=True)

    result = export_dual_templates(
        [standard_a, standard_b, target],
        output_dir=output_dir,
        salary_template_path=salary_template,
        final_tool_template_path=tool_template,
        export_prefix='housing_burden_baseline',
    )

    salary_artifact = next(item for item in result.artifacts if item.template_type == 'salary')
    tool_artifact = next(item for item in result.artifacts if item.template_type == 'final_tool')

    salary_wb = load_workbook(salary_artifact.file_path, data_only=False)
    salary_sheet = salary_wb[salary_wb.sheetnames[0]]
    assert float(salary_sheet['H4'].value) == 2340.0
    assert float(salary_sheet['P4'].value) == 975.0
    assert float(salary_sheet['R4'].value) == 800.0
    salary_wb.close()

    tool_wb = load_workbook(tool_artifact.file_path, data_only=False)
    tool_sheet = tool_wb[tool_wb.sheetnames[0]]
    assert float(tool_sheet['N9'].value) == 2340.0
    assert float(tool_sheet['V9'].value) == 975.0
    assert float(tool_sheet['X9'].value) == 800.0
    tool_wb.close()


def test_export_dual_templates_defaults_social_burden_to_zero_without_explicit_rule() -> None:
    salary_template = find_template('薪酬')
    tool_template = find_template('最终版')
    sample_path = find_sample('深圳创造欢乐')

    standardized = standardize_workbook(sample_path, region='shenzhen', company_name='创造欢乐')
    trimmed = type(standardized)(
        source_file=standardized.source_file,
        sheet_name=standardized.sheet_name,
        raw_header_signature=standardized.raw_header_signature,
        records=standardized.records[:1],
        filtered_rows=standardized.filtered_rows,
        unmapped_headers=standardized.unmapped_headers,
    )

    def build_record(
        employee_id: str,
        *,
        person_name: str,
        medical_company: Decimal,
        medical_personal: Decimal = Decimal('134.54'),
        unemployment_personal: Decimal = Decimal('7'),
        pension_personal: Decimal = Decimal('382'),
    ):
        record = build_normalized_models(trimmed, batch_id='batch-1', source_file_id='source-1')[0]
        record.person_name = person_name
        record.employee_id = employee_id
        record.id_number = f'44010019900101{employee_id[-4:]}'
        record.medical_company = medical_company
        record.medical_personal = medical_personal
        record.unemployment_personal = unemployment_personal
        record.pension_personal = pension_personal
        record.personal_total_amount = medical_personal + unemployment_personal + pension_personal
        record.raw_payload = {
            'merged_sources': [
                {
                    'source_kind': 'social_security',
                    'source_file_name': '深圳社保账单.xlsx',
                    'source_row_number': 12,
                }
            ]
        }
        return record

    target = build_record('S1003', person_name='社保目标', medical_company=Decimal('403.62'))

    output_dir = ARTIFACTS_ROOT / 'social_burden_default_zero'
    output_dir.mkdir(parents=True, exist_ok=True)

    result = export_dual_templates(
        [target],
        output_dir=output_dir,
        salary_template_path=salary_template,
        final_tool_template_path=tool_template,
        export_prefix='social_burden_default_zero',
    )

    salary_artifact = next(item for item in result.artifacts if item.template_type == 'salary')
    tool_artifact = next(item for item in result.artifacts if item.template_type == 'final_tool')

    salary_wb = load_workbook(salary_artifact.file_path, data_only=False)
    salary_sheet = salary_wb[salary_wb.sheetnames[0]]
    assert float(salary_sheet['J2'].value) == 403.62
    assert float(salary_sheet['Q2'].value) == 0.0
    salary_wb.close()

    tool_wb = load_workbook(tool_artifact.file_path, data_only=False)
    tool_sheet = tool_wb[tool_wb.sheetnames[0]]
    assert float(tool_sheet['P7'].value) == 403.62
    assert float(tool_sheet['W7'].value) == 0.0
    assert tool_sheet['AA7'].data_type == 'f'
    tool_wb.close()


def test_export_dual_templates_sums_duplicate_social_records_from_distinct_sources() -> None:
    salary_template = find_template('薪酬')
    tool_template = find_template('最终版')
    sample_path = find_sample('深圳创造欢乐')

    standardized = standardize_workbook(sample_path, region='shenzhen', company_name='创造欢乐')
    trimmed = type(standardized)(
        source_file=standardized.source_file,
        sheet_name=standardized.sheet_name,
        raw_header_signature=standardized.raw_header_signature,
        records=standardized.records[:1],
        filtered_rows=standardized.filtered_rows,
        unmapped_headers=standardized.unmapped_headers,
    )

    base = build_normalized_models(trimmed, batch_id='batch-1', source_file_id='source-1')[0]
    base.person_name = '补缴样本'
    base.employee_id = 'S2001'
    base.id_number = '440100199001012001'
    base.region = 'xiamen'
    base.pension_personal = Decimal('323.44')
    base.pension_company = Decimal('646.88')
    base.supplementary_pension_company = Decimal('0')
    base.medical_personal = Decimal('88.66')
    base.medical_company = Decimal('332.48')
    base.unemployment_personal = Decimal('20.22')
    base.unemployment_company = Decimal('20.22')
    base.injury_company = Decimal('17.66')
    base.maternity_amount = Decimal('31.03')
    base.personal_total_amount = Decimal('432.32')
    base.company_total_amount = Decimal('1048.27')
    base.total_amount = Decimal('1480.59')
    base.raw_payload = {
        'merged_sources': [
            {
                'source_kind': 'social_security',
                'source_file_name': '厦门202602社保账单.xlsx',
                'source_row_number': 10,
            }
        ]
    }

    supplement = build_normalized_models(trimmed, batch_id='batch-1', source_file_id='source-2')[0]
    supplement.person_name = '补缴样本'
    supplement.employee_id = 'S2001'
    supplement.id_number = '440100199001012001'
    supplement.region = 'xiamen'
    supplement.pension_personal = Decimal('323.44')
    supplement.pension_company = Decimal('646.88')
    supplement.supplementary_pension_company = Decimal('0')
    supplement.medical_personal = Decimal('0')
    supplement.medical_company = Decimal('0')
    supplement.unemployment_personal = Decimal('20.22')
    supplement.unemployment_company = Decimal('20.22')
    supplement.injury_company = Decimal('17.66')
    supplement.maternity_amount = Decimal('0')
    supplement.personal_total_amount = Decimal('343.66')
    supplement.company_total_amount = Decimal('715.79')
    supplement.total_amount = Decimal('1059.45')
    supplement.raw_payload = {
        'merged_sources': [
            {
                'source_kind': 'social_security',
                'source_file_name': '厦门202602社保账单（补缴1月入职2人）.xlsx',
                'source_row_number': 11,
            }
        ]
    }

    output_dir = ARTIFACTS_ROOT / 'duplicate_social_sum'
    output_dir.mkdir(parents=True, exist_ok=True)

    result = export_dual_templates(
        [base, supplement],
        output_dir=output_dir,
        salary_template_path=salary_template,
        final_tool_template_path=tool_template,
        export_prefix='duplicate_social_sum',
    )

    salary_artifact = next(item for item in result.artifacts if item.template_type == 'salary')
    salary_wb = load_workbook(salary_artifact.file_path, data_only=False)
    salary_sheet = salary_wb[salary_wb.sheetnames[0]]
    assert float(salary_sheet['C2'].value) == 88.66
    assert float(salary_sheet['D2'].value) == 40.44
    assert float(salary_sheet['G2'].value) == 646.88
    assert float(salary_sheet['I2'].value) == 1293.76
    assert float(salary_sheet['J2'].value) == 363.51
    assert float(salary_sheet['K2'].value) == 40.44
    assert float(salary_sheet['L2'].value) == 35.32
    assert float(salary_sheet['M2'].value) == 31.03
    salary_wb.close()


def test_export_dual_templates_uses_highest_repeated_housing_baseline_for_large_amounts() -> None:
    salary_template = find_template('薪酬')
    tool_template = find_template('最终版')
    sample_path = find_sample('深圳创造欢乐')

    standardized = standardize_workbook(sample_path, region='shenzhen', company_name='创造欢乐')
    trimmed = type(standardized)(
        source_file=standardized.source_file,
        sheet_name=standardized.sheet_name,
        raw_header_signature=standardized.raw_header_signature,
        records=standardized.records[:1],
        filtered_rows=standardized.filtered_rows,
        unmapped_headers=standardized.unmapped_headers,
    )

    def build_record(employee_id: str, *, company_housing: Decimal):
        record = build_normalized_models(trimmed, batch_id='batch-1', source_file_id='source-1')[0]
        record.person_name = employee_id
        record.employee_id = employee_id
        record.id_number = f'44010019900101{employee_id[-4:]}'
        record.housing_fund_personal = company_housing
        record.housing_fund_company = company_housing
        record.housing_fund_total = company_housing * Decimal('2')
        record.raw_payload = {
            'merged_sources': [
                {
                    'source_kind': 'housing_fund',
                    'source_file_name': '深圳公积金账单.xlsx',
                    'source_row_number': int(employee_id[-1]) + 1,
                }
            ]
        }
        return record

    records = [
        build_record('H1001', company_housing=Decimal('118')),
        build_record('H1002', company_housing=Decimal('118')),
        build_record('H1003', company_housing=Decimal('175')),
        build_record('H1004', company_housing=Decimal('175')),
        build_record('H1005', company_housing=Decimal('1250')),
    ]

    output_dir = ARTIFACTS_ROOT / 'housing_burden_highest_baseline'
    output_dir.mkdir(parents=True, exist_ok=True)

    result = export_dual_templates(
        records,
        output_dir=output_dir,
        salary_template_path=salary_template,
        final_tool_template_path=tool_template,
        export_prefix='housing_burden_highest_baseline',
    )

    salary_artifact = next(item for item in result.artifacts if item.template_type == 'salary')
    salary_wb = load_workbook(salary_artifact.file_path, data_only=False)
    salary_sheet = salary_wb[salary_wb.sheetnames[0]]
    assert float(salary_sheet['R6'].value) == 1075.0
    salary_wb.close()


def test_export_dual_templates_derives_housing_amount_from_total_when_personal_value_is_ratio() -> None:
    salary_template = find_template('薪酬')
    tool_template = find_template('最终版')
    sample_path = find_sample('深圳创造欢乐')

    standardized = standardize_workbook(sample_path, region='shenzhen', company_name='创造欢乐')
    trimmed = type(standardized)(
        source_file=standardized.source_file,
        sheet_name=standardized.sheet_name,
        raw_header_signature=standardized.raw_header_signature,
        records=standardized.records[:1],
        filtered_rows=standardized.filtered_rows,
        unmapped_headers=standardized.unmapped_headers,
    )

    record = build_normalized_models(trimmed, batch_id='batch-1', source_file_id='source-1')[0]
    record.person_name = '比例样本'
    record.employee_id = 'H2001'
    record.id_number = '440100199001012201'
    record.housing_fund_personal = Decimal('0.07')
    record.housing_fund_company = Decimal('517')
    record.housing_fund_total = Decimal('1034')
    record.raw_payload = {
        'merged_sources': [
            {
                'source_kind': 'housing_fund',
                'source_file_name': '上海合并账单.xlsx',
                'source_row_number': 20,
            }
        ]
    }

    output_dir = ARTIFACTS_ROOT / 'housing_ratio_fix'
    output_dir.mkdir(parents=True, exist_ok=True)

    result = export_dual_templates(
        [record],
        output_dir=output_dir,
        salary_template_path=salary_template,
        final_tool_template_path=tool_template,
        export_prefix='housing_ratio_fix',
    )

    salary_artifact = next(item for item in result.artifacts if item.template_type == 'salary')
    salary_wb = load_workbook(salary_artifact.file_path, data_only=False)
    salary_sheet = salary_wb[salary_wb.sheetnames[0]]
    assert float(salary_sheet['H2'].value) == 517.0
    assert float(salary_sheet['P2'].value) == 517.0
    salary_wb.close()


def test_export_dual_templates_zeroes_housing_burden_when_no_reliable_baseline_exists() -> None:
    salary_template = find_template('薪酬')
    tool_template = find_template('最终版')
    sample_path = find_sample('深圳创造欢乐')

    standardized = standardize_workbook(sample_path, region='shenzhen', company_name='创造欢乐')
    trimmed = type(standardized)(
        source_file=standardized.source_file,
        sheet_name=standardized.sheet_name,
        raw_header_signature=standardized.raw_header_signature,
        records=standardized.records[:1],
        filtered_rows=standardized.filtered_rows,
        unmapped_headers=standardized.unmapped_headers,
    )
    record = build_normalized_models(trimmed, batch_id='batch-1', source_file_id='source-1')[0]
    record.person_name = '单条样本'
    record.employee_id = 'E2001'
    record.id_number = '440100199001012001'
    record.personal_total_amount = Decimal('0')
    record.housing_fund_personal = Decimal('500')
    record.housing_fund_company = Decimal('500')
    record.housing_fund_total = Decimal('1000')
    record.raw_payload = {
        'merged_sources': [
            {
                'source_kind': 'housing_fund',
                'source_file_name': '单条住房公积金.xlsx',
                'source_row_number': 8,
            }
        ]
    }

    output_dir = ARTIFACTS_ROOT / 'housing_burden_fallback'
    output_dir.mkdir(parents=True, exist_ok=True)

    result = export_dual_templates(
        [record],
        output_dir=output_dir,
        salary_template_path=salary_template,
        final_tool_template_path=tool_template,
        export_prefix='housing_burden_fallback',
    )

    salary_artifact = next(item for item in result.artifacts if item.template_type == 'salary')
    tool_artifact = next(item for item in result.artifacts if item.template_type == 'final_tool')

    salary_wb = load_workbook(salary_artifact.file_path, data_only=False)
    salary_sheet = salary_wb[salary_wb.sheetnames[0]]
    assert float(salary_sheet['H2'].value) == 500.0
    assert float(salary_sheet['R2'].value) == 0.0
    salary_wb.close()

    tool_wb = load_workbook(tool_artifact.file_path, data_only=False)
    tool_sheet = tool_wb[tool_wb.sheetnames[0]]
    assert float(tool_sheet['N7'].value) == 500.0
    assert float(tool_sheet['X7'].value) == 0.0
    tool_wb.close()


def test_export_dual_templates_filters_housing_only_rows_when_split_is_only_inferred() -> None:
    salary_template = find_template('薪酬')
    tool_template = find_template('最终版')
    sample_path = find_sample('深圳创造欢乐')

    standardized = standardize_workbook(sample_path, region='shenzhen', company_name='创造欢乐')
    trimmed = type(standardized)(
        source_file=standardized.source_file,
        sheet_name=standardized.sheet_name,
        raw_header_signature=standardized.raw_header_signature,
        records=standardized.records[:1],
        filtered_rows=standardized.filtered_rows,
        unmapped_headers=standardized.unmapped_headers,
    )
    record = build_normalized_models(trimmed, batch_id='batch-1', source_file_id='source-1')[0]
    record.person_name = '推导公积金'
    record.employee_id = 'H3001'
    record.id_number = '440100199001013001'
    clear_social_amounts(record)
    record.housing_fund_personal = Decimal('175')
    record.housing_fund_company = Decimal('175')
    record.housing_fund_total = Decimal('350')
    record.raw_payload = {
        'housing_fund_inference_notes': ['split_from_total_and_rates'],
        'merged_sources': [
            {
                'source_kind': 'housing_fund',
                'source_file_name': '深圳创造欢乐202602公积金账单.xlsx',
                'source_row_number': 73,
                'raw_values': {
                    '姓名': '推导公积金',
                    '证件号码': '440100199001013001',
                    '个人账号': '123456',
                    '缴存基数（元）': '3500',
                    '单位缴存比例': '0.05',
                    '个人缴存比例': '0.05',
                    '金额合计（元）': '350',
                },
            }
        ],
    }

    output_dir = ARTIFACTS_ROOT / 'housing_only_inferred_filter'
    output_dir.mkdir(parents=True, exist_ok=True)

    result = export_dual_templates(
        [record],
        output_dir=output_dir,
        salary_template_path=salary_template,
        final_tool_template_path=tool_template,
        export_prefix='housing_only_inferred_filter',
    )

    salary_artifact = next(item for item in result.artifacts if item.template_type == 'salary')
    tool_artifact = next(item for item in result.artifacts if item.template_type == 'final_tool')
    assert salary_artifact.row_count == 0
    assert tool_artifact.row_count == 0


def test_export_dual_templates_keeps_housing_only_rows_when_source_has_explicit_split_columns() -> None:
    salary_template = find_template('薪酬')
    tool_template = find_template('最终版')
    sample_path = find_sample('深圳创造欢乐')

    standardized = standardize_workbook(sample_path, region='shenzhen', company_name='创造欢乐')
    trimmed = type(standardized)(
        source_file=standardized.source_file,
        sheet_name=standardized.sheet_name,
        raw_header_signature=standardized.raw_header_signature,
        records=standardized.records[:1],
        filtered_rows=standardized.filtered_rows,
        unmapped_headers=standardized.unmapped_headers,
    )
    record = build_normalized_models(trimmed, batch_id='batch-1', source_file_id='source-1')[0]
    record.person_name = '显式公积金'
    record.employee_id = 'H3002'
    record.id_number = '440100199001013002'
    clear_social_amounts(record)
    record.housing_fund_personal = Decimal('175')
    record.housing_fund_company = Decimal('175')
    record.housing_fund_total = Decimal('350')
    record.raw_payload = {
        'merged_sources': [
            {
                'source_kind': 'housing_fund',
                'source_file_name': '深圳无限增长202602公积金账单.xlsx',
                'source_row_number': 6,
                'raw_values': {
                    '姓名': '显式公积金',
                    '证件号码': '440100199001013002',
                    '个人账号': '654321',
                    '缴存基数（元）': '3500',
                    '单位': 175,
                    '个人': 175,
                },
            }
        ],
    }

    output_dir = ARTIFACTS_ROOT / 'housing_only_explicit_keep'
    output_dir.mkdir(parents=True, exist_ok=True)

    result = export_dual_templates(
        [record],
        output_dir=output_dir,
        salary_template_path=salary_template,
        final_tool_template_path=tool_template,
        export_prefix='housing_only_explicit_keep',
    )

    salary_artifact = next(item for item in result.artifacts if item.template_type == 'salary')
    tool_artifact = next(item for item in result.artifacts if item.template_type == 'final_tool')
    assert salary_artifact.row_count == 1
    assert tool_artifact.row_count == 1
