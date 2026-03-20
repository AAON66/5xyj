from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from backend.app.core.config import ROOT_DIR
from backend.app.validators import classify_row, filter_candidate_rows


SAMPLES_DIR = ROOT_DIR / "data" / "samples"


SHENZHEN_KEYWORD = "\u6df1\u5733\u521b\u9020\u6b22\u4e50"
WUHAN_KEYWORD = "\u6b66\u6c49"
XIAMEN_KEYWORD = "\u53a6\u95e8202602\u793e\u4fdd\u8d26\u5355.xlsx"


def find_sample(keyword: str) -> Path:
    for path in sorted(SAMPLES_DIR.glob("*.xlsx")):
        if keyword in path.name:
            return path
    pytest.skip(f"Sample containing {keyword!r} was not found in {SAMPLES_DIR}.")


def load_row_values(path: Path, sheet_name: str, row_number: int) -> list[object]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        return list(next(sheet.iter_rows(min_row=row_number, max_row=row_number, values_only=True)))
    finally:
        workbook.close()


def test_classify_row_filters_real_shenzhen_group_header() -> None:
    sample_path = find_sample(SHENZHEN_KEYWORD)

    header_row = load_row_values(sample_path, sheet_name="\u7533\u62a5\u660e\u7ec6", row_number=3)
    detail_row = load_row_values(sample_path, sheet_name="\u7533\u62a5\u660e\u7ec6", row_number=4)

    header_decision = classify_row(header_row, row_number=3)
    detail_decision = classify_row(detail_row, row_number=4)

    assert header_decision.keep is False
    assert header_decision.reason == "group_header"
    assert header_decision.first_value == "\u5728\u804c\u4eba\u5458"

    assert detail_decision.keep is True
    assert detail_decision.reason == "detail_row"


def test_classify_row_filters_real_wuhan_total_row() -> None:
    sample_path = find_sample(WUHAN_KEYWORD)

    total_row = load_row_values(sample_path, sheet_name="Sheet1", row_number=4)
    decision = classify_row(total_row, row_number=4)

    assert decision.keep is False
    assert decision.reason == "summary_total"
    assert decision.first_value == "\u5408\u8ba1"


def test_classify_row_filters_real_xiamen_total_row() -> None:
    sample_path = find_sample(XIAMEN_KEYWORD)

    total_row = load_row_values(
        sample_path,
        sheet_name="\u804c\u5de5\u793e\u4fdd\u5bf9\u8d26\u5355\u660e\u7ec6\u67e5\u8be2",
        row_number=37,
    )
    decision = classify_row(total_row, row_number=37)

    assert decision.keep is False
    assert decision.reason == "summary_total"
    assert decision.first_value == "\u5408\u8ba1"


@pytest.mark.parametrize(
    ("row_values", "expected_reason"),
    [
        (["\u5c0f\u8ba1", None, None], "summary_subtotal"),
        (["\u9000\u4f11\u4eba\u5458", None, None], "group_header"),
        (["\u5bb6\u5c5e\u7edf\u7b79\u4eba\u5458", None, None], "group_header"),
        ([None, "", "--"], "blank_row"),
    ],
)
def test_classify_row_filters_known_non_detail_shapes(row_values: list[object], expected_reason: str) -> None:
    decision = classify_row(row_values, row_number=12)

    assert decision.keep is False
    assert decision.reason == expected_reason


def test_classify_row_keeps_sparse_detail_row_with_amounts() -> None:
    decision = classify_row(["\u5f20\u4e09", "", None, "3500.00"], row_number=9)

    assert decision.keep is True
    assert decision.reason == "detail_row"
    assert decision.first_value == "\u5f20\u4e09"


def test_filter_candidate_rows_returns_kept_and_filtered_rows() -> None:
    rows = [
        (3, ["\u5728\u804c\u4eba\u5458", None, None]),
        (4, ["\u5f20\u4e09", "440101199001010011", 3500.0]),
        (5, ["\u5408\u8ba1", None, 3500.0]),
        (6, ["\u5c0f\u8ba1", None, None]),
    ]

    result = filter_candidate_rows(rows)

    assert [row.row_number for row in result.kept_rows] == [4]
    assert [row.values[0] for row in result.kept_rows] == ["\u5f20\u4e09"]
    assert [(row.row_number, row.reason) for row in result.filtered_rows] == [
        (3, "group_header"),
        (5, "summary_total"),
        (6, "summary_subtotal"),
    ]
