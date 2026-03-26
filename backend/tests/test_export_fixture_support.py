from __future__ import annotations

import shutil
from pathlib import Path

import pytest
from _pytest.outcomes import Failed
from openpyxl import load_workbook

from backend.app.core.config import Settings, ROOT_DIR
from backend.tests.support import export_fixtures


ARTIFACTS_ROOT = ROOT_DIR / '.test_artifacts' / 'export_fixture_support'


def _prepare_artifact_dir(name: str) -> Path:
    target = ARTIFACTS_ROOT / name
    if target.exists():
        shutil.rmtree(target)
    target.mkdir(parents=True, exist_ok=True)
    return target


def test_resolve_required_export_templates_uses_repo_manifest_defaults() -> None:
    templates = export_fixtures.resolve_required_export_templates()

    assert templates.manifest == export_fixtures.REGRESSION_TEMPLATE_MANIFEST
    assert templates.salary == export_fixtures.REGRESSION_TEMPLATE_DIR / 'salary-template.xlsx'
    assert templates.final_tool == export_fixtures.REGRESSION_TEMPLATE_DIR / 'final-tool-template.xlsx'


def test_resolve_required_export_templates_prefers_valid_explicit_paths() -> None:
    artifacts_dir = _prepare_artifact_dir('explicit_paths')
    defaults = export_fixtures.create_placeholder_template_pair(artifacts_dir / 'defaults')
    explicit = export_fixtures.create_placeholder_template_pair(artifacts_dir / 'explicit')
    manifest_copy = artifacts_dir / 'manifest.json'
    shutil.copy2(defaults.manifest, manifest_copy)

    settings = Settings(
        salary_template_path=str(explicit.salary),
        final_tool_template_path=str(explicit.final_tool),
    )

    original_dir = export_fixtures.REGRESSION_TEMPLATE_DIR
    original_manifest = export_fixtures.REGRESSION_TEMPLATE_MANIFEST
    export_fixtures.REGRESSION_TEMPLATE_DIR = defaults.manifest.parent
    export_fixtures.REGRESSION_TEMPLATE_MANIFEST = manifest_copy
    try:
        resolved = export_fixtures.resolve_required_export_templates(settings)
    finally:
        export_fixtures.REGRESSION_TEMPLATE_DIR = original_dir
        export_fixtures.REGRESSION_TEMPLATE_MANIFEST = original_manifest

    assert resolved.salary == explicit.salary
    assert resolved.final_tool == explicit.final_tool
    assert resolved.manifest == manifest_copy


def test_resolve_required_export_templates_fails_loudly_when_manifest_assets_are_missing() -> None:
    fixture_dir = _prepare_artifact_dir('missing_manifest_assets')
    manifest = fixture_dir / 'manifest.json'
    manifest.write_text(
        '{"salary":"salary-template.xlsx","final_tool":"final-tool-template.xlsx"}',
        encoding='utf-8',
    )

    original_dir = export_fixtures.REGRESSION_TEMPLATE_DIR
    original_manifest = export_fixtures.REGRESSION_TEMPLATE_MANIFEST
    export_fixtures.REGRESSION_TEMPLATE_DIR = fixture_dir
    export_fixtures.REGRESSION_TEMPLATE_MANIFEST = manifest
    try:
        with pytest.raises(Failed, match='Missing required export templates'):
            export_fixtures.resolve_required_export_templates()
    finally:
        export_fixtures.REGRESSION_TEMPLATE_DIR = original_dir
        export_fixtures.REGRESSION_TEMPLATE_MANIFEST = original_manifest


def test_require_sample_workbook_fails_loudly_for_missing_sample() -> None:
    with pytest.raises(Failed, match='Missing required sample workbook'):
        export_fixtures.require_sample_workbook('missing-sample-keyword')


def test_create_placeholder_template_pair_creates_valid_workbooks() -> None:
    templates = export_fixtures.create_placeholder_template_pair(_prepare_artifact_dir('placeholder_pair'))

    assert templates.salary.exists()
    assert templates.final_tool.exists()
    assert templates.manifest.exists()

    salary_workbook = load_workbook(templates.salary, data_only=False)
    tool_workbook = load_workbook(templates.final_tool, data_only=False)
    assert salary_workbook[salary_workbook.sheetnames[0]]['A1'].value == '员工姓名'
    assert str(tool_workbook[tool_workbook.sheetnames[0]]['AA7'].value).startswith('=')
    assert str(tool_workbook[tool_workbook.sheetnames[0]]['AO7'].value).startswith('=')
    salary_workbook.close()
    tool_workbook.close()
