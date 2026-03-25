from __future__ import annotations

from typing import Optional

import shutil
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from backend.app.core.config import ROOT_DIR, get_settings
from backend.app.exporters import export_dual_templates
from backend.app.services import build_normalized_models, standardize_workbook

ARTIFACTS_ROOT = ROOT_DIR / ".test_artifacts" / "template_exporter_regression"
DESKTOP_ROOT = Path.home() / "Desktop" / "202602社保公积金台账" / "202602社保公积金汇总"
SAMPLES_DIR = ROOT_DIR / "data" / "samples"


@dataclass(frozen=True, slots=True)
class ExportRegressionCase:
    keyword: str
    region: str
    company_name: str
    min_records: int = 1


EXPORT_REGRESSION_CASES = [
    ExportRegressionCase("\u5e7f\u5206", "guangzhou", "\u5e7f\u5206\u793a\u4f8b"),
    ExportRegressionCase("\u676d\u5dde\u805a\u53d8", "hangzhou", "\u676d\u5dde\u805a\u53d8"),
    ExportRegressionCase("\u53a6\u95e8202602\u793e\u4fdd\u8d26\u5355.xlsx", "xiamen", "\u53a6\u95e8\u793a\u4f8b"),
    ExportRegressionCase("\u6df1\u5733\u521b\u9020\u6b22\u4e50", "shenzhen", "\u521b\u9020\u6b22\u4e50", min_records=2),
    ExportRegressionCase("\u6b66\u6c49", "wuhan", "\u6b66\u6c49\u793a\u4f8b"),
    ExportRegressionCase("\u957f\u6c99", "changsha", "\u957f\u6c99\u793a\u4f8b\u516c\u53f8"),
]


def find_template(keyword: str) -> Path:
    settings = get_settings()
    configured = [settings.salary_template_file, settings.final_tool_template_file]
    for path in configured:
        if path is not None and path.exists() and keyword in path.name:
            return path
    if DESKTOP_ROOT.exists():
        for path in sorted(DESKTOP_ROOT.glob("*.xlsx")):
            if keyword in path.name:
                return path
    pytest.skip(f"Template containing {keyword!r} was not found in configured paths or {DESKTOP_ROOT}.")



def find_sample(keyword: str) -> Path:
    for path in sorted(SAMPLES_DIR.glob("*.xlsx")):
        if keyword in path.name:
            return path
    pytest.skip(f"Sample containing {keyword!r} was not found in {SAMPLES_DIR}.")



def decimal_or_zero(value: Optional[Decimal]) -> Decimal:
    return value if value is not None else Decimal("0")



def build_export_records(case: ExportRegressionCase):
    sample_path = find_sample(case.keyword)
    standardized = standardize_workbook(sample_path, region=case.region, company_name=case.company_name)
    assert len(standardized.records) >= case.min_records
    trimmed = type(standardized)(
        source_file=standardized.source_file,
        sheet_name=standardized.sheet_name,
        raw_header_signature=standardized.raw_header_signature,
        records=standardized.records[: case.min_records],
        filtered_rows=standardized.filtered_rows,
        unmapped_headers=standardized.unmapped_headers,
    )
    records = build_normalized_models(trimmed, batch_id=f"batch-{case.region}", source_file_id=f"source-{case.region}")
    for index, record in enumerate(records, start=1):
        record.employee_id = f"{index:05d}"
    return sample_path, records


@pytest.mark.parametrize("case", EXPORT_REGRESSION_CASES, ids=[case.region for case in EXPORT_REGRESSION_CASES])
def test_export_dual_templates_on_cross_region_real_samples(case: ExportRegressionCase) -> None:
    salary_template = find_template("\u85aa\u916c")
    tool_template = find_template("\u6700\u7ec8\u7248")
    _sample_path, records = build_export_records(case)

    output_dir = ARTIFACTS_ROOT / case.region
    if output_dir.exists():
        shutil.rmtree(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    result = export_dual_templates(
        records,
        output_dir=output_dir,
        salary_template_path=salary_template,
        final_tool_template_path=tool_template,
        export_prefix=f"regression_{case.region}",
    )

    assert result.status == "completed"
    salary_artifact = next(item for item in result.artifacts if item.template_type == "salary")
    tool_artifact = next(item for item in result.artifacts if item.template_type == "final_tool")
    assert salary_artifact.status == "completed"
    assert tool_artifact.status == "completed"
    assert salary_artifact.row_count == len(records)
    assert tool_artifact.row_count == len(records)
    assert salary_artifact.file_path is not None and Path(salary_artifact.file_path).exists()
    assert tool_artifact.file_path is not None and Path(tool_artifact.file_path).exists()

    first = records[0]
    salary_wb = load_workbook(salary_artifact.file_path, data_only=False)
    salary_sheet = salary_wb[salary_wb.sheetnames[0]]
    assert salary_sheet["A2"].value == first.person_name
    assert salary_sheet["B2"].value == first.employee_id
    assert float(salary_sheet["Q2"].value) >= 0.0
    salary_wb.close()

    tool_wb = load_workbook(tool_artifact.file_path, data_only=False)
    tool_sheet = tool_wb[tool_wb.sheetnames[0]]
    assert tool_sheet["A7"].value == case.company_name
    assert tool_sheet["B7"].value
    assert tool_sheet["D7"].value == first.id_number
    assert tool_sheet["E7"].value == first.employee_id
    assert float(tool_sheet["W7"].value) >= 0.0
    tool_wb.close()
