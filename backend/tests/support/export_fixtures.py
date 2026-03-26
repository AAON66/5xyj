from __future__ import annotations

from dataclasses import dataclass
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
import pytest

from backend.app.core.config import ROOT_DIR, Settings


REGRESSION_TEMPLATE_DIR = ROOT_DIR / 'data' / 'templates' / 'regression'
REGRESSION_TEMPLATE_MANIFEST = REGRESSION_TEMPLATE_DIR / 'manifest.json'
SAMPLES_DIR = ROOT_DIR / 'data' / 'samples'
HOUSING_SAMPLES_DIR = SAMPLES_DIR / '公积金'
SALARY_TEMPLATE_FILENAME = 'salary-template.xlsx'
FINAL_TOOL_TEMPLATE_FILENAME = 'final-tool-template.xlsx'


@dataclass(frozen=True, slots=True)
class RequiredExportTemplates:
    salary: Path
    final_tool: Path
    manifest: Path


def resolve_required_export_templates(settings: Settings | None = None) -> RequiredExportTemplates:
    manifest_entries = _load_manifest_entries(REGRESSION_TEMPLATE_MANIFEST)

    salary = (settings.salary_template_file if settings is not None else None) or (
        REGRESSION_TEMPLATE_DIR / manifest_entries['salary']
    )
    final_tool = (settings.final_tool_template_file if settings is not None else None) or (
        REGRESSION_TEMPLATE_DIR / manifest_entries['final_tool']
    )

    missing = [str(path) for path in (salary, final_tool) if not path.exists()]
    if missing:
        pytest.fail(f"Missing required export templates: {', '.join(missing)}", pytrace=False)

    return RequiredExportTemplates(
        salary=salary,
        final_tool=final_tool,
        manifest=REGRESSION_TEMPLATE_MANIFEST,
    )


def require_sample_workbook(keyword: str, *, housing: bool = False) -> Path:
    sample_root = HOUSING_SAMPLES_DIR if housing else SAMPLES_DIR
    for path in sorted(sample_root.glob('*.xlsx')):
        if keyword in path.name:
            return path
    pytest.fail(
        f"Missing required sample workbook containing {keyword!r} in {sample_root}",
        pytrace=False,
    )


def create_placeholder_template_pair(target_dir: Path) -> RequiredExportTemplates:
    target_dir.mkdir(parents=True, exist_ok=True)

    salary_template = target_dir / SALARY_TEMPLATE_FILENAME
    final_tool_template = target_dir / FINAL_TOOL_TEMPLATE_FILENAME
    manifest_path = target_dir / 'manifest.json'

    _create_salary_placeholder_workbook(salary_template)
    _create_final_tool_placeholder_workbook(final_tool_template)
    manifest_path.write_text(
        json.dumps(
            {
                'salary': SALARY_TEMPLATE_FILENAME,
                'final_tool': FINAL_TOOL_TEMPLATE_FILENAME,
            },
            ensure_ascii=False,
            indent=2,
        )
        + '\n',
        encoding='utf-8',
    )

    return RequiredExportTemplates(
        salary=salary_template,
        final_tool=final_tool_template,
        manifest=manifest_path,
    )


def _load_manifest_entries(manifest_path: Path) -> dict[str, str]:
    if not manifest_path.exists():
        pytest.fail(f'Missing required export manifest: {manifest_path}', pytrace=False)

    try:
        payload: Any = json.loads(manifest_path.read_text(encoding='utf-8'))
    except json.JSONDecodeError as exc:
        pytest.fail(f'Invalid export manifest {manifest_path}: {exc}', pytrace=False)

    if not isinstance(payload, dict):
        pytest.fail(f'Invalid export manifest {manifest_path}: expected an object payload', pytrace=False)

    salary = payload.get('salary')
    final_tool = payload.get('final_tool')
    if not isinstance(salary, str) or not salary:
        pytest.fail(f'Invalid export manifest {manifest_path}: missing salary entry', pytrace=False)
    if not isinstance(final_tool, str) or not final_tool:
        pytest.fail(f'Invalid export manifest {manifest_path}: missing final_tool entry', pytrace=False)
    return {'salary': salary, 'final_tool': final_tool}


def _create_salary_placeholder_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    headers = [
        '员工姓名',
        '工号',
        '个人医疗保险',
        '个人失业保险',
        '个人大病医疗',
        '个人残疾保障金',
        '个人养老保险',
        '个人公积金',
        '公司养老保险',
        '公司医疗保险',
        '公司失业保险',
        '公司工伤保险',
        '公司生育保险',
        '公司大病医疗',
        '公司残疾保障金',
        '公司公积金',
        '个人社保承担额',
        '个人公积金承担额',
    ]
    for index, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=index, value=header)
        sheet.cell(row=2, column=index, value=None)
    workbook.save(path)
    workbook.close()


def _create_final_tool_placeholder_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    headers = [f'列{index}' for index in range(1, 42)]
    for index, header in enumerate(headers, start=1):
        sheet.cell(row=6, column=index, value=header)
        sheet.cell(row=7, column=index, value=None)
    sheet['AA7'] = '=SUM(A7:Z7)'
    sheet['AO7'] = '=SUM(A7:AN7)'
    workbook.save(path)
    workbook.close()


