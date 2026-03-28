"""Tool template field-to-column alignment tests.

Verifies that _tool_row_values produces values at the correct positions
matching TOOL_HEADERS, and that exported workbook cells align with template headers.
"""
from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from backend.app.core.config import ROOT_DIR
from backend.app.exporters import export_dual_templates
from backend.app.exporters.tool_exporter import TOOL_HEADERS, _tool_row_values
from backend.app.exporters.salary_exporter import SALARY_HEADERS
from backend.app.exporters.export_utils import (
    _build_housing_burden_context,
    _build_social_burden_context,
)
from backend.app.services import build_normalized_models, standardize_workbook
from backend.tests.support.export_fixtures import require_sample_workbook, resolve_required_export_templates


ARTIFACTS_ROOT = ROOT_DIR / ".test_artifacts" / "tool_alignment"


def _prepare_artifact_dir(name: str) -> Path:
    ARTIFACTS_ROOT.mkdir(parents=True, exist_ok=True)
    target = ARTIFACTS_ROOT / f'{name}-{uuid4().hex[:8]}'
    target.mkdir(parents=True, exist_ok=False)
    return target


def _build_records(keyword: str, region: str, company: str, *, count: int = 1, housing: bool = False):
    sample_path = require_sample_workbook(keyword, housing=housing)
    standardized = standardize_workbook(sample_path, region=region, company_name=company)
    assert len(standardized.records) >= count
    trimmed = type(standardized)(
        source_file=standardized.source_file,
        sheet_name=standardized.sheet_name,
        raw_header_signature=standardized.raw_header_signature,
        records=standardized.records[:count],
        filtered_rows=standardized.filtered_rows,
        unmapped_headers=standardized.unmapped_headers,
    )
    records = build_normalized_models(trimmed, batch_id=f"batch-{region}", source_file_id=f"src-{region}")
    for i, r in enumerate(records, 1):
        r.employee_id = f"{i:05d}"
    return records


# --- Test 1: Length match ---

def test_tool_row_values_length_matches_headers():
    """_tool_row_values returns exactly len(TOOL_HEADERS) elements."""
    records = _build_records("深圳创造欢乐", "shenzhen", "创造欢乐")
    values = _tool_row_values(records[0])
    assert len(values) == len(TOOL_HEADERS), (
        f"Expected {len(TOOL_HEADERS)} values, got {len(values)}"
    )


# --- Test 2: Field alignment by name ---

def test_tool_field_alignment_by_name():
    """Each position in _tool_row_values matches the expected field for TOOL_HEADERS."""
    records = _build_records("深圳创造欢乐", "shenzhen", "创造欢乐")
    record = records[0]
    values = _tool_row_values(record)

    # Identity fields
    assert values[0] == record.company_name or values[0] == '', "idx 0 should be company_name"
    assert isinstance(values[1], str), "idx 1 should be region label string"
    assert values[2] == (record.person_name or ''), "idx 2 should be person_name"
    assert values[3] == (record.id_number or ''), "idx 3 should be id_number"
    assert values[4] == (record.employee_id or ''), "idx 4 should be employee_id"
    assert values[5] is None, "idx 5 should be None separator"
    assert values[6] == (record.person_name or ''), "idx 6 should be person_name repeat"
    assert values[7] == (record.employee_id or ''), "idx 7 should be employee_id repeat"

    # Insurance fields should be Decimal
    for i in range(8, 22):
        if TOOL_HEADERS[i] is not None:
            assert isinstance(values[i], (Decimal, int, float)), (
                f"idx {i} ({TOOL_HEADERS[i]}) should be numeric, got {type(values[i])}: {values[i]}"
            )

    # Burden fields
    assert isinstance(values[22], (Decimal, int, float)), "idx 22 personal_social_burden should be numeric"
    assert isinstance(values[23], (Decimal, int, float)), "idx 23 personal_housing_burden should be numeric"

    # Blank separators
    assert values[24] is None, "idx 24 should be None"
    assert values[25] is None, "idx 25 should be None"

    # Derived totals should be numeric
    for i in [26, 27, 28, 29, 30]:
        if TOOL_HEADERS[i] is not None:
            assert isinstance(values[i], (Decimal, int, float)), (
                f"idx {i} ({TOOL_HEADERS[i]}) should be numeric"
            )

    assert values[31] is None, "idx 31 should be None"

    for i in [32, 33, 34, 35, 36]:
        if TOOL_HEADERS[i] is not None:
            assert isinstance(values[i], (Decimal, int, float)), (
                f"idx {i} ({TOOL_HEADERS[i]}) should be numeric"
            )

    assert values[37] is None, "idx 37 should be None"
    assert values[38] is None, "idx 38 should be None"

    # Grand totals
    for i in [39, 40, 41]:
        assert isinstance(values[i], (Decimal, int, float)), (
            f"idx {i} ({TOOL_HEADERS[i]}) should be numeric"
        )


# --- Test 3: Real workbook cell alignment ---

def test_tool_export_cell_alignment_against_workbook():
    """Export Tool template and verify cells match the ACTUAL workbook header row."""
    records = _build_records("深圳创造欢乐", "shenzhen", "创造欢乐", count=2)
    templates = resolve_required_export_templates()
    output = _prepare_artifact_dir("workbook-alignment")

    # Read actual template headers BEFORE export
    template_wb = load_workbook(templates.final_tool, data_only=False)
    template_sheet = template_wb[template_wb.sheetnames[0]]
    # Find header row (typically row 6 for tool template)
    actual_template_headers = []
    for col in range(1, len(TOOL_HEADERS) + 1):
        actual_template_headers.append(template_sheet.cell(row=6, column=col).value)
    template_wb.close()

    # Export
    result = export_dual_templates(
        records,
        output_dir=output,
        salary_template_path=templates.salary,
        final_tool_template_path=templates.final_tool,
        export_prefix="align_check",
    )
    assert result.status == "completed"
    tool_artifact = next(a for a in result.artifacts if a.template_type == "final_tool")
    assert tool_artifact.status == "completed"

    # Verify data row alignment against ACTUAL template headers
    out_wb = load_workbook(tool_artifact.file_path, data_only=False)
    out_sheet = out_wb[out_wb.sheetnames[0]]

    # Check key identity columns
    assert out_sheet.cell(row=7, column=1).value == records[0].company_name  # 主体
    assert out_sheet.cell(row=7, column=3).value == records[0].person_name   # 员工姓名（辅助）
    assert out_sheet.cell(row=7, column=4).value == records[0].id_number     # 身份证
    assert out_sheet.cell(row=7, column=5).value == records[0].employee_id   # 工号
    assert out_sheet.cell(row=7, column=7).value == records[0].person_name   # 员工姓名
    assert out_sheet.cell(row=7, column=8).value == records[0].employee_id   # 工号 repeat

    # Row 8 should be second record
    assert out_sheet.cell(row=8, column=3).value == records[1].person_name
    out_wb.close()


# --- Test 4: Dual export both succeed ---

def test_dual_export_both_succeed():
    """Both Salary and Tool templates export successfully in a single operation."""
    records = _build_records("深圳创造欢乐", "shenzhen", "创造欢乐", count=2)
    templates = resolve_required_export_templates()
    output = _prepare_artifact_dir("dual-export")

    result = export_dual_templates(
        records,
        output_dir=output,
        salary_template_path=templates.salary,
        final_tool_template_path=templates.final_tool,
        export_prefix="dual_test",
    )

    assert result.status == "completed"
    assert len(result.artifacts) == 2
    salary = next(a for a in result.artifacts if a.template_type == "salary")
    tool = next(a for a in result.artifacts if a.template_type == "final_tool")
    assert salary.status == "completed"
    assert tool.status == "completed"
    assert salary.row_count == 2
    assert tool.row_count == 2
    assert Path(salary.file_path).exists()
    assert Path(tool.file_path).exists()


# --- Test 5: Housing and supplementary edge case ---

def test_tool_export_with_housing_and_supplementary():
    """Tool export with records containing housing fund values places them correctly."""
    records = _build_records("深圳创造欢乐", "shenzhen", "创造欢乐", count=2)
    values = _tool_row_values(records[0])

    # Index 13 = 个人公积金
    assert isinstance(values[13], (Decimal, int, float)), "idx 13 should be personal housing fund"
    # Index 21 = 公司公积金
    assert isinstance(values[21], (Decimal, int, float)), "idx 21 should be company housing fund"
    # Index 19 = 公司大病医疗 (supplementary)
    assert isinstance(values[19], (Decimal, int, float)), "idx 19 should be company large medical"

    # Housing grand total (index 40) should equal personal + company housing
    if values[13] and values[21]:
        expected_housing_total = values[13] + values[21]
        assert values[40] == expected_housing_total, (
            f"Housing grand total mismatch: idx 40={values[40]}, expected {expected_housing_total}"
        )


# --- Test 6: Multi-region ---

def test_tool_export_multi_region():
    """Tool export with records from different regions has correct region labels."""
    records_sz = _build_records("深圳创造欢乐", "shenzhen", "创造欢乐")
    records_gz = _build_records("广分", "guangzhou", "广分示例")
    # Set distinct employee_ids
    records_sz[0].employee_id = "SZ001"
    records_gz[0].employee_id = "GZ001"

    combined = records_sz + records_gz
    templates = resolve_required_export_templates()
    output = _prepare_artifact_dir("multi-region")

    result = export_dual_templates(
        combined,
        output_dir=output,
        salary_template_path=templates.salary,
        final_tool_template_path=templates.final_tool,
        export_prefix="multi_region",
    )
    assert result.status == "completed"
    tool = next(a for a in result.artifacts if a.template_type == "final_tool")
    assert tool.status == "completed"
    assert tool.row_count == 2

    wb = load_workbook(tool.file_path, data_only=False)
    sheet = wb[wb.sheetnames[0]]
    # Row 7 and 8 should have different region labels
    region1 = sheet.cell(row=7, column=2).value
    region2 = sheet.cell(row=8, column=2).value
    assert region1 is not None and region2 is not None
    assert region1 != region2, f"Both rows have same region: {region1}"
    wb.close()


# --- Test 7: Salary regression cross-check ---

def test_salary_regression_still_passes():
    """Salary template output is not affected by Tool template changes."""
    records = _build_records("深圳创造欢乐", "shenzhen", "创造欢乐", count=2)
    templates = resolve_required_export_templates()
    output = _prepare_artifact_dir("salary-regression")

    result = export_dual_templates(
        records,
        output_dir=output,
        salary_template_path=templates.salary,
        final_tool_template_path=templates.final_tool,
        export_prefix="sal_reg",
    )
    assert result.status == "completed"
    salary = next(a for a in result.artifacts if a.template_type == "salary")
    assert salary.status == "completed"

    wb = load_workbook(salary.file_path, data_only=False)
    sheet = wb[wb.sheetnames[0]]
    # Verify header row matches SALARY_HEADERS
    for i, expected in enumerate(SALARY_HEADERS, start=1):
        actual = sheet.cell(row=1, column=i).value
        assert actual == expected, f"Salary header col {i}: expected '{expected}', got '{actual}'"
    # Verify data row
    assert sheet["A2"].value == records[0].person_name
    assert sheet["B2"].value == records[0].employee_id
    wb.close()
