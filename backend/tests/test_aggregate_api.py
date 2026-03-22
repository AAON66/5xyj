from __future__ import annotations

import json
import shutil
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from backend.app.core.config import ROOT_DIR, Settings, get_settings
from backend.app.core.database import create_database_engine, create_session_factory
from backend.app.dependencies import get_db
from backend.app.main import create_app
from backend.app.models import Base
from backend.app.services import infer_region_from_filename, standardize_housing_fund_workbook, standardize_workbook

ARTIFACTS_ROOT = ROOT_DIR / '.test_artifacts' / 'aggregate_api'
SAMPLES_DIR = ROOT_DIR / 'data' / 'samples'
HOUSING_SAMPLES_DIR = ROOT_DIR / 'data' / 'samples' / '\u516c\u79ef\u91d1'
DESKTOP_ROOT = Path.home() / 'Desktop' / '\u0032\u0030\u0032\u0036\u0030\u0032\u793e\u4fdd\u516c\u79ef\u91d1\u53f0\u8d26' / '\u0032\u0030\u0032\u0036\u0030\u0032\u793e\u4fdd\u516c\u79ef\u91d1\u6c47\u603b'

SAMPLE_KEYWORD = '\u6df1\u5733\u521b\u9020\u6b22\u4e50'
COMPANY_NAME = '\u521b\u9020\u6b22\u4e50'
APP_NAME = '\u5feb\u901f\u805a\u5408\u6d4b\u8bd5'


def find_sample(keyword: str, *, housing: bool = False) -> Path:
    sample_root = HOUSING_SAMPLES_DIR if housing else SAMPLES_DIR
    for path in sorted(sample_root.glob('*.xlsx')):
        if keyword in path.name:
            return path
    pytest.skip(f'Sample containing {keyword!r} was not found in {sample_root}.')


def find_template(keyword: str) -> Path:
    settings = get_settings()
    configured = [settings.salary_template_file, settings.final_tool_template_file]
    for path in configured:
        if path is not None and path.exists() and keyword in path.name:
            return path
    if DESKTOP_ROOT.exists():
        for path in sorted(DESKTOP_ROOT.glob('*.xlsx')):
            if keyword in path.name:
                return path
    pytest.skip(f'Template containing {keyword!r} was not found in configured paths or {DESKTOP_ROOT}.')


def build_test_context(test_name: str, *, salary_template: Path, final_tool_template: Path):
    artifacts_dir = ARTIFACTS_ROOT / test_name
    if artifacts_dir.exists():
        shutil.rmtree(artifacts_dir)
    artifacts_dir.mkdir(parents=True, exist_ok=True)

    database_path = artifacts_dir / 'aggregate.db'
    settings = Settings(
        app_name=APP_NAME,
        app_version='0.2.0',
        database_url=f'sqlite:///{database_path.as_posix()}',
        upload_dir=str(artifacts_dir / 'uploads'),
        samples_dir=str(artifacts_dir / 'samples'),
        templates_dir=str(artifacts_dir / 'templates'),
        outputs_dir=str(artifacts_dir / 'outputs'),
        salary_template_path=str(salary_template),
        final_tool_template_path=str(final_tool_template),
        log_format='plain',
    )

    engine = create_database_engine(settings)
    session_factory = create_session_factory(settings)
    Base.metadata.create_all(engine)

    app = create_app(settings)

    def override_get_db():
        db: Session = session_factory()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_get_db
    return TestClient(app), settings, session_factory


def make_employee_master_csv(sample_path: Path) -> bytes:
    standardized = standardize_workbook(sample_path, region='shenzhen', company_name=COMPANY_NAME)
    first = standardized.records[0]
    return (
        '\u5de5\u53f7,\u59d3\u540d,\u8eab\u4efd\u8bc1\u53f7,\u516c\u53f8\u540d\u79f0,\u90e8\u95e8,\u662f\u5426\u5728\u804c\n'
        f"E9001,{first.values['person_name']},{first.values['id_number']},{COMPANY_NAME},\u8fd0\u8425,\u662f\n"
    ).encode('utf-8-sig')


def make_employee_master_csv_for_merged_files(social_sample: Path, housing_sample: Path) -> tuple[bytes, str]:
    social = standardize_workbook(social_sample, region='shenzhen', company_name=COMPANY_NAME)
    housing = standardize_housing_fund_workbook(housing_sample, region='shenzhen', company_name=COMPANY_NAME)
    social_by_id = {record.values.get('id_number'): record for record in social.records}
    for housing_record in housing.records:
        id_number = housing_record.values.get('id_number')
        if id_number in social_by_id:
            social_record = social_by_id[id_number]
            csv = (
                '\u5de5\u53f7,\u59d3\u540d,\u8eab\u4efd\u8bc1\u53f7,\u516c\u53f8\u540d\u79f0,\u90e8\u95e8,\u662f\u5426\u5728\u804c\n'
                f"E9001,{social_record.values['person_name']},{id_number},{COMPANY_NAME},\u8fd0\u8425,\u662f\n"
            ).encode('utf-8-sig')
            return csv, str(id_number)
    pytest.skip('Could not find a common employee between the Shenzhen social and housing fund samples.')


def test_aggregate_endpoint_exports_both_templates_without_employee_master() -> None:
    salary_template = find_template('\u85aa\u916c')
    tool_template = find_template('\u6700\u7ec8\u7248')
    sample_path = find_sample(SAMPLE_KEYWORD)
    client, _settings, _session_factory = build_test_context(
        'aggregate_without_master',
        salary_template=salary_template,
        final_tool_template=tool_template,
    )

    with client:
        response = client.post(
            '/api/v1/aggregate',
            data={'batch_name': 'quick-aggregate-no-master'},
            files=[
                (
                    'files',
                    (
                        sample_path.name,
                        sample_path.read_bytes(),
                        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    ),
                )
            ],
        )

    assert response.status_code == 201
    payload = response.json()['data']
    assert payload['batch_name'] == 'quick-aggregate-no-master'
    assert payload['status'] == 'exported'
    assert payload['export_status'] == 'completed'
    assert payload['blocked_reason'] is None
    assert payload['employee_master'] is None
    assert payload['matched_count'] == 0
    assert payload['unmatched_count'] >= 1
    assert len(payload['source_files']) == 1
    assert payload['source_files'][0]['source_kind'] == 'social_security'
    assert payload['source_files'][0]['region'] == 'shenzhen'
    assert payload['source_files'][0]['company_name'] == COMPANY_NAME
    assert payload['source_files'][0]['normalized_record_count'] >= 1
    assert len(payload['artifacts']) == 2
    assert all(item['status'] == 'completed' for item in payload['artifacts'])
    assert all(Path(item['file_path']).exists() for item in payload['artifacts'])


def test_aggregate_endpoint_detects_region_from_workbook_when_filename_is_generic() -> None:
    salary_template = find_template('\u85aa\u916c')
    tool_template = find_template('\u6700\u7ec8\u7248')
    sample_path = find_sample(SAMPLE_KEYWORD)
    client, _settings, _session_factory = build_test_context(
        'aggregate_region_from_workbook',
        salary_template=salary_template,
        final_tool_template=tool_template,
    )

    with client:
        response = client.post(
            '/api/v1/aggregate',
            data={'batch_name': 'quick-aggregate-region-from-workbook'},
            files=[
                (
                    'files',
                    (
                        'generic.xlsx',
                        sample_path.read_bytes(),
                        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    ),
                )
            ],
        )

    assert response.status_code == 201
    payload = response.json()['data']
    assert payload['source_files'][0]['region'] == 'shenzhen'
    assert payload['source_files'][0]['normalized_record_count'] >= 1


def test_aggregate_endpoint_imports_employee_master_and_matches_records() -> None:
    salary_template = find_template('\u85aa\u916c')
    tool_template = find_template('\u6700\u7ec8\u7248')
    sample_path = find_sample(SAMPLE_KEYWORD)
    employee_csv = make_employee_master_csv(sample_path)
    client, _settings, _session_factory = build_test_context(
        'aggregate_with_master',
        salary_template=salary_template,
        final_tool_template=tool_template,
    )

    with client:
        response = client.post(
            '/api/v1/aggregate',
            data={'batch_name': 'quick-aggregate-with-master'},
            files=[
                (
                    'files',
                    (
                        sample_path.name,
                        sample_path.read_bytes(),
                        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    ),
                ),
                ('employee_master_file', ('employee_master.csv', employee_csv, 'text/csv')),
            ],
        )

    assert response.status_code == 201
    payload = response.json()['data']
    assert payload['batch_name'] == 'quick-aggregate-with-master'
    assert payload['status'] == 'exported'
    assert payload['export_status'] == 'completed'
    assert payload['employee_master'] is not None
    assert payload['employee_master']['created_count'] == 1
    assert payload['matched_count'] >= 1
    assert payload['source_files'][0]['source_kind'] == 'social_security'
    assert payload['source_files'][0]['region'] == 'shenzhen'
    assert payload['source_files'][0]['company_name'] == COMPANY_NAME
    assert len(payload['artifacts']) == 2
    assert all(item['status'] == 'completed' for item in payload['artifacts'])
    assert all(Path(item['file_path']).exists() for item in payload['artifacts'])




def test_aggregate_download_endpoint_returns_generated_artifacts() -> None:
    salary_template = find_template('\u85aa\u916c')
    tool_template = find_template('\u6700\u7ec8\u7248')
    sample_path = find_sample(SAMPLE_KEYWORD)
    client, _settings, _session_factory = build_test_context(
        'aggregate_downloads',
        salary_template=salary_template,
        final_tool_template=tool_template,
    )

    with client:
        aggregate_response = client.post(
            '/api/v1/aggregate',
            data={'batch_name': 'quick-aggregate-downloads'},
            files=[
                (
                    'files',
                    (
                        sample_path.name,
                        sample_path.read_bytes(),
                        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    ),
                )
            ],
        )
        assert aggregate_response.status_code == 201
        batch_id = aggregate_response.json()['data']['batch_id']

        salary_download = client.get(f'/api/v1/imports/{batch_id}/export/salary/download')
        assert salary_download.status_code == 200
        assert 'attachment;' in salary_download.headers.get('content-disposition', '')
        assert salary_download.headers['content-type'].startswith(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        assert salary_download.content

        tool_download = client.get(f'/api/v1/imports/{batch_id}/export/final_tool/download')
        assert tool_download.status_code == 200
        assert 'attachment;' in tool_download.headers.get('content-disposition', '')
        assert tool_download.content


def test_aggregate_endpoint_merges_housing_fund_into_dual_exports() -> None:
    salary_template = find_template('\u85aa\u916c')
    tool_template = find_template('\u6700\u7ec8\u7248')
    social_sample = find_sample(SAMPLE_KEYWORD)
    housing_sample = find_sample(SAMPLE_KEYWORD, housing=True)
    employee_csv, matched_id_number = make_employee_master_csv_for_merged_files(social_sample, housing_sample)
    client, _settings, _session_factory = build_test_context(
        'aggregate_with_housing_fund',
        salary_template=salary_template,
        final_tool_template=tool_template,
    )

    with client:
        response = client.post(
            '/api/v1/aggregate',
            data={'batch_name': 'quick-aggregate-with-housing'},
            files=[
                ('files', (social_sample.name, social_sample.read_bytes(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')),
                ('housing_fund_files', (housing_sample.name, housing_sample.read_bytes(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')),
                ('employee_master_file', ('employee_master.csv', employee_csv, 'text/csv')),
            ],
        )

    assert response.status_code == 201
    payload = response.json()['data']
    assert payload['export_status'] == 'completed'
    assert len(payload['source_files']) == 2
    assert sorted(item['source_kind'] for item in payload['source_files']) == ['housing_fund', 'social_security']

    salary_artifact = next(item for item in payload['artifacts'] if item['template_type'] == 'salary')
    tool_artifact = next(item for item in payload['artifacts'] if item['template_type'] == 'final_tool')
    salary_wb = load_workbook(salary_artifact['file_path'], data_only=False)
    salary_sheet = salary_wb[salary_wb.sheetnames[0]]
    target_row = None
    for row in range(2, salary_sheet.max_row + 1):
        if salary_sheet[f'B{row}'].value == 'E9001':
            target_row = row
            break
    assert target_row is not None
    assert float(salary_sheet[f'H{target_row}'].value) > 0
    assert float(salary_sheet[f'P{target_row}'].value) > 0
    assert float(salary_sheet[f'R{target_row}'].value) > 0
    salary_wb.close()

    tool_wb = load_workbook(tool_artifact['file_path'], data_only=False)
    tool_sheet = tool_wb[tool_wb.sheetnames[0]]
    tool_row = None
    for row in range(7, tool_sheet.max_row + 1):
        if tool_sheet[f'D{row}'].value == matched_id_number and tool_sheet[f'E{row}'].value == 'E9001':
            tool_row = row
            break
    assert tool_row is not None
    assert float(tool_sheet[f'N{tool_row}'].value) > 0
    assert float(tool_sheet[f'V{tool_row}'].value) > 0
    tool_wb.close()


def test_aggregate_stream_endpoint_emits_progress_events() -> None:
    salary_template = find_template('\u85aa\u916c')
    tool_template = find_template('\u6700\u7ec8\u7248')
    sample_path = find_sample(SAMPLE_KEYWORD)
    client, _settings, _session_factory = build_test_context(
        'aggregate_stream',
        salary_template=salary_template,
        final_tool_template=tool_template,
    )

    with client, client.stream(
        'POST',
        '/api/v1/aggregate/stream',
        data={'batch_name': 'quick-aggregate-stream'},
        files=[
            (
                'files',
                (
                    sample_path.name,
                    sample_path.read_bytes(),
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                ),
            )
        ],
    ) as response:
        assert response.status_code == 200
        events = []
        for line in response.iter_lines():
            if not line:
                continue
            if isinstance(line, bytes):
                line = line.decode('utf-8')
            events.append(json.loads(line))

    progress_stages = [event['stage'] for event in events if event['event'] == 'progress']
    assert progress_stages[:3] == ['employee_import', 'employee_import', 'batch_upload']
    assert 'parse' in progress_stages
    assert 'validate' in progress_stages
    assert 'match' in progress_stages
    assert 'export' in progress_stages

    result_event = next(event for event in events if event['event'] == 'result')
    assert result_event['data']['batch_name'] == 'quick-aggregate-stream'
    assert result_event['data']['export_status'] == 'completed'
    assert len(result_event['data']['artifacts']) == 2


def test_infer_region_from_filename_prefers_explicit_shenzhen_label() -> None:
    assert infer_region_from_filename('\u6df1\u5733\u96f6\u4e00\u88c2\u53d8202602\u793e\u4fdd\u660e\u7ec6.xlsx') == 'shenzhen'
    assert infer_region_from_filename('202602\u6708\u96f6\u4e00\u88c2\u53d8\uff08\u6df1\u5733\uff09\u79d1\u6280\u6709\u9650\u516c\u53f8\u793e\u4fdd\u8d26\u5355.xlsx') == 'shenzhen'
    assert infer_region_from_filename('\u6df1\u5733\u88c2\u53d8202602\u516c\u79ef\u91d1\u8d26\u5355.xlsx') == 'shenzhen'



def test_aggregate_stream_endpoint_reports_intermediate_upload_and_parse_progress() -> None:
    salary_template = find_template('薪酬')
    tool_template = find_template('最终版')
    sample_path = find_sample(SAMPLE_KEYWORD)
    client, _settings, _session_factory = build_test_context(
        'aggregate_stream_granular_progress',
        salary_template=salary_template,
        final_tool_template=tool_template,
    )

    with client, client.stream(
        'POST',
        '/api/v1/aggregate/stream',
        data={'batch_name': 'quick-aggregate-stream-granular'},
        files=[
            ('files', ('sample-a.xlsx', sample_path.read_bytes(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')),
            ('files', ('sample-b.xlsx', sample_path.read_bytes(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')),
        ],
    ) as response:
        assert response.status_code == 200
        events = []
        for line in response.iter_lines():
            if not line:
                continue
            if isinstance(line, bytes):
                line = line.decode('utf-8')
            events.append(json.loads(line))

    upload_events = [event for event in events if event['event'] == 'progress' and event['stage'] == 'batch_upload']
    parse_events = [event for event in events if event['event'] == 'progress' and event['stage'] == 'parse']

    assert len(upload_events) >= 5
    assert any('\u6b63\u5728\u4fdd\u5b58\u6587\u4ef6 1/2' in event['message'] for event in upload_events)
    assert any('\u6587\u4ef6 1/2 \u5df2\u4fdd\u5b58' in event['message'] for event in upload_events)
    assert any('\u6b63\u5728\u4fdd\u5b58\u6587\u4ef6 2/2' in event['message'] for event in upload_events)
    assert any('\u6587\u4ef6 2/2 \u5df2\u4fdd\u5b58' in event['message'] for event in upload_events)
    assert len(parse_events) >= 3
    assert any('1/2' in event['message'] for event in parse_events)
    assert any('2/2' in event['message'] for event in parse_events)
    assert any(event.get('parse_summary', {}).get('worker_count', 0) >= 1 for event in parse_events)
    assert any(event.get('parse_summary', {}).get('saved_count', 0) >= 1 for event in parse_events)
    assert any(event.get('parse_files') for event in parse_events)
    detail_index = next(index for index, event in enumerate(events) if event['event'] == 'progress' and event['stage'] == 'parse' and '1/2' in event['message'])
    completion_index = next(index for index, event in enumerate(events) if event['event'] == 'progress' and event['stage'] == 'parse' and '\u5df2\u5b8c\u6210 2 \u4e2a\u6587\u4ef6\u7684\u89e3\u6790' in event['message'])
    assert detail_index < completion_index
    assert upload_events[-1]['percent'] >= upload_events[0]['percent']
    assert parse_events[-1]['percent'] >= parse_events[0]['percent']



def test_aggregate_stream_endpoint_reports_region_detection_for_generic_filename() -> None:
    salary_template = find_template('\u85aa\u916c')
    tool_template = find_template('\u6700\u7ec8\u7248')
    sample_path = find_sample(SAMPLE_KEYWORD)
    client, _settings, _session_factory = build_test_context(
        'aggregate_stream_region_detection_progress',
        salary_template=salary_template,
        final_tool_template=tool_template,
    )

    with client, client.stream(
        'POST',
        '/api/v1/aggregate/stream',
        data={'batch_name': 'quick-aggregate-stream-region-detection'},
        files=[
            (
                'files',
                (
                    'generic.xlsx',
                    sample_path.read_bytes(),
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                ),
            )
        ],
    ) as response:
        assert response.status_code == 200
        events = []
        for line in response.iter_lines():
            if not line:
                continue
            if isinstance(line, bytes):
                line = line.decode('utf-8')
            events.append(json.loads(line))

    upload_events = [event for event in events if event['event'] == 'progress' and event['stage'] == 'batch_upload']
    assert any('\u6b63\u5728\u8bc6\u522b\u6587\u4ef6 1/1 \u7684\u5730\u533a\u4fe1\u606f' in event['message'] for event in upload_events)
