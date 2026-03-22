from __future__ import annotations

import io
import json
import shutil
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy.orm import Session

from backend.app.core.config import ROOT_DIR, Settings
from backend.app.core.database import create_database_engine, create_session_factory
from backend.app.dependencies import get_db
from backend.app.main import create_app
from backend.app.models import Base
from backend.app.services import standardize_workbook


ARTIFACTS_ROOT = ROOT_DIR / '.test_artifacts' / 'employee_master_api'
SAMPLES_DIR = ROOT_DIR / 'data' / 'samples'
GUANGZHOU_KEYWORD = '\u5e7f\u5206'
GUANGZHOU_COMPANY = '\u5e7f\u5206\u793a\u4f8b'
APP_NAME = '\u5458\u5de5\u4e3b\u6863\u6d4b\u8bd5'


def build_test_context(test_name: str):
    artifacts_dir = ARTIFACTS_ROOT / test_name
    if artifacts_dir.exists():
        shutil.rmtree(artifacts_dir)
    artifacts_dir.mkdir(parents=True, exist_ok=True)

    database_path = artifacts_dir / 'employee.db'
    settings = Settings(
        app_name=APP_NAME,
        app_version='0.2.0',
        database_url=f'sqlite:///{database_path.as_posix()}',
        upload_dir=str(artifacts_dir / 'uploads'),
        samples_dir=str(artifacts_dir / 'samples'),
        templates_dir=str(artifacts_dir / 'templates'),
        outputs_dir=str(artifacts_dir / 'outputs'),
        log_format='plain',
    )

    engine = create_database_engine(settings)
    session_factory = create_session_factory(settings)
    Base.metadata.create_all(engine)

    app = create_app(settings)

    def override_get_db():
        db: Session = session_factory()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_get_db
    return TestClient(app), settings, session_factory


def make_csv_bytes() -> bytes:
    return (
        '工号,姓名,身份证号,公司名称,部门,是否在职\n'
        'E1001,张三,440101199001010011,广分示例,运营,是\n'
        'E1002,李四,440101199202020022,广分示例,人事,否\n'
    ).encode('utf-8-sig')


def make_update_workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '员工主档'
    sheet.append(['工号', '姓名', '身份证号', '公司', '部门名称', 'active'])
    sheet.append(['E1001', '张三', '440101199001010011', '广分示例', '平台运营', 'true'])
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def make_offset_alias_workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '\u82b1\u540d\u518c'
    sheet.append(['2026\u5e742\u6708\u5458\u5de5\u82b1\u540d\u518c'])
    sheet.append(['\u5458\u5de5\u7f16\u53f7', '\u5458\u5de5\u59d3\u540d', '\u8bc1\u4ef6\u53f7\u7801(\u8eab\u4efd\u8bc1)', '\u6240\u5c5e\u6cd5\u4eba\u516c\u53f8', '\u7ec4\u7ec7\u67b6\u6784', '\u4efb\u804c\u72b6\u6001'])
    sheet.append(['E2001', '\u738b\u4e94', '440101199303030033', '\u96f6\u4e00\u88c2\u53d8', '\u589e\u957f\u4e2d\u5fc3', '\u5728\u804c'])
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def make_slash_id_header_workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '花名册'
    sheet.append(['姓名', '工号', '证件/证件号码', '公司名称'])
    sheet.append(['赵六', 'E3001', '440101199404040044', '广分示例'])
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def find_sample(keyword: str) -> Path:
    for path in sorted(SAMPLES_DIR.glob('*.xlsx')):
        if keyword in path.name:
            return path
    pytest.skip(f'Sample containing {keyword!r} was not found in {SAMPLES_DIR}.')


def create_batch(client: TestClient, sample_path: Path) -> str:
    response = client.post(
        '/api/v1/imports',
        data={
            'batch_name': 'employee-batch',
            'regions': json.dumps(['guangzhou'], ensure_ascii=False),
            'company_names': json.dumps([GUANGZHOU_COMPANY], ensure_ascii=False),
        },
        files=[('files', (sample_path.name, sample_path.read_bytes(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))],
    )
    assert response.status_code == 201
    return response.json()['data']['id']


def import_default_master(client: TestClient):
    response = client.post(
        '/api/v1/employees/import',
        files=[('file', ('employee_master.csv', make_csv_bytes(), 'text/csv'))],
    )
    assert response.status_code == 201
    return response.json()['data']


def test_import_employee_master_csv_and_list_records() -> None:
    client, _settings, _session_factory = build_test_context('csv_import')

    with client:
        payload = import_default_master(client)
        list_all_response = client.get('/api/v1/employees', params={'active_only': 'false'})
        list_active_response = client.get('/api/v1/employees', params={'active_only': 'true'})
        search_response = client.get('/api/v1/employees', params={'query': '张三', 'active_only': 'false'})

    assert payload['file_name'] == 'employee_master.csv'
    assert payload['total_rows'] == 2
    assert payload['imported_count'] == 2
    assert payload['created_count'] == 2
    assert payload['updated_count'] == 0
    assert payload['errors'] == []

    all_items = list_all_response.json()['data']
    assert all_items['total'] == 2
    assert [item['employee_id'] for item in all_items['items']] == ['E1001', 'E1002']

    active_items = list_active_response.json()['data']
    assert active_items['total'] == 1
    assert active_items['items'][0]['employee_id'] == 'E1001'

    search_items = search_response.json()['data']
    assert search_items['total'] == 1
    assert search_items['items'][0]['person_name'] == '张三'


def test_import_employee_master_xlsx_updates_existing_record_and_writes_audit() -> None:
    client, _settings, _session_factory = build_test_context('xlsx_update')

    with client:
        import_default_master(client)
        second_import = client.post(
            '/api/v1/employees/import',
            files=[('file', ('employee_master.xlsx', make_update_workbook_bytes(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))],
        )
        list_response = client.get('/api/v1/employees', params={'query': 'E1001', 'active_only': 'false'})
        employee_id = list_response.json()['data']['items'][0]['id']
        audits_response = client.get(f'/api/v1/employees/{employee_id}/audits')

    assert second_import.status_code == 201
    second_payload = second_import.json()['data']
    assert second_payload['total_rows'] == 1
    assert second_payload['created_count'] == 0
    assert second_payload['updated_count'] == 1

    employee = list_response.json()['data']['items'][0]
    assert employee['employee_id'] == 'E1001'
    assert employee['department'] == '平台运营'
    assert employee['active'] is True

    audits = audits_response.json()['data']['items']
    assert audits[0]['action'] == 'import_update'
    assert audits[-1]['action'] == 'import_create'


def test_import_employee_master_detects_offset_header_and_alias_columns() -> None:
    client, _settings, _session_factory = build_test_context('offset_alias_import')

    with client:
        response = client.post(
            '/api/v1/employees/import',
            files=[('file', ('\u82b1\u540d\u518c.xlsx', make_offset_alias_workbook_bytes(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))],
        )
        list_response = client.get('/api/v1/employees', params={'query': 'E2001', 'active_only': 'false'})

    assert response.status_code == 201
    payload = response.json()['data']
    assert payload['created_count'] == 1
    employee = list_response.json()['data']['items'][0]
    assert employee['employee_id'] == 'E2001'
    assert employee['person_name'] == '\u738b\u4e94'
    assert employee['id_number'] == '440101199303030033'
    assert employee['company_name'] == '\u96f6\u4e00\u88c2\u53d8'
    assert employee['department'] == '\u589e\u957f\u4e2d\u5fc3'
    assert employee['active'] is True


def test_import_employee_master_maps_slash_style_id_number_header() -> None:
    client, _settings, _session_factory = build_test_context('slash_id_header_import')

    with client:
        response = client.post(
            '/api/v1/employees/import',
            files=[('file', ('花名册.xlsx', make_slash_id_header_workbook_bytes(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))],
        )
        list_response = client.get('/api/v1/employees', params={'query': 'E3001', 'active_only': 'false'})

    assert response.status_code == 201
    employee = list_response.json()['data']['items'][0]
    assert employee['employee_id'] == 'E3001'
    assert employee['person_name'] == '赵六'
    assert employee['id_number'] == '440101199404040044'
    assert employee['company_name'] == '广分示例'


def test_update_and_status_endpoints_change_employee_and_write_audit() -> None:
    client, _settings, _session_factory = build_test_context('update_status')

    with client:
        import_default_master(client)
        employee = client.get('/api/v1/employees', params={'query': 'E1001', 'active_only': 'false'}).json()['data']['items'][0]
        update_response = client.patch(
            f"/api/v1/employees/{employee['id']}",
            json={
                'person_name': '张三丰',
                'id_number': employee['id_number'],
                'company_name': '广分示例',
                'department': '平台运营中心',
                'active': True,
            },
        )
        status_response = client.post(
            f"/api/v1/employees/{employee['id']}/status",
            json={'active': False, 'note': '手动停用测试'},
        )
        audits_response = client.get(f"/api/v1/employees/{employee['id']}/audits")

    assert update_response.status_code == 200
    assert update_response.json()['data']['person_name'] == '张三丰'
    assert update_response.json()['data']['department'] == '平台运营中心'

    assert status_response.status_code == 200
    assert status_response.json()['data']['active'] is False

    audits = audits_response.json()['data']['items']
    assert audits[0]['action'] == 'status_change'
    assert audits[0]['note'] == '手动停用测试'
    assert audits[1]['action'] == 'manual_update'


def test_delete_employee_master_removes_record_when_no_match_history() -> None:
    client, _settings, _session_factory = build_test_context('delete_employee')

    with client:
        import_default_master(client)
        employee = client.get('/api/v1/employees', params={'query': 'E1002', 'active_only': 'false'}).json()['data']['items'][0]
        delete_response = client.delete(f"/api/v1/employees/{employee['id']}")
        list_response = client.get('/api/v1/employees', params={'active_only': 'false'})

    assert delete_response.status_code == 204
    remaining = list_response.json()['data']['items']
    assert [item['employee_id'] for item in remaining] == ['E1001']


def test_delete_employee_master_is_blocked_when_match_history_exists() -> None:
    client, _settings, _session_factory = build_test_context('delete_blocked')
    sample_path = find_sample(GUANGZHOU_KEYWORD)
    standardized = standardize_workbook(sample_path, region='guangzhou', company_name=GUANGZHOU_COMPANY)
    first = standardized.records[0]
    employee_csv = (
        '工号,姓名,身份证号,公司名称,部门,是否在职\n'
        f"E9001,{first.values['person_name']},{first.values['id_number']},{GUANGZHOU_COMPANY},运营,是\n"
    ).encode('utf-8-sig')

    with client:
        import_response = client.post('/api/v1/employees/import', files=[('file', ('matching_master.csv', employee_csv, 'text/csv'))])
        employee_id = import_response.json()['data']['items'][0]['id']
        batch_id = create_batch(client, sample_path)
        match_response = client.post(f'/api/v1/imports/{batch_id}/match')
        delete_response = client.delete(f'/api/v1/employees/{employee_id}')
        audits_response = client.get(f'/api/v1/employees/{employee_id}/audits')

    assert match_response.status_code == 200
    assert delete_response.status_code == 409
    assert 'match history' in delete_response.json()['error']['message'].lower()
    audits = audits_response.json()['data']['items']
    assert audits[0]['action'] in {'import_create', 'import_update'}


def test_imported_employee_master_records_are_used_by_match_endpoint() -> None:
    client, _settings, _session_factory = build_test_context('import_then_match')
    sample_path = find_sample(GUANGZHOU_KEYWORD)
    standardized = standardize_workbook(sample_path, region='guangzhou', company_name=GUANGZHOU_COMPANY)
    first = standardized.records[0]
    employee_csv = (
        '工号,姓名,身份证号,公司名称,部门,是否在职\n'
        f"E9001,{first.values['person_name']},{first.values['id_number']},{GUANGZHOU_COMPANY},运营,是\n"
    ).encode('utf-8-sig')

    with client:
        import_response = client.post(
            '/api/v1/employees/import',
            files=[('file', ('matching_master.csv', employee_csv, 'text/csv'))],
        )
        batch_id = create_batch(client, sample_path)
        match_response = client.post(f'/api/v1/imports/{batch_id}/match')
        match_get_response = client.get(f'/api/v1/imports/{batch_id}/match')

    assert import_response.status_code == 201
    assert match_response.status_code == 200
    payload = match_response.json()['data']
    assert payload['employee_master_available'] is True
    assert payload['employee_master_count'] == 1
    assert payload['matched_count'] >= 1

    results = payload['source_files'][0]['results']
    indexed = {item['source_row_number']: item for item in results}
    assert indexed[first.source_row_number]['match_status'] == 'matched'
    assert indexed[first.source_row_number]['employee_id'] == 'E9001'

    persisted = match_get_response.json()['data']
    assert persisted['matched_count'] >= 1


@pytest.mark.parametrize(
    ('filename', 'content_type', 'payload'),
    [
        ('employee_master.txt', 'text/plain', b'not-supported'),
        ('employee_master.csv', 'text/csv', '姓名,公司名称\n张三,广分示例\n'.encode('utf-8-sig')),
    ],
)
def test_employee_import_endpoint_rejects_invalid_files(filename: str, content_type: str, payload: bytes) -> None:
    client, _settings, _session_factory = build_test_context(f"invalid_{filename.replace('.', '_')}")

    with client:
        response = client.post('/api/v1/employees/import', files=[('file', (filename, payload, content_type))])

    assert response.status_code == 400
    assert response.json()['error']['code'] == 'http_error'
